from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.contrib.auth.models import User, Group, Permission
from django.contrib.auth.password_validation import validate_password
from django.db import transaction
from django.contrib.auth.decorators import permission_required
from django.core.mail import send_mail
from django.conf import settings
from .models import Admission, AcademicYear, Class, Section
from .utils import role_required
from django.http import HttpResponseForbidden, JsonResponse
from django.utils.crypto import get_random_string
from django.contrib.auth.hashers import make_password
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login

# =================================================================
# ===================== IMPORTS =====================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import User, Group, Permission
from django.http import HttpResponseForbidden
from django.db.models import Q, Sum, Count
from .forms import RoleForm, AssignRoleForm
from .models import UserRole, RoleActivityLog





from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum
from django.views.decorators.csrf import csrf_protect
from datetime import date, datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import (
    Student, Transaction, Teacher, Staff, StudentPerformance, FeeVoucher,
    AutomationJob, AutomationJobDetail, get_dashboard_stats, StudentBalance,
    NotificationQueue, SalaryVoucher, FeeGenerationSettings, FeeGenerationLog,
    SalaryAutomationSettings, SalaryAutomationJob, SalaryAutomationJobDetail
)
from .forms import StudentRegistrationForm
from .services import (
    FeeGenerationService, SalaryAutomationService, NotificationDispatcherService,
    FixtureAutomationService
)



# ===================================================
# 🔹 STEP 1 — DEFAULT ROLES CREATOR
# ===================================================
def setup_default_roles():
    default_roles = {
        "Teacher": [
            "view_teacher_dashboard",
            "add_result", "change_result", "view_result",
            "add_assignment", "view_assignment",
            "add_quiz", "view_quiz",
        ],
        "Student": [
            "view_student_dashboard",
            "view_result", "view_assignment", "view_quiz",
        ],
        "Parent": [
            "view_parent_dashboard",
            "view_result", "view_student_progress",
        ],
        "Admin": ["__all__"],
    }

    for role_name, perms in default_roles.items():
        group, created = Group.objects.get_or_create(name=role_name)
        if "__all__" in perms:
            group.permissions.set(Permission.objects.all())
        else:
            permissions = Permission.objects.filter(codename__in=perms)
            group.permissions.set(permissions)
        group.save()

    print("✅ Default roles setup complete!")


# ===================================================
# 🔹 STEP 2 — CREATE CUSTOM ROLE
# ===================================================
def _detect_user_profile_type(user):
    profile_checks = (
        ("Teacher", "teacher"),
        ("Student", "student"),
        ("Parent", "parent"),
        ("Employee", "employee_profile"),
    )
    for label, attr in profile_checks:
        try:
            if hasattr(user, attr) and getattr(user, attr):
                return label
        except Exception:
            continue
    if user.is_superuser:
        return "Admin"
    if user.is_staff:
        return "Staff"
    return "User"


def _group_permissions_by_model(permissions):
    grouped_permissions = {}
    for perm in permissions:
        model_name = perm.content_type.model.replace('_', ' ').title()
        grouped_permissions.setdefault(model_name, []).append(perm)
    return grouped_permissions


def _split_full_name(full_name):
    parts = (full_name or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _record_role_activity(action_by, action_type, message, target_user=None, target_role=None, metadata=None):
    RoleActivityLog.objects.create(
        action_by=action_by if getattr(action_by, "is_authenticated", False) else None,
        action_type=action_type,
        target_user=target_user,
        target_role=target_role,
        message=message,
        metadata=metadata or {},
    )


def _role_management_context(active_tab="users", role_form=None, assign_form=None):
    groups = Group.objects.prefetch_related('permissions').annotate(user_count=Count('user')).order_by('name')
    users = User.objects.prefetch_related('groups').order_by('-date_joined', 'username')
    permissions = Permission.objects.select_related('content_type').order_by(
        'content_type__app_label',
        'content_type__model',
        'codename'
    )

    users_with_meta = []
    for user in users:
        user_roles = list(user.groups.all())
        users_with_meta.append({
            "user": user,
            "roles": user_roles,
            "primary_role": user_roles[0].name if user_roles else "No Role",
            "profile_type": _detect_user_profile_type(user),
        })

    role_rows = []
    for group in groups:
        role_name = group.name.lower()
        role_rows.append({
            "group": group,
            "role_type": "System" if role_name in {"admin", "teacher", "student", "parent"} else "Custom",
            "permissions_count": group.permissions.count(),
            "users_count": group.user_count,
        })

    permission_rows = []
    for perm in permissions:
        permission_rows.append({
            "permission": perm,
            "module": perm.content_type.app_label.replace("_", " ").title(),
            "model": perm.content_type.model.replace("_", " ").title(),
        })

    return {
        "active_tab": active_tab,
        "role_form": role_form or RoleForm(),
        "assign_form": assign_form or AssignRoleForm(),
        "users_with_meta": users_with_meta,
        "role_rows": role_rows,
        "role_cards": role_rows,
        "groups": groups,
        "permission_groups": _group_permissions_by_model(permissions),
        "permission_rows": permission_rows,
        "activity_logs": RoleActivityLog.objects.select_related("action_by", "target_user", "target_role")[:50],
        "stats": {
            "total_users": users.count(),
            "active_users": users.filter(is_active=True).count(),
            "total_roles": groups.count(),
            "total_permissions": permissions.count(),
        },
        "recent_users": users[:8],
        "recent_roles": groups[:8],
    }


@login_required
@permission_required('auth.view_group', raise_exception=True)
def user_role_management(request):
    active_tab = request.POST.get('active_tab') or request.GET.get('tab') or "users"
    role_form = RoleForm()
    assign_form = AssignRoleForm()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create_role':
            if not request.user.has_perm('auth.add_group'):
                return HttpResponseForbidden("You do not have permission to create roles.")
            active_tab = "roles"
            role_form = RoleForm(request.POST)
            if role_form.is_valid():
                role_name = role_form.cleaned_data['name'].strip()
                permissions = role_form.cleaned_data['permissions']
                if Group.objects.filter(name__iexact=role_name).exists():
                    messages.error(request, f"Role '{role_name}' already exists.")
                    _record_role_activity(
                        request.user,
                        RoleActivityLog.ACTION_VALIDATION_FAILED,
                        f"Duplicate role creation attempt for '{role_name}'.",
                        metadata={"role_name": role_name},
                    )
                else:
                    group = Group.objects.create(name=role_name)
                    group.permissions.set(permissions)
                    _record_role_activity(
                        request.user,
                        RoleActivityLog.ACTION_ROLE_CREATED,
                        f"Role '{role_name}' created with {permissions.count()} permission(s).",
                        target_role=group,
                        metadata={"permissions": list(permissions.values_list("codename", flat=True))},
                    )
                    messages.success(request, f"Role '{role_name}' created successfully.")
                    return redirect('user_role_management')

        elif action == 'assign_role':
            if not request.user.has_perm('auth.change_user'):
                return HttpResponseForbidden("You do not have permission to assign roles.")
            active_tab = "assign"
            assign_form = AssignRoleForm(request.POST)
            if assign_form.is_valid():
                role = assign_form.cleaned_data['role']
                first_name, last_name = _split_full_name(assign_form.cleaned_data.get('full_name'))
                user = User.objects.create_user(
                    username=assign_form.cleaned_data['username'],
                    email=assign_form.cleaned_data['email'],
                    password=assign_form.cleaned_data['password'],
                    first_name=first_name,
                    last_name=last_name,
                    is_active=assign_form.cleaned_data.get('is_active', True),
                    is_staff=assign_form.cleaned_data.get('is_staff') or role.name.lower() == 'admin',
                )
                user.groups.set([role])
                UserRole.objects.update_or_create(user=user, defaults={'role': role})
                _record_role_activity(
                    request.user,
                    RoleActivityLog.ACTION_USER_CREATED,
                    f"User '{user.username}' created and assigned to '{role.name}'.",
                    target_user=user,
                    target_role=role,
                    metadata={
                        "email": user.email,
                        "is_active": user.is_active,
                        "is_staff": user.is_staff,
                    },
                )
                _record_role_activity(
                    request.user,
                    RoleActivityLog.ACTION_ROLE_ASSIGNED,
                    f"Role '{role.name}' assigned to '{user.username}'.",
                    target_user=user,
                    target_role=role,
                )
                messages.success(request, f"User '{user.username}' created with role '{role.name}'.")
                return redirect('user_role_management')

    return render(
        request,
        'admin_panel/user_role_management.html',
        _role_management_context(active_tab=active_tab, role_form=role_form, assign_form=assign_form)
    )


@login_required
@permission_required('auth.add_group', raise_exception=True)
def create_role(request):
    if request.method == 'POST':
        form = RoleForm(request.POST)
        if form.is_valid():
            role_name = form.cleaned_data['name'].strip()
            permissions = form.cleaned_data['permissions']

            if Group.objects.filter(name__iexact=role_name).exists():
                messages.error(request, f"❌ Role '{role_name}' already exists.")
                _record_role_activity(
                    request.user,
                    RoleActivityLog.ACTION_VALIDATION_FAILED,
                    f"Duplicate role creation attempt for '{role_name}'.",
                    metadata={"role_name": role_name},
                )
                return redirect('create_role')

            group = Group.objects.create(name=role_name)
            group.permissions.set(permissions)
            group.save()
            _record_role_activity(
                request.user,
                RoleActivityLog.ACTION_ROLE_CREATED,
                f"Role '{role_name}' created with {permissions.count()} permission(s).",
                target_role=group,
                metadata={"permissions": list(permissions.values_list("codename", flat=True))},
            )

            messages.success(request, f"✅ Role '{role_name}' created successfully!")
            return redirect('user_role_management')
    else:
        form = RoleForm()

    return render(request, 'admin_panel/user_role_management.html', _role_management_context(active_tab="roles", role_form=form))


# ===================================================
# 🔹 STEP 3 — LIST ALL ROLES
# ===================================================
@login_required
@permission_required('auth.view_group', raise_exception=True)
def list_roles(request):
    groups = Group.objects.prefetch_related('permissions').all()
    return render(request, 'admin_panel/user_role_management.html', _role_management_context(active_tab="roles"))

# ===================================================
# 🔹 HELPER — Safe Role Assignment
# ===================================================
def assign_role_to_user(user, group):
    user.groups.set([group])
    UserRole.objects.update_or_create(
        user=user,
        defaults={'role': group}
    )


from teacher_dashboard.models import Assignment, Quiz, LectureNote
from django.contrib.auth.decorators import login_required

@login_required
def admin_view_assignments(request):
    assignments = Assignment.objects.all().order_by('-created_at')
    return render(request, 'admin_panel/admin_view_assignments.html', {'assignments': assignments})

@login_required
def admin_view_quizzes(request):
    quizzes = Quiz.objects.all().order_by('-due_date')
    return render(request, 'admin_panel/admin_view_quizzes.html', {'quizzes': quizzes})

@login_required
def admin_view_diaries(request):
    return render(request, 'admin_panel/admin_view_diaries.html', {})

@login_required
def admin_view_lecture_notes(request):
    notes = LectureNote.objects.all().order_by('-created_at')
    return render(request, 'admin_panel/admin_view_lecture_notes.html', {'notes': notes})


# ===================================================
# 🔹 STEP 4 — ASSIGN ROLE TO USER
# ===================================================
@login_required
@permission_required('auth.change_user', raise_exception=True)
def assign_role(request):
    users_with_roles = User.objects.prefetch_related('groups').all()

    if request.method == 'POST':
        form = AssignRoleForm(request.POST)
        if form.is_valid():
            role = form.cleaned_data['role']
            first_name, last_name = _split_full_name(form.cleaned_data.get('full_name'))
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password'],
                first_name=first_name,
                last_name=last_name,
                is_active=form.cleaned_data.get('is_active', True),
                is_staff=form.cleaned_data.get('is_staff') or role.name.lower() == 'admin',
            )

            user.groups.clear()
            user.groups.add(role)
            user.save()
            UserRole.objects.update_or_create(user=user, defaults={'role': role})
            _record_role_activity(
                request.user,
                RoleActivityLog.ACTION_USER_CREATED,
                f"User '{user.username}' created and assigned to '{role.name}'.",
                target_user=user,
                target_role=role,
                metadata={
                    "email": user.email,
                    "is_active": user.is_active,
                    "is_staff": user.is_staff,
                },
            )
            _record_role_activity(
                request.user,
                RoleActivityLog.ACTION_ROLE_ASSIGNED,
                f"Role '{role.name}' assigned to '{user.username}'.",
                target_user=user,
                target_role=role,
            )

            messages.success(request, f"User '{user.username}' created with role '{role.name}'.")
            return redirect('user_role_management')
    else:
        form = AssignRoleForm()

    context = _role_management_context(active_tab="assign", assign_form=form)
    context['users_with_roles'] = users_with_roles
    return render(request, 'admin_panel/user_role_management.html', context)


# ===================================================
# 🔹 STEP 5 — DASHBOARDS
# ===================================================
@login_required
@permission_required('teacher_dashboard.view_teacher_dashboard', raise_exception=True)
def teacher_dashboard(request):
    return render(request, 'teacher_dashboard/dashboard.html')


@login_required
@permission_required('student_dashboard.view_student_dashboard', raise_exception=True)
def student_profile(request):
    return render(request, 'student_profile/dashboard.html')


@login_required
@permission_required('parent_dashboard.view_parent_dashboard', raise_exception=True)
def parent_dashboard(request):
    return render(request, 'parent_dashboard/dashboard.html')


def _safe_reverse(view_name, fallback="#"):
    from django.urls import NoReverseMatch, reverse

    try:
        return reverse(view_name)
    except NoReverseMatch:
        return fallback


def _safe_value(callback, default=0):
    try:
        return callback()
    except Exception:
        return default


def _format_automation_activity(value):
    from datetime import date as date_class, datetime as datetime_class
    from django.utils import timezone
    from django.utils.dateformat import format as date_format

    if not value:
        return "No activity yet"
    if isinstance(value, datetime_class):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return date_format(value, "F j, Y, g:i a")
    if isinstance(value, date_class):
        return date_format(value, "F j, Y")
    return str(value)


def build_automation_overview_data():
    from django.utils import timezone
    from .models import (
        AppraisalCycle, AutomationJob, AutomationJobDetail, FeeGenerationSettings,
        FeeVoucher, KpiRule, LeaveApplication, NotificationQueue,
        SalaryAutomationJob, SalaryAutomationJobDetail, SalaryAutomationSettings,
        SalaryVoucher, TeacherAbsence, TeacherAbsenceResult, TeacherAppraisalSubmission,
        TeacherFixture, TeacherFixtureNotificationLog,
    )

    today = timezone.localdate()
    month_start = today.replace(day=1)

    def card(**kwargs):
        defaults = {
            "status": "Ready",
            "status_type": "neutral",
            "primary_value": 0,
            "primary_label": "",
            "secondary_value": 0,
            "secondary_label": "",
            "last_activity": "No activity yet",
            "url": "#",
            "action_label": "Open",
            "icon": "lni lni-cog",
        }
        defaults.update(kwargs)
        return defaults

    total_fixtures = _safe_value(lambda: TeacherFixture.objects.count())
    today_fixtures = _safe_value(lambda: TeacherFixture.objects.filter(fixture_date=today).count())
    fixture_review_count = _safe_value(
        lambda: TeacherFixture.objects.filter(
            fixture_status__in=["uncovered", "cancelled", "declined"]
        ).count()
    ) + _safe_value(lambda: TeacherAbsenceResult.objects.filter(status="unassigned").count())
    last_fixture = _safe_value(
        lambda: TeacherFixture.objects.order_by("-fixture_date", "-created_at").first(),
        None,
    )
    fixture_status = "Needs Review" if fixture_review_count else ("Active" if total_fixtures else "Ready")
    fixture_status_type = "warning" if fixture_review_count else ("success" if total_fixtures else "neutral")

    approved_leaves = _safe_value(
        lambda: LeaveApplication.objects.filter(status="approved", start_date__gte=month_start).count()
    )
    leave_absences = _safe_value(lambda: TeacherAbsence.objects.filter(source="leave").count())
    leave_review_count = _safe_value(
        lambda: TeacherAbsence.objects.filter(source="leave", status__in=["partial", "failed"]).count()
    )
    latest_leave = _safe_value(
        lambda: LeaveApplication.objects.filter(status="approved").order_by("-action_date", "-applied_at").first(),
        None,
    )
    leave_status = "Needs Review" if leave_review_count else ("Active" if leave_absences else "Ready")
    leave_status_type = "warning" if leave_review_count else ("success" if leave_absences else "neutral")

    fee_settings = _safe_value(lambda: FeeGenerationSettings.objects.first(), None)
    latest_fee_job = _safe_value(lambda: AutomationJob.objects.order_by("-started_at").first(), None)
    fee_failed = _safe_value(lambda: AutomationJobDetail.objects.filter(status="FAILED").count())
    fee_vouchers = _safe_value(lambda: FeeVoucher.objects.count())
    fee_status = "Needs Review" if fee_failed else ("Active" if getattr(fee_settings, "auto_enabled", False) else "Ready")
    fee_status_type = "warning" if fee_failed else ("success" if getattr(fee_settings, "auto_enabled", False) else "neutral")

    salary_settings = _safe_value(lambda: SalaryAutomationSettings.objects.first(), None)
    latest_salary_job = _safe_value(lambda: SalaryAutomationJob.objects.order_by("-started_at").first(), None)
    salary_failed = _safe_value(lambda: SalaryAutomationJobDetail.objects.filter(status="FAILED").count())
    salary_vouchers = _safe_value(lambda: SalaryVoucher.objects.count())
    salary_status = "Needs Review" if salary_failed else ("Active" if getattr(salary_settings, "auto_enabled", False) else "Ready")
    salary_status_type = "warning" if salary_failed else ("success" if getattr(salary_settings, "auto_enabled", False) else "neutral")

    pending_notifications = _safe_value(lambda: NotificationQueue.objects.filter(status="PENDING").count())
    failed_notifications = _safe_value(lambda: NotificationQueue.objects.filter(status="FAILED").count())
    fixture_failed_notifications = _safe_value(
        lambda: TeacherFixtureNotificationLog.objects.filter(status="failed").count()
    )
    sent_notifications = _safe_value(lambda: NotificationQueue.objects.filter(status="SENT").count()) + _safe_value(
        lambda: TeacherFixtureNotificationLog.objects.filter(status="sent").count()
    )
    latest_notification = _safe_value(
        lambda: NotificationQueue.objects.order_by("-created_at").first(),
        None,
    )
    notification_review_count = failed_notifications + fixture_failed_notifications
    notification_status = "Needs Review" if notification_review_count else ("Pending" if pending_notifications else "Active")
    notification_status_type = "warning" if notification_review_count or pending_notifications else "success"

    scheduler_jobs = 0
    scheduler_last = None
    scheduler_status = "Ready"
    scheduler_status_type = "neutral"
    try:
        from django_apscheduler.models import DjangoJob, DjangoJobExecution

        scheduler_jobs = DjangoJob.objects.count()
        scheduler_last = DjangoJobExecution.objects.order_by("-run_time").first()
        scheduler_status = "Active" if scheduler_jobs else "Ready"
        scheduler_status_type = "success" if scheduler_jobs else "neutral"
    except Exception:
        scheduler_jobs = 0

    exam_schedules = 0
    seating_count = 0
    latest_seating = None
    try:
        from exam_system.models import ExamSchedule, ExamSeatingPlan

        exam_schedules = ExamSchedule.objects.count()
        seating_count = ExamSeatingPlan.objects.count()
        latest_seating = ExamSeatingPlan.objects.select_related("schedule").order_by("-schedule__exam_date", "-id").first()
    except Exception:
        pass
    exam_status = "Active" if seating_count else ("Pending" if exam_schedules else "Ready")
    exam_status_type = "success" if seating_count else ("neutral" if exam_schedules else "neutral")

    ai_documents = 0
    latest_ai_document = None
    try:
        from ai_tutor.models import AIKnowledgeDocument

        ai_documents = AIKnowledgeDocument.objects.filter(approval_status="approved").count()
        latest_ai_document = AIKnowledgeDocument.objects.filter(approval_status="approved").order_by("-updated_at", "-created_at").first()
    except Exception:
        pass
    ai_status = "Ready" if ai_documents else "Pending"
    ai_status_type = "success" if ai_documents else "neutral"

    open_cycle = _safe_value(lambda: AppraisalCycle.objects.filter(is_open=True).first(), None)
    auto_kpi_count = len(getattr(KpiRule, "AUTO_KPI_KEYS", []))
    auto_metric_submissions = _safe_value(lambda: TeacherAppraisalSubmission.objects.exclude(auto_metrics={}).count())
    latest_appraisal = _safe_value(
        lambda: TeacherAppraisalSubmission.objects.order_by("-submitted_at", "-id").first(),
        None,
    )
    appraisal_status = "Active" if open_cycle else "Ready"
    appraisal_status_type = "success" if open_cycle else "neutral"

    return [
        card(
            key="fixture",
            title="Timetable / Fixture Automation",
            status=fixture_status,
            status_type=fixture_status_type,
            primary_value=total_fixtures,
            primary_label="Total Fixtures",
            secondary_value=today_fixtures,
            secondary_label="Today",
            last_activity=_format_automation_activity(getattr(last_fixture, "created_at", None) or getattr(last_fixture, "fixture_date", None)),
            url=_safe_reverse("timetable_automation"),
            action_label="Open Automation",
            icon="lni lni-calendar",
        ),
        card(
            key="leave_fixture",
            title="HR Leave To Fixture Automation",
            status=leave_status,
            status_type=leave_status_type,
            primary_value=approved_leaves,
            primary_label="Approved This Month",
            secondary_value=leave_absences,
            secondary_label="Leave Absences",
            last_activity=_format_automation_activity(getattr(latest_leave, "action_date", None) or getattr(latest_leave, "applied_at", None)),
            url=_safe_reverse("leave_list"),
            action_label="Open Leaves",
            icon="lni lni-user",
        ),
        card(
            key="fee",
            title="Fee Automation",
            status=fee_status,
            status_type=fee_status_type,
            primary_value=fee_vouchers,
            primary_label="Fee Vouchers",
            secondary_value=fee_failed,
            secondary_label="Failed Records",
            last_activity=_format_automation_activity(getattr(latest_fee_job, "started_at", None)),
            url=_safe_reverse("fee-automation"),
            action_label="Open Fees",
            icon="lni lni-wallet",
        ),
        card(
            key="salary",
            title="Salary Automation",
            status=salary_status,
            status_type=salary_status_type,
            primary_value=salary_vouchers,
            primary_label="Salary Vouchers",
            secondary_value=salary_failed,
            secondary_label="Failed Records",
            last_activity=_format_automation_activity(getattr(latest_salary_job, "started_at", None)),
            url=_safe_reverse("salary-automation"),
            action_label="Open Salary",
            icon="lni lni-briefcase",
        ),
        card(
            key="notifications",
            title="Notification Automation",
            status=notification_status,
            status_type=notification_status_type,
            primary_value=pending_notifications,
            primary_label="Pending",
            secondary_value=sent_notifications,
            secondary_label="Sent",
            last_activity=_format_automation_activity(getattr(latest_notification, "created_at", None)),
            url=_safe_reverse("notification-queue"),
            action_label="Open Queue",
            icon="lni lni-alarm",
        ),
        card(
            key="scheduler",
            title="Scheduler Automation",
            status=scheduler_status,
            status_type=scheduler_status_type,
            primary_value=scheduler_jobs,
            primary_label="Scheduled Jobs",
            secondary_value="Ready",
            secondary_label="Configured",
            last_activity=_format_automation_activity(getattr(scheduler_last, "run_time", None)),
            url=_safe_reverse("automation-settings"),
            action_label="Open Settings",
            icon="lni lni-timer",
        ),
        card(
            key="exam_seating",
            title="Exam Seating Auto Generation",
            status=exam_status,
            status_type=exam_status_type,
            primary_value=exam_schedules,
            primary_label="Exam Schedules",
            secondary_value=seating_count,
            secondary_label="Generated Seats",
            last_activity=_format_automation_activity(getattr(getattr(latest_seating, "schedule", None), "exam_date", None)),
            url=_safe_reverse("exam_plan_list"),
            action_label="Open Exams",
            icon="lni lni-grid-alt",
        ),
        card(
            key="ai_knowledge",
            title="AI Tutor Knowledge Indexing",
            status=ai_status,
            status_type=ai_status_type,
            primary_value=ai_documents,
            primary_label="Approved Documents",
            secondary_value="Index",
            secondary_label="Management Command",
            last_activity=_format_automation_activity(getattr(latest_ai_document, "updated_at", None)),
            url="#",
            action_label="Index Command",
            icon="lni lni-bulb",
        ),
        card(
            key="kpi_appraisal",
            title="KPI / Appraisal Auto Metrics",
            status=appraisal_status,
            status_type=appraisal_status_type,
            primary_value=auto_kpi_count,
            primary_label="Auto KPI Rules",
            secondary_value=auto_metric_submissions,
            secondary_label="Metric Submissions",
            last_activity=_format_automation_activity(getattr(latest_appraisal, "submitted_at", None)),
            url=_safe_reverse("admin_appraisal_list"),
            action_label="Open Appraisal",
            icon="lni lni-stats-up",
        ),
    ]


def build_admin_dashboard_stats():
    from django.contrib.auth.models import Group
    from .models import Admission, NotificationQueue

    admin_group = Group.objects.filter(name__iexact="Admin").first()
    admin_users = User.objects.filter(is_superuser=True)
    if admin_group:
        admin_users = admin_users | User.objects.filter(groups=admin_group)

    try:
        from student_profile.models import Student as PortalStudent

        students_count = PortalStudent.objects.count()
    except Exception:
        students_count = _safe_value(lambda: Admission.objects.count())

    try:
        from teacher_dashboard.models import Teacher as PortalTeacher

        teachers_count = PortalTeacher.objects.count()
    except Exception:
        from .models import Teacher

        teachers_count = _safe_value(lambda: Teacher.objects.count())

    return {
        "students_count": students_count,
        "teachers_count": teachers_count,
        "admins_count": _safe_value(lambda: admin_users.distinct().count()),
        "notifications_pending": _safe_value(lambda: NotificationQueue.objects.filter(status="PENDING").count()),
        "pending_admissions": _safe_value(lambda: Admission.objects.filter(admission_status="pending").count()),
    }


def _dashboard_card(key, title, primary_value, primary_label, secondary_value, secondary_label, url="#",
                    status="Ready", status_type="neutral", action_label="Open", icon="lni lni-cog",
                    extra_value=None, extra_label=None):
    return {
        "key": key,
        "title": title,
        "primary_value": primary_value,
        "primary_label": primary_label,
        "secondary_value": secondary_value,
        "secondary_label": secondary_label,
        "extra_value": extra_value,
        "extra_label": extra_label,
        "url": url,
        "status": status,
        "status_type": status_type,
        "action_label": action_label,
        "icon": icon,
    }


def _dashboard_status(review_count=0, active_count=0):
    if review_count:
        return "Needs Review", "warning"
    if active_count:
        return "Active", "success"
    return "Ready", "neutral"


def build_attention_required_cards():
    from django.utils import timezone
    from .models import (
        AutomationJobDetail, LeaveApplication, NotificationQueue,
        SalaryAutomationJobDetail, TeacherAbsence, TeacherAbsenceResult,
        TeacherFixtureNotificationLog,
    )

    today = timezone.localdate()
    pending_leaves = _safe_value(lambda: LeaveApplication.objects.filter(status="pending").count())
    uncovered_fixtures = _safe_value(lambda: TeacherAbsenceResult.objects.filter(status="unassigned").count())
    absent_today = _safe_value(lambda: TeacherAbsence.objects.filter(absence_date=today).count())
    failed_notifications = _safe_value(lambda: NotificationQueue.objects.filter(status="FAILED").count()) + _safe_value(
        lambda: TeacherFixtureNotificationLog.objects.filter(status="failed").count()
    )
    failed_fee_jobs = _safe_value(lambda: AutomationJobDetail.objects.filter(status="FAILED").count())
    failed_salary_jobs = _safe_value(lambda: SalaryAutomationJobDetail.objects.filter(status="FAILED").count())
    pending_papers = 0
    try:
        from exam_system.models import GeneratedPaper

        pending_papers = GeneratedPaper.objects.filter(status="REVIEW").count()
    except Exception:
        pending_papers = 0

    return [
        _dashboard_card("pending_leaves", "Pending Leaves", pending_leaves, "Applications", absent_today, "Absent Today",
                        _safe_reverse("leave_list"), *(_dashboard_status(pending_leaves, absent_today)),
                        action_label="Review Leaves", icon="lni lni-user"),
        _dashboard_card("uncovered_fixtures", "Uncovered Fixtures", uncovered_fixtures, "Need Substitute", absent_today,
                        "Teacher Absences", _safe_reverse("timetable_automation"),
                        *(_dashboard_status(uncovered_fixtures, absent_today)), action_label="Open Timetable",
                        icon="lni lni-calendar"),
        _dashboard_card("failed_notifications", "Failed Notifications", failed_notifications, "Failed", 0, "Retry Queue",
                        _safe_reverse("notification-queue"), *(_dashboard_status(failed_notifications, 1)),
                        action_label="Open Queue", icon="lni lni-alarm"),
        _dashboard_card("pending_papers", "Paper Approvals", pending_papers, "In Review", 0, "Pending Action",
                        _safe_reverse("exam_plan_list"), *(_dashboard_status(pending_papers, 1)),
                        action_label="Open Exams", icon="lni lni-write"),
        _dashboard_card("failed_jobs", "Failed Automation Jobs", failed_fee_jobs + failed_salary_jobs, "Fee/Salary",
                        failed_fee_jobs, "Fee Failures", _safe_reverse("automation-logs"),
                        *(_dashboard_status(failed_fee_jobs + failed_salary_jobs, 1)), action_label="Open Logs",
                        icon="lni lni-warning"),
    ]


def build_module_summary_cards():
    from django.utils import timezone
    from .models import (
        AppraisalCycle,
        Admission, AssignedPeriod, AutomationJob, AutomationJobDetail, Class, ClassGroup,
        ClassTeacher, Employee, FeeVoucher, KpiRule, LeaveApplication, NotificationQueue,
        SalaryAutomationJob, SalaryAutomationJobDetail, SalaryVoucher, Section, Subject,
        TeacherAbsence, TeacherAbsenceResult, TeacherAppraisalSubmission, TeacherFixture,
        TeacherFixtureNotificationLog,
    )

    today = timezone.localdate()
    month_start = today.replace(day=1)

    cards = []

    pending_admissions = _safe_value(lambda: Admission.objects.filter(admission_status="pending").count())
    month_admissions = _safe_value(lambda: Admission.objects.filter(admission_date__gte=month_start).count())
    cards.append(_dashboard_card(
        "admissions", "Admissions", pending_admissions, "Pending", month_admissions, "This Month",
        _safe_reverse("admission_list"), *(_dashboard_status(pending_admissions, month_admissions)),
        action_label="Open Admissions", icon="lni lni-users"
    ))

    total_classes = _safe_value(lambda: Class.objects.count())
    total_subjects = _safe_value(lambda: Subject.objects.count())
    missing_class_teachers = max(total_classes - _safe_value(lambda: ClassTeacher.objects.count()), 0)
    cards.append(_dashboard_card(
        "academics", "Academics", total_classes, "Classes", total_subjects, "Subjects",
        _safe_reverse("class_list"), *(_dashboard_status(missing_class_teachers, total_classes)),
        action_label="Open Academics", icon="lni lni-book",
        extra_value=_safe_value(lambda: Section.objects.count()) + _safe_value(lambda: ClassGroup.objects.count()),
        extra_label="Sections/Groups"
    ))

    try:
        from teacher_dashboard.models import Teacher as PortalTeacher

        total_teachers = PortalTeacher.objects.count()
        active_teachers = PortalTeacher.objects.filter(status="active").count()
        missing_employee = PortalTeacher.objects.filter(employee__isnull=True).count()
    except Exception:
        from .models import Teacher as PortalTeacher

        total_teachers = _safe_value(lambda: PortalTeacher.objects.count())
        active_teachers = total_teachers
        missing_employee = 0
    cards.append(_dashboard_card(
        "teacher_gateway", "Teacher Gateway", total_teachers, "Teachers", active_teachers, "Active",
        _safe_reverse("teacher_list"), *(_dashboard_status(missing_employee, total_teachers)),
        action_label="Open Teachers", icon="lni lni-graduation",
        extra_value=missing_employee, extra_label="Missing Employee Link"
    ))

    try:
        from student_profile.models import Student as PortalStudent

        total_students = PortalStudent.objects.count()
        missing_class_section = PortalStudent.objects.filter(class_fk__isnull=True).count() + PortalStudent.objects.filter(section__isnull=True).count()
    except Exception:
        total_students = _safe_value(lambda: Admission.objects.count())
        missing_class_section = _safe_value(lambda: Admission.objects.filter(class_fk__isnull=True).count()) + _safe_value(lambda: Admission.objects.filter(section__isnull=True).count())
    try:
        from parent_dashboard.models import Parent

        parent_accounts = Parent.objects.count()
    except Exception:
        parent_accounts = 0
    cards.append(_dashboard_card(
        "student_gateway", "Student Gateway", total_students, "Students", missing_class_section, "Missing Class/Section",
        _safe_reverse("admission_list"), *(_dashboard_status(missing_class_section, total_students)),
        action_label="Open Students", icon="lni lni-user",
        extra_value=parent_accounts, extra_label="Parent Accounts"
    ))

    assigned_periods = _safe_value(lambda: AssignedPeriod.objects.count())
    today_fixtures = _safe_value(lambda: TeacherFixture.objects.filter(fixture_date=today).count())
    uncovered = _safe_value(lambda: TeacherAbsenceResult.objects.filter(status="unassigned").count())
    cards.append(_dashboard_card(
        "timetable", "Timetable / Fixture", assigned_periods, "Assigned Periods", today_fixtures, "Today Fixtures",
        _safe_reverse("timetable_automation"), *(_dashboard_status(uncovered, assigned_periods)),
        action_label="Open Timetable", icon="lni lni-calendar",
        extra_value=uncovered, extra_label="Uncovered"
    ))

    total_employees = _safe_value(lambda: Employee.objects.count())
    pending_leaves = _safe_value(lambda: LeaveApplication.objects.filter(status="pending").count())
    approved_this_month = _safe_value(lambda: LeaveApplication.objects.filter(status="approved", start_date__gte=month_start).count())
    cards.append(_dashboard_card(
        "hr_leave", "HR / Leave", total_employees, "Employees", pending_leaves, "Pending Leaves",
        _safe_reverse("leave_list"), *(_dashboard_status(pending_leaves, total_employees)),
        action_label="Open HR", icon="lni lni-briefcase",
        extra_value=approved_this_month, extra_label="Approved This Month"
    ))

    total_vouchers = _safe_value(lambda: FeeVoucher.objects.count())
    unpaid_vouchers = _safe_value(lambda: FeeVoucher.objects.filter(status__in=["UNPAID", "PARTIAL", "OVERDUE"]).count())
    fee_failures = _safe_value(lambda: AutomationJobDetail.objects.filter(status="FAILED").count())
    cards.append(_dashboard_card(
        "fees", "Fees", total_vouchers, "Vouchers", unpaid_vouchers, "Unpaid/Overdue",
        _safe_reverse("fee-automation"), *(_dashboard_status(fee_failures + unpaid_vouchers, total_vouchers)),
        action_label="Open Fees", icon="lni lni-wallet",
        extra_value=fee_failures, extra_label="Failed Records"
    ))

    salary_vouchers = _safe_value(lambda: SalaryVoucher.objects.count())
    unpaid_salaries = _safe_value(lambda: SalaryVoucher.objects.exclude(status__iexact="PAID").count())
    salary_failures = _safe_value(lambda: SalaryAutomationJobDetail.objects.filter(status="FAILED").count())
    cards.append(_dashboard_card(
        "salary", "Salary", salary_vouchers, "Vouchers", unpaid_salaries, "Unpaid",
        _safe_reverse("salary-automation"), *(_dashboard_status(salary_failures + unpaid_salaries, salary_vouchers)),
        action_label="Open Salary", icon="lni lni-coin",
        extra_value=salary_failures, extra_label="Failed Records"
    ))

    pending_notifications = _safe_value(lambda: NotificationQueue.objects.filter(status="PENDING").count())
    sent_notifications = _safe_value(lambda: NotificationQueue.objects.filter(status="SENT").count()) + _safe_value(
        lambda: TeacherFixtureNotificationLog.objects.filter(status="sent").count()
    )
    failed_notifications = _safe_value(lambda: NotificationQueue.objects.filter(status="FAILED").count()) + _safe_value(
        lambda: TeacherFixtureNotificationLog.objects.filter(status="failed").count()
    )
    cards.append(_dashboard_card(
        "notifications", "Notifications", pending_notifications, "Pending", sent_notifications, "Sent",
        _safe_reverse("notification-queue"), *(_dashboard_status(failed_notifications + pending_notifications, sent_notifications)),
        action_label="Open Queue", icon="lni lni-alarm",
        extra_value=failed_notifications, extra_label="Failed"
    ))

    exam_plans = exam_schedules = generated_seats = pending_questions = pending_papers = 0
    try:
        from exam_system.models import ExamPlan, ExamSchedule, ExamSeatingPlan, GeneratedPaper, Question

        exam_plans = ExamPlan.objects.count()
        exam_schedules = ExamSchedule.objects.filter(exam_date__gte=today).count()
        generated_seats = ExamSeatingPlan.objects.count()
        pending_questions = Question.objects.filter(human_approved=False).count()
        pending_papers = GeneratedPaper.objects.filter(status="REVIEW").count()
    except Exception:
        pass
    cards.append(_dashboard_card(
        "exam_system", "Exam System", exam_plans, "Exam Plans", exam_schedules, "Upcoming Exams",
        _safe_reverse("exam_plan_list"), *(_dashboard_status(pending_questions + pending_papers, exam_plans)),
        action_label="Open Exams", icon="lni lni-pencil-alt",
        extra_value=generated_seats, extra_label="Generated Seats"
    ))

    lecture_notes = assignments = quizzes = pending_submissions = 0
    try:
        from teacher_dashboard.models import Assignment, LectureNote, Quiz
        from student_profile.models import AssignmentSubmission

        lecture_notes = LectureNote.objects.count()
        assignments = Assignment.objects.count()
        quizzes = Quiz.objects.count()
        pending_submissions = AssignmentSubmission.objects.filter(marks__isnull=True).count()
    except Exception:
        pass
    cards.append(_dashboard_card(
        "lms_content", "LMS / Teacher Content", lecture_notes, "Lecture Notes", assignments + quizzes, "Assignments/Quizzes",
        _safe_reverse("admin_view_assignments"), *(_dashboard_status(pending_submissions, lecture_notes + assignments + quizzes)),
        action_label="Open Content", icon="lni lni-library",
        extra_value=pending_submissions, extra_label="Pending Marks"
    ))

    ai_sessions = ai_docs = ai_attempts = weak_topics = 0
    try:
        from ai_tutor.models import AIKnowledgeDocument, AIPracticeAttempt, AITutorSession, StudentTopicMastery

        ai_sessions = AITutorSession.objects.count()
        ai_docs = AIKnowledgeDocument.objects.filter(approval_status="approved").count()
        ai_attempts = AIPracticeAttempt.objects.count()
        weak_topics = StudentTopicMastery.objects.filter(mastery_level__in=["weak", "developing"]).count()
    except Exception:
        pass
    cards.append(_dashboard_card(
        "ai_tutor", "AI Tutor", ai_sessions, "AI Sessions", ai_docs, "Knowledge Docs",
        "#", *(_dashboard_status(weak_topics, ai_sessions + ai_docs)),
        action_label="View Status", icon="lni lni-bulb",
        extra_value=ai_attempts, extra_label="Practice Attempts"
    ))

    open_cycles = _safe_value(lambda: AppraisalCycle.objects.filter(is_open=True).count())
    submitted_appraisals = _safe_value(lambda: TeacherAppraisalSubmission.objects.filter(status="submitted").count())
    auto_kpis = len(getattr(KpiRule, "AUTO_KPI_KEYS", []))
    cards.append(_dashboard_card(
        "appraisal", "Appraisal / KPI", open_cycles, "Open Cycles", submitted_appraisals, "Submitted",
        _safe_reverse("admin_appraisal_list"), *(_dashboard_status(0, open_cycles + submitted_appraisals)),
        action_label="Open Appraisal", icon="lni lni-stats-up",
        extra_value=auto_kpis, extra_label="Auto KPI Rules"
    ))

    latest_fee_job = _safe_value(lambda: AutomationJob.objects.order_by("-started_at").first(), None)
    latest_salary_job = _safe_value(lambda: SalaryAutomationJob.objects.order_by("-started_at").first(), None)
    failed_jobs = fee_failures + salary_failures
    cards.append(_dashboard_card(
        "system_health", "System Health", failed_jobs, "Failed Jobs", getattr(latest_fee_job, "status", "None"), "Latest Fee Job",
        _safe_reverse("automation-logs"), *(_dashboard_status(failed_jobs, 1)),
        action_label="Open Logs", icon="lni lni-cog",
        extra_value=getattr(latest_salary_job, "status", "None"), extra_label="Latest Salary Job"
    ))

    return cards


def build_recent_activity_items():
    from .models import Admission, LeaveApplication, TeacherFixture

    items = []
    latest_admission = _safe_value(lambda: Admission.objects.order_by("-admission_date", "-id").first(), None)
    if latest_admission:
        items.append({
            "title": "Latest admission",
            "description": latest_admission.name,
            "meta": _format_automation_activity(latest_admission.admission_date),
            "url": _safe_reverse("admission_list"),
        })

    latest_leave = _safe_value(lambda: LeaveApplication.objects.order_by("-applied_at").first(), None)
    if latest_leave:
        items.append({
            "title": "Latest leave request",
            "description": str(latest_leave.employee),
            "meta": latest_leave.status.title(),
            "url": _safe_reverse("leave_list"),
        })

    latest_fixture = _safe_value(lambda: TeacherFixture.objects.order_by("-created_at").first(), None)
    if latest_fixture:
        items.append({
            "title": "Latest fixture",
            "description": str(latest_fixture),
            "meta": latest_fixture.fixture_status.title(),
            "url": _safe_reverse("timetable_automation"),
        })

    return items


def build_quick_action_cards():
    return [
        {"title": "Register Student", "url": _safe_reverse("registration"), "icon": "lni lni-user"},
        {"title": "Add Teacher", "url": _safe_reverse("teacher_add"), "icon": "lni lni-graduation"},
        {"title": "Timetable Automation", "url": _safe_reverse("timetable_automation"), "icon": "lni lni-calendar"},
        {"title": "Fee Automation", "url": _safe_reverse("fee-automation"), "icon": "lni lni-wallet"},
        {"title": "Exam Plans", "url": _safe_reverse("exam_plan_list"), "icon": "lni lni-pencil-alt"},
        {"title": "KPI Builder", "url": _safe_reverse("admin_kpi_builder"), "icon": "lni lni-stats-up"},
    ]


def build_dashboard_summary_data():
    return {
        "stats": build_admin_dashboard_stats(),
        "attention_cards": build_attention_required_cards(),
        "module_cards": build_module_summary_cards(),
        "recent_activity": build_recent_activity_items(),
        "quick_actions": build_quick_action_cards(),
    }


def _reference_date_range(period="last_7_days", start_date=None, end_date=None):
    from datetime import datetime, timedelta
    from django.utils import timezone

    today = timezone.localdate()
    labels = {
        "today": "Today",
        "week": "This Week",
        "this_month": "This Month",
        "last_7_days": "Last 7 Days",
        "custom": "Custom Range",
    }
    if period == "today":
        start = end = today
    elif period == "week":
        start, end = today - timedelta(days=today.weekday()), today
    elif period == "this_month":
        start, end = today.replace(day=1), today
    elif period == "custom":
        try:
            start = datetime.strptime(str(start_date), "%Y-%m-%d").date()
            end = datetime.strptime(str(end_date), "%Y-%m-%d").date()
        except (TypeError, ValueError):
            start, end = today - timedelta(days=6), today
            period = "last_7_days"
        if start > end:
            start, end = end, start
        if (end - start).days > 366:
            start = end - timedelta(days=366)
    else:
        start, end = today - timedelta(days=6), today
        period = "last_7_days"

    duration = (end - start).days + 1
    previous_end = start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=duration - 1)
    return start, end, previous_start, previous_end, labels.get(period, "Last 7 Days"), period


def _reference_delta(current, previous):
    current = float(current or 0)
    previous = float(previous or 0)
    if previous == 0:
        return 100.0 if current > 0 else 0.0
    return round(((current - previous) / previous) * 100, 1)


def _reference_money(value):
    value = float(value or 0)
    if abs(value) >= 1000000:
        return f"PKR {value / 1000000:.1f}M"
    if abs(value) >= 1000:
        return f"PKR {value / 1000:.1f}K"
    return f"PKR {value:,.0f}"


def _reference_buckets(start, end):
    from datetime import timedelta

    if (end - start).days <= 31:
        buckets = []
        current = start
        while current <= end:
            buckets.append({"key": current.isoformat(), "label": current.strftime("%a" if (end - start).days <= 7 else "%b %d"), "start": current, "end": current})
            current += timedelta(days=1)
        return buckets, "day"

    buckets = []
    current = start.replace(day=1)
    while current <= end:
        next_month = current.replace(year=current.year + 1, month=1, day=1) if current.month == 12 else current.replace(month=current.month + 1, day=1)
        bucket_start = max(current, start)
        bucket_end = min(next_month - timedelta(days=1), end)
        buckets.append({"key": current.strftime("%Y-%m"), "label": current.strftime("%b %Y"), "start": bucket_start, "end": bucket_end})
        current = next_month
    return buckets, "month"


def _reference_bucket_key(value, mode):
    if hasattr(value, "date") and not hasattr(value, "day"):
        value = value.date()
    return value.strftime("%Y-%m") if mode == "month" else value.isoformat()


def build_reference_dashboard_data(period="last_7_days", start_date=None, end_date=None):
    from collections import defaultdict
    from datetime import timedelta
    from django.db.models import Count, Q, Sum
    from django.db.models.functions import TruncDate
    from django.utils import timezone
    from student_profile.models import Student as PortalStudent
    from teacher_dashboard.models import Assignment, Attendance, Teacher as PortalTeacher
    from exam_system.models import ExamSchedule, GeneratedPaper
    from edupilot_core.models import (
        AutomationJobDetail as CoreAutomationJobDetail,
        FeeVoucher as CoreFeeVoucher,
        NotificationQueue as CoreNotificationQueue,
        StudentBalance as CoreStudentBalance,
        StudentLedger as CoreStudentLedger,
        Transaction as CoreTransaction,
    )
    from .models import (
        Admission, AssignedPeriod, Employee, InventoryItem, LeaveApplication,
        PurchaseRequest, RouteVehicleAssignment, TeacherAbsence, TransportTrip,
        Vehicle, VehicleMaintenance, Vendor,
    )

    start, end, previous_start, previous_end, range_label, normalized_period = _reference_date_range(period, start_date, end_date)
    today = timezone.localdate()
    buckets, bucket_mode = _reference_buckets(start, end)
    bucket_keys = [bucket["key"] for bucket in buckets]
    chart_labels = [bucket["label"] for bucket in buckets]

    def sum_value(queryset, field):
        return queryset.aggregate(total=Sum(field))["total"] or 0

    def values_for_buckets(rows, date_key, value_key):
        values = defaultdict(float)
        for row in rows:
            value_date = row[date_key]
            if value_date is not None:
                values[_reference_bucket_key(value_date, bucket_mode)] += float(row[value_key] or 0)
        return [round(values[key], 2) for key in bucket_keys]

    total_students = PortalStudent.objects.count()
    active_teachers = PortalTeacher.objects.filter(status="active").count()
    teacher_hires = PortalTeacher.objects.filter(joining_date__range=(start, end)).count()
    previous_teacher_hires = PortalTeacher.objects.filter(joining_date__range=(previous_start, previous_end)).count()

    today_attendance = Attendance.objects.filter(date=today)
    marked_today = today_attendance.count()
    present_today = today_attendance.filter(status="present").count()
    absent_today = today_attendance.filter(status="absent").count()
    leave_today = today_attendance.filter(status="leave").count()
    attendance_rate = round((present_today / marked_today) * 100, 1) if marked_today else 0
    previous_day = today - timedelta(days=1)
    previous_marked = Attendance.objects.filter(date=previous_day).count()
    previous_present = Attendance.objects.filter(date=previous_day, status="present").count()
    previous_attendance_rate = round((previous_present / previous_marked) * 100, 1) if previous_marked else 0

    billed = sum_value(CoreFeeVoucher.objects.filter(issue_date__range=(start, end)), "net_amount")
    previous_billed = sum_value(CoreFeeVoucher.objects.filter(issue_date__range=(previous_start, previous_end)), "net_amount")
    received = sum_value(CoreStudentLedger.objects.filter(date__range=(start, end)), "credit")
    previous_received = sum_value(CoreStudentLedger.objects.filter(date__range=(previous_start, previous_end)), "credit")
    fee_rate = round((float(received) / float(billed)) * 100, 1) if billed else 0
    previous_fee_rate = round((float(previous_received) / float(previous_billed)) * 100, 1) if previous_billed else 0

    pending_admissions = Admission.objects.filter(admission_status="pending").count()
    pending_leaves = LeaveApplication.objects.filter(status="pending").count()
    pending_papers = GeneratedPaper.objects.filter(status="REVIEW").count()
    pending_purchases = PurchaseRequest.objects.filter(status="pending").count()
    pending_approvals = pending_admissions + pending_leaves + pending_papers + pending_purchases
    active_vehicles = Vehicle.objects.filter(status="active").count()
    vehicle_additions = Vehicle.objects.filter(created_at__date__range=(start, end)).count()
    previous_vehicle_additions = Vehicle.objects.filter(created_at__date__range=(previous_start, previous_end)).count()

    kpis = [
        {"key": "students", "label": "Total Students", "value": total_students, "display": f"{total_students:,}", "delta": None, "trend_label": "current enrolment", "icon": "fa-user-graduate", "tone": "teal", "url": _safe_reverse("admission_list")},
        {"key": "teachers", "label": "Teaching Staff", "value": active_teachers, "display": f"{active_teachers:,}", "delta": _reference_delta(teacher_hires, previous_teacher_hires), "trend_label": "new hires vs prior period", "icon": "fa-chalkboard-teacher", "tone": "blue", "url": _safe_reverse("teacher_list")},
        {"key": "attendance", "label": "Attendance Today", "value": attendance_rate, "display": f"{attendance_rate:.1f}%", "delta": round(attendance_rate - previous_attendance_rate, 1), "trend_label": "points vs yesterday", "icon": "fa-user-check", "tone": "green", "url": _safe_reverse("ai_analytics_dashboard")},
        {"key": "fees", "label": "Fee Collected", "value": fee_rate, "display": f"{fee_rate:.1f}%", "delta": round(fee_rate - previous_fee_rate, 1), "trend_label": "points vs prior period", "icon": "fa-wallet", "tone": "amber", "url": _safe_reverse("fee-automation")},
        {"key": "approvals", "label": "Pending Approvals", "value": pending_approvals, "display": f"{pending_approvals:,}", "delta": None, "trend_label": "requires action", "icon": "fa-user-clock", "tone": "rose", "url": _safe_reverse("leave_list")},
        {"key": "vehicles", "label": "Active Vehicles", "value": active_vehicles, "display": f"{active_vehicles:,}", "delta": _reference_delta(vehicle_additions, previous_vehicle_additions), "trend_label": "new vehicles vs prior period", "icon": "fa-bus", "tone": "purple", "url": _safe_reverse("operation_vehicle_list")},
    ]

    attendance_rows = Attendance.objects.filter(date__range=(start, end)).values("date", "status").annotate(total=Count("id"))
    attendance_map = defaultdict(lambda: {"present": 0, "absent": 0, "leave": 0})
    for row in attendance_rows:
        attendance_map[_reference_bucket_key(row["date"], bucket_mode)][row["status"]] += row["total"]
    attendance_percentages = []
    for key in bucket_keys:
        values = attendance_map[key]
        total = values["present"] + values["absent"] + values["leave"]
        attendance_percentages.append(round((values["present"] / total) * 100, 1) if total else 0)

    weekday = today.strftime("%A")
    classes_today = AssignedPeriod.objects.filter(day__iexact=weekday).values("class_fk_id", "section_id").distinct().count()
    exams_scheduled = ExamSchedule.objects.filter(exam_date__range=(start, end)).count()
    assignments_due = Assignment.objects.filter(due_date__range=(start, end)).count()
    range_attendance_total = sum(sum(values.values()) for values in attendance_map.values())
    range_present_total = sum(values["present"] for values in attendance_map.values())
    range_attendance_rate = round((range_present_total / range_attendance_total) * 100, 1) if range_attendance_total else 0

    active_employees = Employee.objects.filter(is_active=True)
    active_employee_count = active_employees.count()
    unlinked_teachers = PortalTeacher.objects.filter(status="active", employee__isnull=True).count()
    workforce_total = active_employee_count + unlinked_teachers
    leave_employee_ids = set(LeaveApplication.objects.filter(status="approved", start_date__lte=today, end_date__gte=today).values_list("employee_id", flat=True))
    absence_rows = TeacherAbsence.objects.filter(absence_date=today).select_related("teacher__employee")
    absent_employee_ids = {row.teacher.employee_id for row in absence_rows if row.teacher.employee_id}
    absent_unlinked = sum(1 for row in absence_rows if not row.teacher.employee_id)
    unavailable_employee_ids = leave_employee_ids | absent_employee_ids
    workforce_leave = len(leave_employee_ids)
    workforce_absent = max(len(unavailable_employee_ids) - workforce_leave, 0) + absent_unlinked
    workforce_available = max(workforce_total - workforce_leave - workforce_absent, 0)

    outstanding = sum_value(CoreStudentBalance.objects.all(), "outstanding_amount")
    expenses = sum_value(CoreTransaction.objects.filter(type__iexact="EXPENSE", date__range=(start, end)), "amount")
    pending_vouchers = CoreFeeVoucher.objects.exclude(status="PAID").count()
    income_rows = CoreStudentLedger.objects.filter(date__range=(start, end)).values("date").annotate(total=Sum("credit"))
    expense_rows = CoreTransaction.objects.filter(type__iexact="EXPENSE", date__range=(start, end)).values("date").annotate(total=Sum("amount"))
    income_series = values_for_buckets(income_rows, "date", "total")
    expense_series = values_for_buckets(expense_rows, "date", "total")

    total_requests = PurchaseRequest.objects.count()
    active_vendors = Vendor.objects.filter(status="active").count()
    inventory_items = list(InventoryItem.objects.filter(status="active").only("quantity", "reorder_level"))
    low_stock = sum(1 for item in inventory_items if item.quantity <= item.reorder_level)
    procurement_value = sum_value(PurchaseRequest.objects.filter(created_at__date__range=(start, end), status__in=["approved", "ordered", "received"]), "estimated_cost")
    received_requests = PurchaseRequest.objects.filter(created_at__date__range=(start, end), status="received").count()
    overdue_requests = PurchaseRequest.objects.filter(needed_by__lt=today).exclude(status__in=["received", "rejected"]).count()
    procurement_rows = PurchaseRequest.objects.filter(created_at__date__range=(start, end), status__in=["approved", "ordered", "received"]).annotate(day=TruncDate("created_at")).values("day").annotate(total=Sum("estimated_cost"))
    procurement_series = values_for_buckets(procurement_rows, "day", "total")

    active_assignments = RouteVehicleAssignment.objects.filter(is_active=True).count()
    fleet_capacity = sum_value(Vehicle.objects.filter(status="active"), "capacity")
    under_maintenance = Vehicle.objects.filter(status="maintenance").count()
    trips_in_range = TransportTrip.objects.filter(service_date__range=(start, end))
    students_transported = sum_value(trips_in_range, "students_transported")
    on_time_trips = trips_in_range.filter(status__in=["on_time", "completed"]).count()
    delayed_trips = trips_in_range.filter(status="delayed").count()
    total_performance_trips = on_time_trips + delayed_trips
    on_time_rate = round((on_time_trips / total_performance_trips) * 100, 1) if total_performance_trips else 0
    route_rows = list(
        trips_in_range.values("route__route_name")
        .annotate(total=Count("id"), on_time=Count("id", filter=Q(status__in=["on_time", "completed"])), delayed=Count("id", filter=Q(status="delayed")))
        .order_by("route__route_name")[:6]
    )
    route_performance = [{
        "name": row["route__route_name"] or "Route",
        "rate": round((row["on_time"] / row["total"]) * 100, 1) if row["total"] else 0,
        "status": "Delayed" if row["delayed"] else "On Time",
    } for row in route_rows]

    expiring_registrations = Vehicle.objects.filter(registration_expiry__range=(today, today + timedelta(days=30))).count()
    maintenance_due = VehicleMaintenance.objects.filter(status="scheduled", service_date__lte=today + timedelta(days=7)).count()
    unpaid_fees = CoreFeeVoucher.objects.exclude(status="PAID").count()
    failed_automation = CoreAutomationJobDetail.objects.filter(status="FAILED").count()
    failed_notifications = CoreNotificationQueue.objects.filter(status="FAILED").count()

    alerts = []
    def add_alert(key, tone, icon, title, detail, action, url, count):
        if count:
            alerts.append({"key": key, "tone": tone, "icon": icon, "title": title, "detail": detail, "action": action, "url": url, "count": count})

    add_alert("approvals", "purple", "fa-clipboard-check", f"{pending_approvals} approvals require review", "Admissions, leaves, exams and purchases are awaiting action.", "Review Approvals", _safe_reverse("leave_list"), pending_approvals)
    add_alert("stock", "amber", "fa-box-open", f"{low_stock} inventory items are low", "Current quantity has reached the configured reorder level.", "View Low Stock", _safe_reverse("operation_inventory_item_list"), low_stock)
    add_alert("procurement", "green", "fa-shopping-cart", f"{overdue_requests} purchase requests are overdue", "Required date has passed and the request is not received.", "Review Requests", _safe_reverse("operation_purchase_request_list"), overdue_requests)
    add_alert("trips", "red", "fa-clock", f"{delayed_trips} transport trips were delayed", f"Recorded within {range_label.lower()}.", "View Trips", _safe_reverse("operation_transport_trip_list"), delayed_trips)
    add_alert("vehicles", "blue", "fa-id-card", f"{expiring_registrations} registrations expire soon", "Vehicle registration expiry is within 30 days.", "View Vehicles", _safe_reverse("operation_vehicle_list"), expiring_registrations)
    add_alert("maintenance", "blue", "fa-tools", f"{maintenance_due} maintenance records are due", "Scheduled service is due within seven days.", "View Maintenance", _safe_reverse("operation_vehicle_maintenance_list"), maintenance_due)
    add_alert("fees", "amber", "fa-file-invoice-dollar", f"{unpaid_fees} fee vouchers are pending", "Includes unpaid, partial and overdue vouchers.", "View Vouchers", _safe_reverse("voucher-management"), unpaid_fees)
    add_alert("attendance", "red", "fa-user-check", "Attendance is below 85% today", f"Current marked attendance is {attendance_rate:.1f}%.", "View Attendance", _safe_reverse("ai_analytics_dashboard"), 1 if marked_today and attendance_rate < 85 else 0)
    add_alert("automation", "red", "fa-exclamation-triangle", f"{failed_automation + failed_notifications} system records failed", "Automation jobs or notifications need review.", "Open Logs", _safe_reverse("automation-logs"), failed_automation + failed_notifications)
    if not alerts:
        alerts.append({"key": "clear", "tone": "green", "icon": "fa-check-circle", "title": "All systems are clear", "detail": "No actionable operational alerts right now.", "action": "View Dashboard", "url": _safe_reverse("admin_panel_dashboard"), "count": 0})

    return {
        "meta": {
            "generated_at": timezone.localtime().strftime("%d %b %Y, %I:%M %p"),
            "period": normalized_period,
            "range_label": range_label,
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
        },
        "kpis": kpis,
        "academics": {
            "metrics": [
                {"label": "Attendance", "value": f"{range_attendance_rate:.1f}%"},
                {"label": "Classes Today", "value": classes_today},
                {"label": "Exams Scheduled", "value": exams_scheduled},
                {"label": "Assignments Due", "value": assignments_due},
            ],
            "chart": {"labels": chart_labels, "attendance": attendance_percentages},
            "primary_url": _safe_reverse("class_list"), "report_url": _safe_reverse("ai_analytics_reports"),
        },
        "hr": {
            "metrics": [
                {"label": "Total Workforce", "value": workforce_total},
                {"label": "Available Today", "value": workforce_available},
                {"label": "On Leave", "value": workforce_leave},
                {"label": "Pending Requests", "value": pending_leaves},
            ],
            "chart": {"labels": ["Available", "On Leave", "Absent"], "values": [workforce_available, workforce_leave, workforce_absent]},
            "primary_url": _safe_reverse("employee_list"), "report_url": _safe_reverse("ai_analytics_reports"),
        },
        "finance": {
            "metrics": [
                {"label": "Fee Received", "value": _reference_money(received)},
                {"label": "Outstanding", "value": _reference_money(outstanding)},
                {"label": "Expenses", "value": _reference_money(expenses)},
                {"label": "Pending Vouchers", "value": pending_vouchers},
            ],
            "chart": {"labels": chart_labels, "income": income_series, "expense": expense_series},
            "primary_url": _safe_reverse("automation-dashboard"), "report_url": _safe_reverse("ai_analytics_reports"),
        },
        "procurement": {
            "metrics": [
                {"label": "Purchase Requests", "value": total_requests},
                {"label": "Pending Approvals", "value": pending_purchases},
                {"label": "Active Vendors", "value": active_vendors},
                {"label": "Low Stock Items", "value": low_stock},
            ],
            "chart": {"labels": chart_labels, "spend": procurement_series},
            "summary": {"value": _reference_money(procurement_value), "received": received_requests, "overdue": overdue_requests},
            "primary_url": _safe_reverse("operation_procurement_dashboard"), "report_url": _safe_reverse("operation_purchase_request_list"),
        },
        "fleet": {
            "metrics": [
                {"label": "Active Vehicles", "value": active_vehicles},
                {"label": "Routes Operating", "value": active_assignments},
                {"label": "Students Transported", "value": int(students_transported)},
                {"label": "Under Maintenance", "value": under_maintenance},
            ],
            "summary": {"on_time": on_time_trips, "delayed": delayed_trips, "on_time_rate": on_time_rate, "capacity": int(fleet_capacity)},
            "routes": route_performance,
            "primary_url": _safe_reverse("operation_transportation_dashboard"), "report_url": _safe_reverse("operation_transport_trip_list"),
        },
        "alerts": alerts[:6],
        "quick_actions": [
            {"label": "Register Student", "icon": "fa-user-plus", "url": _safe_reverse("registration")},
            {"label": "Add Teacher", "icon": "fa-chalkboard-teacher", "url": _safe_reverse("teacher_add")},
            {"label": "Upload Students", "icon": "fa-file-upload", "url": _safe_reverse("bulk_upload_students")},
            {"label": "Upload Teachers", "icon": "fa-file-upload", "url": _safe_reverse("bulk_upload_teachers")},
            {"label": "Automation", "icon": "fa-cogs", "url": _safe_reverse("automation-dashboard")},
            {"label": "Reports", "icon": "fa-chart-bar", "url": _safe_reverse("ai_analytics_reports")},
        ],
    }


def build_admin_dashboard_context():
    dashboard_summary = build_dashboard_summary_data()
    return {
        "admin_dashboard_stats": dashboard_summary["stats"],
        "attention_cards": dashboard_summary["attention_cards"],
        "module_summary_cards": dashboard_summary["module_cards"],
        "recent_activity_items": dashboard_summary["recent_activity"],
        "quick_action_cards": dashboard_summary["quick_actions"],
        "automation_overview_cards": build_automation_overview_data(),
        "reference_dashboard_data": build_reference_dashboard_data(),
    }


def build_ai_analytics_data(filters=None):
    from django.utils import timezone
    from django.db.models import Count, Sum
    from datetime import timedelta
    from admin_ai.services import PERIOD_LABELS, resolve_date_range
    from student_profile.models import Student as PortalStudent
    from teacher_dashboard.models import Attendance, Assignment, LectureNote, Quiz, Teacher as PortalTeacher
    from edupilot_core.models import AutomationJob, FeeVoucher, NotificationQueue, SalaryAutomationJob
    from .models import (
        Admission, AssignedPeriod, Class, LeaveApplication, Section, Subject,
        TeacherFixture, TeacherFixtureNotificationLog,
    )

    filters = filters or {"period": "this_month"}
    today = timezone.localdate()
    month_start = today.replace(day=1)
    selected_period = filters.get("period") or "this_month"
    class_filter = filters.get("class_id")
    section_filter = filters.get("section_id")
    chart_start, chart_end, period_label = resolve_date_range(
        selected_period,
        filters.get("start_date"),
        filters.get("end_date"),
    )

    def scoped_attendance_qs(qs):
        if class_filter:
            qs = qs.filter(class_fk_id=class_filter)
        if section_filter:
            qs = qs.filter(section_id=section_filter)
        return qs

    def scoped_admission_qs(qs):
        if class_filter:
            qs = qs.filter(class_fk_id=class_filter)
        if section_filter:
            qs = qs.filter(section_id=section_filter)
        return qs

    total_students = _safe_value(lambda: PortalStudent.objects.count())
    total_teachers = _safe_value(lambda: PortalTeacher.objects.count())
    total_classes = _safe_value(lambda: Class.objects.count())
    total_subjects = _safe_value(lambda: Subject.objects.count())

    attendance_today = _safe_value(lambda: scoped_attendance_qs(Attendance.objects.filter(date=today)).count())
    absent_today = _safe_value(lambda: scoped_attendance_qs(Attendance.objects.filter(date=today, status="absent")).count())
    present_today = _safe_value(lambda: scoped_attendance_qs(Attendance.objects.filter(date=today, status="present")).count())
    attendance_rate = round((present_today / attendance_today) * 100, 1) if attendance_today else 0

    admissions_month = _safe_value(lambda: scoped_admission_qs(Admission.objects.filter(admission_date__gte=month_start)).count())
    pending_admissions = _safe_value(lambda: Admission.objects.filter(admission_status="pending").count())
    approved_admissions = _safe_value(lambda: Admission.objects.filter(admission_status="approved").count())

    fee_total = _safe_value(lambda: FeeVoucher.objects.aggregate(total=Sum("net_amount"))["total"], 0) or 0
    fee_paid = _safe_value(lambda: FeeVoucher.objects.filter(status="PAID").aggregate(total=Sum("net_amount"))["total"], 0) or 0
    unpaid_vouchers = _safe_value(lambda: FeeVoucher.objects.exclude(status="PAID").count())
    fee_collection_rate = round((float(fee_paid) / float(fee_total)) * 100, 1) if fee_total else 0

    assigned_periods = _safe_value(lambda: AssignedPeriod.objects.count())
    fixtures_total = _safe_value(lambda: TeacherFixture.objects.count())
    fixtures_today = _safe_value(lambda: TeacherFixture.objects.filter(fixture_date=today).count())
    uncovered_fixtures = _safe_value(lambda: TeacherFixture.objects.filter(fixture_status__in=["uncovered", "cancelled"]).count())

    pending_leaves = _safe_value(lambda: LeaveApplication.objects.filter(status="pending").count())
    approved_leaves_month = _safe_value(lambda: LeaveApplication.objects.filter(status="approved", action_date__date__gte=month_start).count())
    failed_notifications = _safe_value(lambda: NotificationQueue.objects.filter(status="FAILED").count()) + _safe_value(
        lambda: TeacherFixtureNotificationLog.objects.filter(status="failed").count()
    )

    assignments_due = _safe_value(lambda: Assignment.objects.filter(due_date__gte=today).count())
    quizzes_due = _safe_value(lambda: Quiz.objects.filter(due_date__gte=today).count())
    lecture_notes = _safe_value(lambda: LectureNote.objects.count())

    exam_schedules = 0
    seating_plans = 0
    try:
        from exam_system.models import ExamSchedule, ExamSeatingPlan
        exam_schedules = ExamSchedule.objects.count()
        seating_plans = ExamSeatingPlan.objects.count()
    except Exception:
        pass

    ai_sessions = 0
    ai_docs = 0
    ai_attempts = 0
    try:
        from ai_tutor.models import AITutorSession, AIKnowledgeDocument, AIPracticeAttempt
        ai_sessions = AITutorSession.objects.count()
        ai_docs = AIKnowledgeDocument.objects.count()
        ai_attempts = AIPracticeAttempt.objects.count()
    except Exception:
        pass

    failed_fee_jobs = _safe_value(lambda: AutomationJob.objects.filter(status="FAILED").count())
    failed_salary_jobs = _safe_value(lambda: SalaryAutomationJob.objects.filter(status="FAILED").count())

    risk_score = 0
    risk_score += min(absent_today * 6, 30)
    risk_score += min(pending_admissions * 4, 20)
    risk_score += min(unpaid_vouchers * 3, 20)
    risk_score += min(uncovered_fixtures * 8, 20)
    risk_score += min(failed_notifications * 5, 10)
    risk_score = min(risk_score, 100)

    risk_breakdown = [
        {"key": "attendance", "label": "Absent today", "value": absent_today, "points": min(absent_today * 6, 30), "max_points": 30, "icon": "fas fa-user-times", "url": "#attendance-intelligence"},
        {"key": "admissions", "label": "Pending admissions", "value": pending_admissions, "points": min(pending_admissions * 4, 20), "max_points": 20, "icon": "fas fa-user-clock", "url": _safe_reverse("admission_list")},
        {"key": "fees", "label": "Unpaid vouchers", "value": unpaid_vouchers, "points": min(unpaid_vouchers * 3, 20), "max_points": 20, "icon": "fas fa-file-invoice-dollar", "url": _safe_reverse("voucher-management")},
        {"key": "fixtures", "label": "Uncovered fixtures", "value": uncovered_fixtures, "points": min(uncovered_fixtures * 8, 20), "max_points": 20, "icon": "fas fa-calendar-times", "url": _safe_reverse("timetable_automation")},
        {"key": "notifications", "label": "Failed notifications", "value": failed_notifications, "points": min(failed_notifications * 5, 10), "max_points": 10, "icon": "fas fa-bell-slash", "url": _safe_reverse("notification-queue")},
    ]

    def status_for(value, warning=1):
        if value >= warning:
            return "Needs Review"
        return "Active"

    kpis = [
        {"key": "students", "title": "Students", "value": total_students, "label": "Active profiles", "tone": "teal", "icon": "fas fa-user-graduate"},
        {"key": "attendance", "title": "Attendance Today", "value": f"{attendance_rate}%", "label": f"{absent_today} absent records", "tone": "green" if absent_today == 0 else "warning", "icon": "fas fa-calendar-check"},
        {"key": "fees", "title": "Fee Collection", "value": f"{fee_collection_rate}%", "label": f"{unpaid_vouchers} unpaid vouchers", "tone": "blue" if unpaid_vouchers == 0 else "warning", "icon": "fas fa-wallet"},
        {"key": "risk", "title": "AI Risk Index", "value": risk_score, "label": "Operational risk score", "tone": "green" if risk_score < 35 else "warning", "icon": "fas fa-brain"},
    ]

    modules = [
        {"title": "Admissions Intelligence", "status": status_for(pending_admissions), "primary_value": admissions_month, "primary_label": "This Month", "secondary_value": pending_admissions, "secondary_label": "Pending", "url": _safe_reverse("admission_list"), "icon": "fas fa-user-plus"},
        {"title": "Timetable / Fixture AI", "status": status_for(uncovered_fixtures), "primary_value": assigned_periods, "primary_label": "Assigned Periods", "secondary_value": fixtures_today, "secondary_label": "Today Fixtures", "url": _safe_reverse("timetable_automation"), "icon": "fas fa-calendar-alt"},
        {"title": "HR & Leave Intelligence", "status": status_for(pending_leaves), "primary_value": pending_leaves, "primary_label": "Pending Leaves", "secondary_value": approved_leaves_month, "secondary_label": "Approved Month", "url": _safe_reverse("leave_list"), "icon": "fas fa-users-cog"},
        {"title": "Academic Content", "status": "Active", "primary_value": assignments_due + quizzes_due, "primary_label": "Due Items", "secondary_value": lecture_notes, "secondary_label": "Lecture Notes", "url": _safe_reverse("admin_view_assignments"), "icon": "fas fa-book-open"},
        {"title": "Exam Intelligence", "status": "Ready", "primary_value": exam_schedules, "primary_label": "Schedules", "secondary_value": seating_plans, "secondary_label": "Seats", "url": _safe_reverse("exam_plan_list"), "icon": "fas fa-file-signature"},
        {"title": "AI Tutor Signals", "status": "Active" if ai_sessions else "Ready", "primary_value": ai_sessions, "primary_label": "Sessions", "secondary_value": ai_docs, "secondary_label": "Knowledge Docs", "url": _safe_reverse("ai_tutor_dashboard", "#"), "icon": "fas fa-robot"},
    ]

    alerts = []
    if absent_today:
        alerts.append({"title": "Attendance attention", "detail": f"{absent_today} absence record(s) found today.", "priority": "High", "tone": "danger", "icon": "fas fa-user-times", "url": "#attendance-intelligence", "action": "Review attendance"})
    if pending_admissions:
        alerts.append({"title": "Admissions pending", "detail": f"{pending_admissions} admission request(s) need review.", "priority": "Medium", "tone": "purple", "icon": "fas fa-user-clock", "url": _safe_reverse("admission_list"), "action": "Review admissions"})
    if uncovered_fixtures:
        alerts.append({"title": "Uncovered fixture risk", "detail": f"{uncovered_fixtures} fixture(s) need substitute review.", "priority": "Critical", "tone": "warning", "icon": "fas fa-calendar-times", "url": _safe_reverse("timetable_automation"), "action": "Open fixtures"})
    if failed_notifications:
        alerts.append({"title": "Notification delivery issue", "detail": f"{failed_notifications} failed delivery log(s).", "priority": "High", "tone": "danger", "icon": "fas fa-bell-slash", "url": _safe_reverse("notification-queue"), "action": "Open delivery logs"})
    if attendance_today and attendance_rate < 85:
        alerts.append({"title": "Low attendance signal", "detail": f"Attendance is {attendance_rate}% today, below the 85% review threshold.", "priority": "Warning", "tone": "warning", "icon": "fas fa-chart-line", "url": "#attendance-intelligence", "action": "View trend"})
    if fee_total and fee_collection_rate < 75:
        alerts.append({"title": "Fee collection below target", "detail": f"Collection is {fee_collection_rate}% against issued vouchers.", "priority": "Info", "tone": "info", "icon": "fas fa-wallet", "url": _safe_reverse("voucher-management"), "action": "View vouchers"})
    if not alerts:
        alerts.append({"title": "All clear", "detail": "No critical operational alerts detected from current data.", "priority": "Low", "tone": "success", "icon": "fas fa-check-circle", "url": "#system-health", "action": "View system health"})

    trends = [
        {"label": "Present", "value": present_today, "max": max(attendance_today, 1), "tone": "green"},
        {"label": "Absent", "value": absent_today, "max": max(attendance_today, 1), "tone": "danger"},
        {"label": "Approved Admissions", "value": approved_admissions, "max": max(approved_admissions + pending_admissions, 1), "tone": "teal"},
        {"label": "Pending Admissions", "value": pending_admissions, "max": max(approved_admissions + pending_admissions, 1), "tone": "warning"},
        {"label": "Paid Fees", "value": float(fee_paid), "max": max(float(fee_total), 1), "tone": "green"},
        {"label": "Unpaid Vouchers", "value": unpaid_vouchers, "max": max(unpaid_vouchers + _safe_value(lambda: FeeVoucher.objects.filter(status='PAID').count()), 1), "tone": "warning"},
    ]

    copilot_examples = [
        "Show attendance today",
        "Which students are at risk?",
        "Compare fee collection with last month",
        "Open pending admissions",
        "Generate executive report",
        "Show teacher workload",
    ]

    search_base = _safe_reverse("admin_search")
    quick_actions = [
        {"title": "Generate Report", "url": _safe_reverse("ai_analytics_reports"), "icon": "fas fa-file-export"},
        {"title": "Compare Data", "url": f"{search_base}?q=compare fee collection with last month", "icon": "fas fa-chart-bar"},
        {"title": "Forecast", "url": f"{search_base}?q=forecast", "icon": "fas fa-chart-line"},
        {"title": "Notify Parents", "url": _safe_reverse("notification-queue"), "icon": "fas fa-bell"},
        {"title": "Open Student Search", "url": f"{search_base}?q=student", "icon": "fas fa-search"},
        {"title": "Automation Center", "url": _safe_reverse("automation-dashboard"), "icon": "fas fa-cogs"},
    ]

    def month_bounds(base_date, offset):
        month = base_date.month + offset
        year = base_date.year + ((month - 1) // 12)
        month = ((month - 1) % 12) + 1
        start = base_date.replace(year=year, month=month, day=1)
        if month == 12:
            end = start.replace(year=year + 1, month=1, day=1)
        else:
            end = start.replace(month=month + 1)
        return start, end

    month_labels = []
    admissions_series = []
    fee_total_series = []
    fee_paid_series = []
    fixture_manual_series = []
    fixture_auto_series = []
    if chart_start and chart_end:
        total_days = (chart_end - chart_start).days
        if total_days <= 45:
            current = chart_start
            while current <= chart_end:
                next_day = current + timedelta(days=1)
                month_labels.append(current.strftime("%b %d"))
                admissions_series.append(_safe_value(lambda s=current, e=next_day: scoped_admission_qs(Admission.objects.filter(admission_date__gte=s, admission_date__lt=e)).count()))
                fee_total_series.append(float(_safe_value(lambda s=current, e=next_day: FeeVoucher.objects.filter(issue_date__gte=s, issue_date__lt=e).aggregate(total=Sum("net_amount"))["total"], 0) or 0))
                fee_paid_series.append(float(_safe_value(lambda s=current, e=next_day: FeeVoucher.objects.filter(issue_date__gte=s, issue_date__lt=e, status="PAID").aggregate(total=Sum("net_amount"))["total"], 0) or 0))
                fixture_manual_series.append(_safe_value(lambda s=current, e=next_day: TeacherFixture.objects.filter(fixture_date__gte=s, fixture_date__lt=e, assignment_mode="manual").count()))
                fixture_auto_series.append(_safe_value(lambda s=current, e=next_day: TeacherFixture.objects.filter(fixture_date__gte=s, fixture_date__lt=e).exclude(assignment_mode="manual").count()))
                current = next_day
        else:
            current = chart_start.replace(day=1)
            while current <= chart_end:
                start = current
                end = start.replace(year=start.year + 1, month=1, day=1) if start.month == 12 else start.replace(month=start.month + 1, day=1)
                month_labels.append(start.strftime("%b %Y"))
                admissions_series.append(_safe_value(lambda s=start, e=end: scoped_admission_qs(Admission.objects.filter(admission_date__gte=s, admission_date__lt=e)).count()))
                fee_total_series.append(float(_safe_value(lambda s=start, e=end: FeeVoucher.objects.filter(issue_date__gte=s, issue_date__lt=e).aggregate(total=Sum("net_amount"))["total"], 0) or 0))
                fee_paid_series.append(float(_safe_value(lambda s=start, e=end: FeeVoucher.objects.filter(issue_date__gte=s, issue_date__lt=e, status="PAID").aggregate(total=Sum("net_amount"))["total"], 0) or 0))
                fixture_manual_series.append(_safe_value(lambda s=start, e=end: TeacherFixture.objects.filter(fixture_date__gte=s, fixture_date__lt=e, assignment_mode="manual").count()))
                fixture_auto_series.append(_safe_value(lambda s=start, e=end: TeacherFixture.objects.filter(fixture_date__gte=s, fixture_date__lt=e).exclude(assignment_mode="manual").count()))
                current = end
    else:
        for offset in range(-5, 1):
            start, end = month_bounds(today, offset)
            month_labels.append(start.strftime("%b"))
            admissions_series.append(_safe_value(lambda s=start, e=end: scoped_admission_qs(Admission.objects.filter(admission_date__gte=s, admission_date__lt=e)).count()))
            fee_total_series.append(float(_safe_value(lambda s=start, e=end: FeeVoucher.objects.filter(issue_date__gte=s, issue_date__lt=e).aggregate(total=Sum("net_amount"))["total"], 0) or 0))
            fee_paid_series.append(float(_safe_value(lambda s=start, e=end: FeeVoucher.objects.filter(issue_date__gte=s, issue_date__lt=e, status="PAID").aggregate(total=Sum("net_amount"))["total"], 0) or 0))
            fixture_manual_series.append(_safe_value(lambda s=start, e=end: TeacherFixture.objects.filter(fixture_date__gte=s, fixture_date__lt=e, assignment_mode="manual").count()))
            fixture_auto_series.append(_safe_value(lambda s=start, e=end: TeacherFixture.objects.filter(fixture_date__gte=s, fixture_date__lt=e).exclude(assignment_mode="manual").count()))

    day_labels = []
    present_series = []
    absent_series = []
    leave_series = []
    if chart_start and chart_end:
        current = chart_start
        while current <= chart_end:
            day_labels.append(current.strftime("%b %d"))
            present_series.append(_safe_value(lambda d=current: scoped_attendance_qs(Attendance.objects.filter(date=d, status="present")).count()))
            absent_series.append(_safe_value(lambda d=current: scoped_attendance_qs(Attendance.objects.filter(date=d, status="absent")).count()))
            leave_series.append(_safe_value(lambda d=current: scoped_attendance_qs(Attendance.objects.filter(date=d, status="leave")).count()))
            current = current + timedelta(days=1)
    else:
        for offset in range(6, -1, -1):
            day = today - timedelta(days=offset)
            day_labels.append(day.strftime("%a"))
            present_series.append(_safe_value(lambda d=day: scoped_attendance_qs(Attendance.objects.filter(date=d, status="present")).count()))
            absent_series.append(_safe_value(lambda d=day: scoped_attendance_qs(Attendance.objects.filter(date=d, status="absent")).count()))
            leave_series.append(_safe_value(lambda d=day: scoped_attendance_qs(Attendance.objects.filter(date=d, status="leave")).count()))

    workload_rows = _safe_value(
        lambda: list(
            AssignedPeriod.objects.values("teacher__name")
            .annotate(total=Count("id"))
            .order_by("-total")[:6]
        ),
        [],
    )
    workload_labels = [row.get("teacher__name") or "Teacher" for row in workload_rows]
    workload_values = [row.get("total", 0) for row in workload_rows]

    charts = {
        "attendance": {
            "labels": day_labels,
            "present": present_series,
            "absent": absent_series,
            "leave": leave_series,
        },
        "fees": {
            "labels": month_labels,
            "total": fee_total_series,
            "paid": fee_paid_series,
        },
        "admissions": {
            "labels": month_labels,
            "values": admissions_series,
        },
        "fixture_mix": {
            "labels": month_labels,
            "manual": fixture_manual_series,
            "auto": fixture_auto_series,
        },
        "attendance_donut": {
            "labels": ["Present", "Absent", "Leave"],
            "values": [present_today, absent_today, _safe_value(lambda: Attendance.objects.filter(date=today, status="leave").count())],
        },
        "teacher_workload": {
            "labels": workload_labels,
            "values": workload_values,
        },
    }

    failed_automation_jobs = failed_fee_jobs + failed_salary_jobs
    health_issue_count = failed_automation_jobs + failed_notifications + uncovered_fixtures
    system_health = {
        "status": "Healthy" if health_issue_count == 0 else "Needs Attention",
        "tone": "success" if health_issue_count == 0 else "warning",
        "issue_count": health_issue_count,
        "items": [
            {"key": "automation", "label": "Failed automation jobs", "value": failed_automation_jobs, "icon": "fas fa-cogs", "url": _safe_reverse("automation-logs")},
            {"key": "notifications", "label": "Failed deliveries", "value": failed_notifications, "icon": "fas fa-bell", "url": _safe_reverse("notification-queue")},
            {"key": "fixtures", "label": "Uncovered fixtures", "value": uncovered_fixtures, "icon": "fas fa-calendar-alt", "url": _safe_reverse("timetable_automation")},
        ],
    }

    snapshot = [
        {"key": "students", "label": "Students", "value": total_students},
        {"key": "teachers", "label": "Teachers", "value": total_teachers},
        {"key": "classes", "label": "Classes", "value": total_classes},
        {"key": "subjects", "label": "Subjects", "value": total_subjects},
        {"key": "admissions", "label": "Admissions", "sublabel": "This Month", "value": admissions_month},
        {"key": "pending_leaves", "label": "Leave", "sublabel": "Pending", "value": pending_leaves},
    ]

    return {
        "generated_at": timezone.now().strftime("%B %d, %Y, %I:%M %p"),
        "meta": {
            "period": selected_period,
            "period_label": period_label,
            "start_date": chart_start.isoformat(),
            "end_date": chart_end.isoformat(),
            "class_id": class_filter or "",
            "section_id": section_filter or "",
        },
        "kpis": kpis,
        "snapshot": snapshot,
        "risk": {"score": risk_score, "tone": "success" if risk_score < 35 else "warning", "breakdown": risk_breakdown},
        "system_health": system_health,
        "modules": modules,
        "alerts": alerts,
        "trends": trends,
        "charts": charts,
        "copilot_examples": copilot_examples,
        "quick_actions": quick_actions,
        "counts": {
            "students": total_students,
            "teachers": total_teachers,
            "classes": total_classes,
            "subjects": total_subjects,
            "admissions": admissions_month,
            "pending_leaves": pending_leaves,
            "failed_jobs": failed_automation_jobs,
        },
    }


@login_required
@permission_required('auth.view_group', raise_exception=True)
def ai_analytics_dashboard(request):
    from admin_ai.services import PERIOD_LABELS
    from admin_ai.models import AdminAIConversation
    from .models import Class, Section
    default_filters = {"period": "this_month"}
    conversation = AdminAIConversation.objects.filter(created_by=request.user).first()
    if not conversation:
        conversation = AdminAIConversation.objects.create(created_by=request.user)
    return render(request, "admin_panel/ai_analytics.html", {
        "ai_analytics": build_ai_analytics_data(default_filters),
        "ai_conversation": conversation,
        "period_options": PERIOD_LABELS.items(),
        "classes": Class.objects.all().order_by("class_name"),
        "sections": Section.objects.all().order_by("section_name"),
    })


@login_required
@permission_required('auth.view_group', raise_exception=True)
def ai_analytics_data(request):
    from admin_ai.services import parse_filter_request
    filters = parse_filter_request(request)
    if filters["period"] == "custom" and (not filters["start_date"] or not filters["end_date"]):
        return JsonResponse({"error": "Choose both a start date and an end date for a custom range."}, status=400)
    if filters["start_date"] and filters["end_date"] and filters["start_date"] > filters["end_date"]:
        return JsonResponse({"error": "Start date cannot be after end date."}, status=400)
    return JsonResponse(build_ai_analytics_data(filters))


@login_required
@permission_required('auth.view_group', raise_exception=True)
def admin_dashboard(request):
    return render(request, "admin_panel/index.html", build_admin_dashboard_context())


# ===================================================
# 🔹 STEP 6 — Access Denied Handler
# ===================================================
def access_denied(request):
    return HttpResponseForbidden("🚫 Access Denied: You don't have permission to view this page.")


# =================================================================
from django.views.decorators.cache import never_cache
from django.contrib.auth import logout
from django.shortcuts import redirect, render
from django.contrib.auth import authenticate, login
from django.contrib import messages


@never_cache
def custom_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, "Invalid username/password OR user not active.")
            return render(request, "registration/login.html")

        if not user.is_active:
            messages.error(request, "Your account is inactive. Contact admin.")
            return render(request, "registration/login.html")

        login(request, user)

        if user.is_superuser or user.groups.filter(name__iexact="Admin").exists():
            return redirect("admin_dashboard")

        if user.groups.filter(name__iexact="Teacher").exists():
            return redirect("teacher_dashboard")

        if user.groups.filter(name__iexact="Student").exists():
            return redirect("student_profile_home")

        if user.groups.filter(name__iexact="Parent").exists():
            return redirect("parent_dashboard")

        return redirect("admin_dashboard")

    return render(request, "registration/login.html")


@never_cache
def custom_logout(request):
    logout(request)
    return redirect('login')


# ---------------- CREATE ADMIN USER ----------------
from .utils import group_required


@group_required('Admin')
def create_admin_user(request):
    if request.method == "POST":
        username = request.POST.get("login_id", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "")
        confirm_password = request.POST.get("confirm_password", "")

        if not username:
            messages.error(request, "Login ID is required.")
            return render(request, 'admin_panel/create_admin_user.html')
        if User.objects.filter(username__iexact=username).exists():
            messages.error(request, "This Login ID is already in use.")
            return render(request, 'admin_panel/create_admin_user.html')
        if password != confirm_password:
            messages.error(request, "Password and confirm password do not match.")
            return render(request, 'admin_panel/create_admin_user.html')
        try:
            validate_password(password)
        except Exception as exc:
            messages.error(request, " ".join(exc.messages) if hasattr(exc, "messages") else str(exc))
            return render(request, 'admin_panel/create_admin_user.html')

        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                is_staff=True
            )
            admin_group, _ = Group.objects.get_or_create(name="Admin")
            user.groups.set([admin_group])

        messages.success(request, "Admin user created successfully.")
        return redirect('admin_dashboard')

    return render(request, 'admin_panel/create_admin_user.html')


@login_required
def admin_panel_dashboard(request):
    return render(request, 'admin_panel/index.html', build_admin_dashboard_context())


@login_required
@permission_required('auth.view_group', raise_exception=True)
def automation_overview_data(request):
    return JsonResponse({"cards": build_automation_overview_data()})


@login_required
@permission_required('auth.view_group', raise_exception=True)
def dashboard_summary_data(request):
    reference_data = build_reference_dashboard_data(
        period=request.GET.get("period", "last_7_days"),
        start_date=request.GET.get("start_date"),
        end_date=request.GET.get("end_date"),
    )
    # Keep legacy keys during the dashboard transition for existing consumers.
    reference_data.update(build_dashboard_summary_data())
    return JsonResponse(reference_data)


def _admin_role_label(user):
    if user.is_superuser:
        return "Super Admin"
    groups = list(user.groups.values_list("name", flat=True))
    return ", ".join(groups) if groups else "Admin"


def _admin_search_result(title, module, url, description="", status=""):
    return {
        "title": str(title),
        "module": module,
        "url": url,
        "description": str(description or ""),
        "status": str(status or ""),
    }


def _build_admin_header_notifications(limit=6):
    from django.utils.dateformat import format as date_format
    from django.utils import timezone
    from .models import (
        Admission, AutomationJobDetail, LeaveApplication, NotificationQueue,
        SalaryAutomationJobDetail, TeacherFixtureNotificationLog,
    )

    items = []
    pending_queue = _safe_value(lambda: NotificationQueue.objects.filter(status="PENDING").count())
    failed_queue = _safe_value(lambda: NotificationQueue.objects.filter(status="FAILED").count())
    failed_fixture_logs = _safe_value(lambda: TeacherFixtureNotificationLog.objects.filter(status="failed").count())
    failed_jobs = _safe_value(lambda: AutomationJobDetail.objects.filter(status="FAILED").count()) + _safe_value(
        lambda: SalaryAutomationJobDetail.objects.filter(status="FAILED").count()
    )
    pending_leaves = _safe_value(lambda: LeaveApplication.objects.filter(status="pending").count())
    pending_admissions = _safe_value(lambda: Admission.objects.filter(admission_status="pending").count())

    def add(title, message, url, status="info", dt=None):
        if dt and timezone.is_aware(dt):
            dt = timezone.localtime(dt)
        items.append({
            "title": title,
            "message": message,
            "time": date_format(dt, "F j, Y, g:i a") if dt else "Now",
            "url": url,
            "status": status,
        })

    latest_queue = _safe_value(lambda: NotificationQueue.objects.order_by("-created_at").first(), None)
    latest_fixture_log = _safe_value(lambda: TeacherFixtureNotificationLog.objects.order_by("-created_at").first(), None)

    if failed_queue or failed_fixture_logs:
        add(
            "Failed notifications",
            f"{failed_queue + failed_fixture_logs} failed notification(s) need review",
            _safe_reverse("notification-queue"),
            "warning",
            getattr(latest_fixture_log, "created_at", None) or getattr(latest_queue, "created_at", None),
        )
    if pending_queue:
        add(
            "Pending notifications",
            f"{pending_queue} pending notification(s) in queue",
            _safe_reverse("notification-queue"),
            "info",
            getattr(latest_queue, "created_at", None),
        )
    if failed_jobs:
        add("Failed automation jobs", f"{failed_jobs} automation job detail(s) failed", _safe_reverse("automation-logs"), "warning")
    if pending_leaves:
        add("Pending leaves", f"{pending_leaves} leave application(s) need approval", _safe_reverse("leave_list"), "info")
    if pending_admissions:
        add("Pending admissions", f"{pending_admissions} admission request(s) need review", _safe_reverse("admission_list"), "info")
    if not items:
        add("All clear", "No pending admin alerts right now", _safe_reverse("admin_panel_dashboard"), "success")

    return {
        "count": pending_queue + failed_queue + failed_fixture_logs + failed_jobs + pending_leaves + pending_admissions,
        "items": items[:limit],
    }


def _build_admin_search_results(query, limit_per_group=5):
    query = (query or "").strip()
    if not query:
        return []

    results = []
    query_lower = query.lower()

    def add_many(module, queryset, title_fn, url, description_fn=None, status_fn=None):
        try:
            for obj in queryset[:limit_per_group]:
                results.append(_admin_search_result(
                    title_fn(obj),
                    module,
                    url(obj) if callable(url) else url,
                    description_fn(obj) if description_fn else "",
                    status_fn(obj) if status_fn else "",
                ))
        except Exception:
            return

    q = Q

    from .models import Admission, Book, Class, Employee, FeeVoucher, LeaveApplication, Section, Subject

    def add_command_results():
        from django.db.models import Sum
        from django.utils import timezone
        from teacher_dashboard.models import Attendance
        from student_profile.models import Student as PortalStudent

        today = timezone.localdate()
        month_start = today.replace(day=1)
        paid_total = _safe_value(
            lambda: FeeVoucher.objects.filter(status="PAID").aggregate(total=Sum("net_amount"))["total"],
            0,
        ) or 0
        fee_total = _safe_value(lambda: FeeVoucher.objects.aggregate(total=Sum("net_amount"))["total"], 0) or 0
        unpaid_count = _safe_value(lambda: FeeVoucher.objects.exclude(status="PAID").count())
        collection_rate = round((float(paid_total) / float(fee_total)) * 100, 1) if fee_total else 0
        attendance_total = _safe_value(lambda: Attendance.objects.filter(date=today).count())
        present_count = _safe_value(lambda: Attendance.objects.filter(date=today, status="present").count())
        absent_count = _safe_value(lambda: Attendance.objects.filter(date=today, status="absent").count())
        leave_count = _safe_value(lambda: Attendance.objects.filter(date=today, status="leave").count())
        pending_admissions = _safe_value(lambda: Admission.objects.filter(admission_status="pending").count())
        students_total = _safe_value(lambda: PortalStudent.objects.count())

        command_matches = False

        if any(word in query_lower for word in ["attendance", "present", "absent", "leave"]):
            command_matches = True
            results.append(_admin_search_result(
                "Attendance Today",
                "AI Analytics",
                _safe_reverse("ai_analytics_dashboard"),
                f"{present_count} present, {absent_count} absent, {leave_count} leave from {attendance_total} marked records.",
                "Live",
            ))

        if any(word in query_lower for word in ["fee", "fees", "collection", "compare", "voucher"]):
            command_matches = True
            results.append(_admin_search_result(
                "Fee Collection Summary",
                "Accounts",
                _safe_reverse("fee-automation"),
                f"{collection_rate}% collected. Paid {paid_total}; total billed {fee_total}; {unpaid_count} unpaid vouchers.",
                "Live",
            ))
            results.append(_admin_search_result(
                "Voucher Management",
                "Accounts",
                _safe_reverse("voucher-management"),
                "Open voucher list for paid, unpaid, partial, and overdue records.",
                "Open",
            ))

        if any(word in query_lower for word in ["report", "executive", "analytics", "summary"]):
            command_matches = True
            results.append(_admin_search_result(
                "Executive AI Analytics Report",
                "AI Analytics",
                _safe_reverse("ai_analytics_reports"),
                f"{students_total} students, {attendance_total} attendance records today, {collection_rate}% fee collection, {pending_admissions} pending admissions.",
                "Ready",
            ))

        if any(word in query_lower for word in ["forecast", "risk", "predict"]):
            command_matches = True
            risk_score = 0
            risk_score += min(absent_count * 6, 30)
            risk_score += min(pending_admissions * 4, 20)
            risk_score += min(unpaid_count * 3, 20)
            results.append(_admin_search_result(
                "Forecast And Risk Signal",
                "AI Analytics",
                _safe_reverse("ai_analytics_dashboard"),
                f"Current operational risk estimate is {min(risk_score, 100)} based on absence, admissions, and fee signals.",
                "Forecast",
            ))

        if any(word in query_lower for word in ["teacher workload", "workload", "period load", "assigned period"]):
            command_matches = True
            results.append(_admin_search_result(
                "Teacher Workload Chart",
                "Timetable",
                _safe_reverse("ai_analytics_dashboard"),
                "Open AI Analytics teacher workload graph based on assigned periods.",
                "Live",
            ))
            results.append(_admin_search_result(
                "Assign Period Teachers",
                "Timetable",
                _safe_reverse("assign_period"),
                "Open period assignment screen for detailed teacher timetable allocation.",
                "Open",
            ))

        if any(word in query_lower for word in ["parent", "notify", "notification", "message"]):
            command_matches = True
            results.append(_admin_search_result(
                "Notification Queue",
                "Accounts",
                _safe_reverse("notification-queue"),
                "Open pending, sent, and failed parent/student notification queue.",
                "Open",
            ))

        if any(word in query_lower for word in ["pending admission", "admission", "admissions"]):
            command_matches = True
            results.append(_admin_search_result(
                "Pending Admissions",
                "Admissions",
                _safe_reverse("admission_list"),
                f"{pending_admissions} admission request(s) need review.",
                "Review",
            ))

        if any(word in query_lower for word in ["student", "students", "at risk"]):
            command_matches = True
            results.append(_admin_search_result(
                "Student Search",
                "Student Gateway",
                _safe_reverse("admin_ai_student_intelligence"),
                f"Search student/admission records. Current portal students: {students_total}.",
                "Open",
            ))

        return command_matches

    command_matched = add_command_results()

    add_many(
        "Admissions",
        Admission.objects.filter(q(name__icontains=query) | q(student_id__icontains=query) | q(email__icontains=query) | q(contact__icontains=query)).order_by("-id"),
        lambda item: item.name,
        _safe_reverse("admission_list"),
        lambda item: f"{item.student_id or 'No ID'} - {getattr(item.class_fk, 'class_name', 'No class')}",
        lambda item: item.admission_status,
    )
    add_many(
        "Teachers",
        __import__("teacher_dashboard.models", fromlist=["Teacher"]).Teacher.objects.filter(q(name__icontains=query) | q(email__icontains=query) | q(phone__icontains=query)).order_by("name"),
        lambda item: item.name,
        lambda item: _safe_reverse("teacher_edit", "#").replace("/0/", f"/{item.pk}/") if _safe_reverse("teacher_edit", "#") != "#" else _safe_reverse("teacher_list"),
        lambda item: item.email,
        lambda item: item.status,
    )
    add_many(
        "Users",
        User.objects.filter(q(username__icontains=query) | q(first_name__icontains=query) | q(last_name__icontains=query) | q(email__icontains=query)).order_by("username"),
        lambda item: item.get_full_name() or item.username,
        _safe_reverse("user_list"),
        lambda item: item.email,
        lambda item: "active" if item.is_active else "inactive",
    )
    add_many("Classes", Class.objects.filter(class_name__icontains=query).order_by("class_name"), lambda item: item.class_name, _safe_reverse("class_list"), lambda item: f"{item.total_students} students")
    add_many("Sections", Section.objects.filter(q(section_name__icontains=query) | q(class_fk__class_name__icontains=query)).select_related("class_fk").order_by("class_fk__class_name"), lambda item: str(item), _safe_reverse("section_list"), lambda item: f"Capacity {item.capacity}")
    add_many("Subjects", Subject.objects.filter(q(name__icontains=query) | q(short_code__icontains=query)).order_by("name"), lambda item: str(item), _safe_reverse("subject_list"), lambda item: getattr(item.class_fk, "class_name", "No class"))
    add_many("Employees", Employee.objects.filter(q(name__icontains=query) | q(email__icontains=query) | q(phone__icontains=query)).order_by("name"), lambda item: item.name, _safe_reverse("employee_list"), lambda item: item.email)
    add_many("Leaves", LeaveApplication.objects.filter(q(employee__name__icontains=query) | q(reason__icontains=query) | q(status__icontains=query)).select_related("employee").order_by("-applied_at"), lambda item: f"{item.employee.name} leave", _safe_reverse("leave_list"), lambda item: f"{item.start_date} to {item.end_date}", lambda item: item.status)
    add_many("Fee Vouchers", FeeVoucher.objects.filter(q(voucher_no__icontains=query) | q(student__full_name__icontains=query) | q(month__icontains=query)).select_related("student").order_by("-id"), lambda item: item.voucher_no, _safe_reverse("voucher-management"), lambda item: f"{item.student.full_name} - {item.month} {item.year}", lambda item: item.status)
    add_many("Books", Book.objects.filter(q(title__icontains=query) | q(subject__name__icontains=query) | q(class_for__class_name__icontains=query)).select_related("subject", "class_for").order_by("-uploaded_at"), lambda item: item.title, lambda item: _safe_reverse("book_detail", "#").replace("/0/", f"/{item.pk}/") if _safe_reverse("book_detail", "#") != "#" else _safe_reverse("book_list"), lambda item: f"{item.class_for} - {item.subject}")

    static_pages = [
        ("Timetable Automation", "Timetable", _safe_reverse("timetable_automation")),
        ("AI Copilot", "AI Analytics", _safe_reverse("admin_ai_copilot")),
        ("Advanced AI Analytics", "AI Analytics", _safe_reverse("ai_analytics_advanced")),
        ("AI Analytics Reports", "AI Analytics", _safe_reverse("ai_analytics_reports")),
        ("Student Insight Profile", "AI Analytics", _safe_reverse("admin_ai_student_intelligence")),
        ("Fixture Management", "Timetable", _safe_reverse("manage_fixture")),
        ("Assign Periods", "Timetable", _safe_reverse("assign_period")),
        ("Fee Automation", "Accounts", _safe_reverse("fee-automation")),
        ("Salary Automation", "Accounts", _safe_reverse("salary-automation")),
        ("Notification Queue", "Accounts", _safe_reverse("notification-queue")),
        ("Question Formats", "Exams", _safe_reverse("format_list")),
        ("Questions", "Exams", _safe_reverse("all_questions")),
        ("Books", "LMS", _safe_reverse("book_list")),
        ("KPI Builder", "Appraisal", _safe_reverse("admin_kpi_builder")),
    ]
    for title, module, url in static_pages:
        if query.lower() in title.lower() or query.lower() in module.lower():
            results.append(_admin_search_result(title, module, url, "Open module"))

    if command_matched and results:
        return results[:30]

    return results[:30]


@login_required
def admin_header_data(request):
    from django.templatetags.static import static

    notifications = _build_admin_header_notifications()
    return JsonResponse({
        "notifications_count": notifications["count"],
        "notifications": notifications["items"],
        "user": {
            "name": request.user.get_full_name() or request.user.username,
            "role": _admin_role_label(request.user),
            "avatar": static("admin_panel/images/dummy-profile.png"),
        },
    })


@login_required
def admin_search_suggestions(request):
    query = request.GET.get("q", "")
    return JsonResponse({"results": _build_admin_search_results(query, limit_per_group=3)[:8]})


@login_required
def admin_search(request):
    query = request.GET.get("q", "")
    return render(request, "admin_panel/search_results.html", {
        "query": query,
        "results": _build_admin_search_results(query, limit_per_group=8),
    })


@login_required
def admin_profile(request):
    return render(request, "admin_panel/admin_profile.html", {
        "profile_user": request.user,
        "role_label": _admin_role_label(request.user),
        "groups": request.user.groups.all(),
    })


# ------------------- user_list -------------------
from collections import OrderedDict
from django.contrib.auth.models import User
from admin_panel.models import UserRole


def user_list(request):
    users = User.objects.all().order_by('email')
    unique = []

    for user in users:
        roles = ", ".join([g.name for g in user.groups.all()]) or "No Role"
        unique.append({
            'username': user.username,
            'email': user.email,
            'role': roles
        })

    return render(request, 'admin_panel/user_list.html', {'users_with_roles': unique})


# ==================== ADMISSION WORK ====================
from django.shortcuts import render, redirect
from .forms import AdmissionForm
from .models import Admission, Class, AcademicYear
from django.core.paginator import Paginator


@permission_required('admin_panel.add_admission', raise_exception=True)
def register_admission(request):
    classes = Class.objects.all()
    active_years = AcademicYear.objects.filter(is_active=True)
    selected_year = None

    if request.method == 'POST':
        selected_year = request.POST.get('academic_year')
        form = AdmissionForm(request.POST)
        form.fields['academic_year'].queryset = active_years
        if form.is_valid():
            admission = form.save(commit=False)
            admission.save()
            messages.success(request, "Admission registered successfully.")
            return redirect('admission_list')
        else:
            print("Form Errors:", form.errors)
    else:
        form = AdmissionForm()
        form.fields['academic_year'].queryset = active_years

    context = {
        'form': form,
        'classes': classes,
        'academic_years': active_years,
        'academic_year': int(selected_year) if selected_year else None,
    }
    return render(request, 'admin_panel/registration.html', context)


@permission_required('admin_panel.view_admission', raise_exception=True)
def admission_list(request):
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if active_year:
        admissions = Admission.objects.filter(academic_year=active_year).order_by('id')
    else:
        admissions = Admission.objects.none()

    return render(request, 'admin_panel/admission_list.html', {'admissions': admissions})


@permission_required('admin_panel.change_admission', raise_exception=True)
def update_admission_status(request, pk):
    if request.method == 'POST':
        admission = get_object_or_404(Admission, pk=pk)
        new_status = request.POST.get('admission_status')
        if new_status in ['pending', 'approved', 'rejected']:
            admission.admission_status = new_status
            admission.save()
    return redirect('admission_list')


def is_section_full(section):
    current_count = Admission.objects.filter(section=section).count()
    return current_count >= section.capacity


# -------------------- reject_reason --------------------
from django.core.mail import send_mail, BadHeaderError


@permission_required('admin_panel.change_admission', raise_exception=True)
def reject_reason(request, admission_id):
    admission = get_object_or_404(Admission, id=admission_id)

    if request.method == 'POST':
        reason = request.POST.get('reason')
        admission.admission_status = 'rejected'
        admission.rejection_reason = reason
        admission.save()

        if admission.email:
            try:
                send_mail(
                    subject="Admission Rejected",
                    message=f"Dear {admission.name},\n\nYour admission has been rejected.\n\nReason: {reason}\n\nRegards,\nSchool Administration",
                    from_email="your@email.com",
                    recipient_list=[admission.email],
                    fail_silently=False,
                )
            except BadHeaderError:
                messages.error(request, "Invalid header found when sending email.")
            except Exception as e:
                messages.error(request, f"Failed to send email: {str(e)}")
        else:
            messages.warning(request, "Student email not found; no email was sent.")

        return redirect('admission_list')

    return render(request, 'admin_panel/reject_reason.html', {'admission': admission})


# -------------------- change_admission_status --------------------

from django.db.models import Max
from parent_dashboard.models import Parent
from student_profile.models import Student


def generate_unique_student_id():
    while True:
        try:
            with transaction.atomic():
                max_id = Student.objects.aggregate(max_num=Max('student_id'))['max_num']
                last_num = int(max_id.replace("STD", "")) if max_id else 0
                new_id = f"STD{last_num + 1:04d}"
                if not Student.objects.filter(student_id=new_id).exists():
                    return new_id
        except Exception:
            continue


@permission_required('admin_panel.change_admission', raise_exception=True)
def approve_admission_credentials(request, admission_id):
    admission = get_object_or_404(Admission, id=admission_id)
    student_login_default = admission.student_id or f"STD{admission.id:04d}"
    parent_login_default = f"PAR{admission.id:04d}" if admission.father_email else ""

    if request.method == "POST":
        student_login_id = request.POST.get("student_login_id", "").strip()
        student_password = request.POST.get("student_password", "")
        student_confirm_password = request.POST.get("student_confirm_password", "")
        parent_login_id = request.POST.get("parent_login_id", "").strip()
        parent_password = request.POST.get("parent_password", "")
        parent_confirm_password = request.POST.get("parent_confirm_password", "")
        errors = []

        if not student_login_id:
            errors.append("Student Login ID is required.")
        if User.objects.filter(username__iexact=student_login_id).exists():
            errors.append("Student Login ID is already in use.")
        if admission.email and Student.objects.filter(email__iexact=admission.email).exists():
            errors.append("A student with this email already exists.")
        if student_password != student_confirm_password:
            errors.append("Student password and confirm password do not match.")
        else:
            try:
                validate_password(student_password)
            except Exception as exc:
                errors.extend(exc.messages if hasattr(exc, "messages") else [str(exc)])

        needs_parent_account = bool(admission.father_email)
        existing_parent_user = User.objects.filter(email__iexact=admission.father_email).first() if needs_parent_account else None
        if needs_parent_account:
            if not parent_login_id:
                errors.append("Parent Login ID is required.")
            existing_login = User.objects.filter(username__iexact=parent_login_id).first()
            if existing_login and existing_login != existing_parent_user:
                errors.append("Parent Login ID is already in use.")
            if parent_password != parent_confirm_password:
                errors.append("Parent password and confirm password do not match.")
            else:
                try:
                    validate_password(parent_password)
                except Exception as exc:
                    errors.extend(exc.messages if hasattr(exc, "messages") else [str(exc)])

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, "admin_panel/approve_admission_credentials.html", {
                "admission": admission,
                "student_login_default": student_login_id or student_login_default,
                "parent_login_default": parent_login_id or parent_login_default,
            })

        student_user = None
        parent_user = None
        student = None

        with transaction.atomic():
            student_user = User.objects.create_user(
                username=student_login_id,
                email=admission.email or "",
                password=student_password,
                first_name=admission.name,
            )
            student_group, _ = Group.objects.get_or_create(name="Student")
            student_user.groups.add(student_group)

            student = Student.objects.create(
                user=student_user,
                academic_year=admission.academic_year,
                student_id=admission.student_id or student_login_id,
                name=admission.name,
                father_name=admission.father_name,
                mother_name=admission.mother_name or "Unknown",
                class_fk=admission.class_fk,
                section=admission.section,
                roll_no=admission.student_id or student_login_id,
                phone=admission.contact,
                gender=(admission.gender or "Male").title(),
                date_of_birth=admission.dob or date.today(),
                email=admission.email or f"{student_login_id}@edupilot.local",
            )

            if needs_parent_account:
                parent_group, _ = Group.objects.get_or_create(name="Parent")
                if existing_parent_user:
                    parent_user = existing_parent_user
                    parent_user.username = parent_login_id
                    parent_user.set_password(parent_password)
                    parent_user.first_name = admission.father_name
                    parent_user.save()
                else:
                    parent_user = User.objects.create_user(
                        username=parent_login_id,
                        email=admission.father_email or "",
                        password=parent_password,
                        first_name=admission.father_name,
                    )
                parent_user.groups.add(parent_group)

                parent, _ = Parent.objects.get_or_create(
                    email=admission.father_email,
                    defaults={
                        "user": parent_user,
                        "full_name": admission.father_name,
                        "phone": admission.father_contact or admission.contact,
                        "address": admission.address,
                        "occupation": admission.father_occupation,
                    }
                )
                parent.user = parent_user
                parent.full_name = admission.father_name or parent.full_name
                parent.phone = admission.father_contact or admission.contact
                parent.address = admission.address
                parent.occupation = admission.father_occupation
                parent.save()
                parent.students.add(student)

            admission.admission_status = "approved"
            admission.save()

        site_url = getattr(settings, "SITE_URL", "http://127.0.0.1:8000").rstrip("/")
        if admission.email:
            try:
                send_mail(
                    "Student Portal Login Details",
                    (
                        f"Hello {admission.name},\n\n"
                        f"Your student portal account has been approved.\n\n"
                        f"Login ID: {student_user.username}\n"
                        f"Password: {student_password}\n"
                        f"Login URL: {site_url}/login/student/\n\n"
                        "Thanks\nEduPilot"
                    ),
                    settings.DEFAULT_FROM_EMAIL,
                    [admission.email],
                    fail_silently=False,
                )
            except Exception as exc:
                messages.warning(request, f"Student account approved, but email failed: {exc}")

        if parent_user and admission.father_email:
            try:
                send_mail(
                    "Parent Portal Login Details",
                    (
                        f"Hello {admission.father_name},\n\n"
                        f"Your parent portal account has been created.\n\n"
                        f"Login ID: {parent_user.username}\n"
                        f"Password: {parent_password}\n"
                        f"Login URL: {site_url}/login/parent/\n\n"
                        "Thanks\nEduPilot"
                    ),
                    settings.DEFAULT_FROM_EMAIL,
                    [admission.father_email],
                    fail_silently=False,
                )
            except Exception as exc:
                messages.warning(request, f"Parent account created, but email failed: {exc}")

        messages.success(request, "Admission approved and portal credentials saved.")
        return redirect("admission_list")

    return render(request, "admin_panel/approve_admission_credentials.html", {
        "admission": admission,
        "student_login_default": student_login_default,
        "parent_login_default": parent_login_default,
    })


from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.models import User, Group
from django.core.mail import send_mail
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.crypto import get_random_string
from django.core.paginator import Paginator
from django.db.models import Q

from .models import Admission
from student_profile.models import Student


@permission_required('admin_panel.change_admission', raise_exception=True)
def change_admission_status(request, admission_id):

    print("\n========== FUNCTION START ==========")

    if request.method != 'POST':
        print(">>> NOT POST - returning")
        return redirect('admission_list')

    admission = get_object_or_404(Admission, id=admission_id)
    status = request.POST.get('admission_status')

    print("STATUS        =", repr(status))
    print("EMAIL         =", admission.email)
    print("FATHER EMAIL  =", admission.father_email)
    print("CURRENT DB STATUS =", admission.admission_status)

    # =========================================================
    # APPROVED
    # =========================================================
    if status == 'approved':
        return redirect('approve_admission_credentials', admission_id=admission.id)

    if status == 'approved':
        print(">>> ENTERED APPROVED BLOCK")

        student_exists = Student.objects.filter(email=admission.email).exists() if admission.email else False
        user_exists    = User.objects.filter(email=admission.email).exists()    if admission.email else False

        print("STUDENT EXISTS =", student_exists)
        print("USER EXISTS    =", user_exists)

        if student_exists:
            print(">>> Student pehle se exist karta hai — sirf status update")
            admission.admission_status = 'approved'
            admission.save()
            messages.warning(
                request,
                f"⚠️ Is email '{admission.email}' ka student pehle se exist karta hai. "
                f"Naya account nahi banaya, status approved kar diya."
            )
            print("========== FUNCTION END ==========\n")
            return redirect('admission_list')

        if user_exists:
            print(">>> User exist karta hai — delete karke fresh banao")
            User.objects.filter(email=admission.email).delete()
            print(">>> Old user deleted")

        username = admission.student_id or f"STD{get_random_string(6)}"
        password = get_random_string(8)

        print("USERNAME =", username)
        print("PASSWORD =", password)

        try:
            with transaction.atomic():

                print(">>> Creating User...")
                student_user = User.objects.create_user(
                    username=username,
                    email=admission.email or "",
                    password=password,
                    first_name=admission.name,
                )
                print(">>> User created:", student_user.username)

                student_group, _ = Group.objects.get_or_create(name='Student')
                student_user.groups.set([student_group])
                print(">>> Group assigned")

                print(">>> Creating Student profile...")
                Student.objects.update_or_create(
                    student_id=username,
                    defaults={
                        "user":          student_user,
                        "academic_year": admission.academic_year,
                        "name":          admission.name,
                        "father_name":   admission.father_name,
                        "mother_name":   admission.mother_name,
                        "class_fk":      admission.class_fk,
                        "roll_no":       "TEMP",
                        "phone":         admission.contact,
                        "gender":        admission.gender,
                        "date_of_birth": admission.dob,
                        "email":         admission.email,
                        "section":       admission.section,
                    }
                )
                print(">>> Student profile created")

            admission.admission_status = 'approved'
            admission.save()
            print(">>> Admission status = approved")
            print(">>> STUDENT CREATED SUCCESSFULLY")

            school_name = getattr(settings, 'SCHOOL_NAME', 'EduPilot School')
            site_url    = getattr(settings, 'SITE_URL', 'http://127.0.0.1:8000')
            login_url   = f"{site_url}/auth/login/"

            # =========================================================
            # STUDENT KO EMAIL BHEJO
            # =========================================================
            if admission.email:
                student_email_body = f"""Dear {admission.name},

Congratulations! Your admission has been approved.

=============================
  STUDENT LOGIN DETAILS
=============================
Username : {username}
Password : {password}
Login URL: {login_url}
=============================

Regards,
{school_name}
"""
                print(">>> Sending student email to:", admission.email)

                try:
                    result = send_mail(
                        subject=f"Admission Approved - {school_name}",
                        message=student_email_body,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[admission.email],
                        fail_silently=False,
                    )
                    print(">>> STUDENT EMAIL RESULT =", result)

                    if result == 1:
                        print(">>> STUDENT EMAIL SENT SUCCESSFULLY ✓")
                        messages.success(
                            request,
                            f"✅ Student '{admission.name}' approved! "
                            f"Login credentials {admission.email} par send ho gaye."
                        )
                    else:
                        print(">>> STUDENT EMAIL NOT SENT (result != 1)")
                        messages.warning(
                            request,
                            f"⚠️ Student bana lekin email nahi gayi. "
                            f"Username: {username} | Password: {password}"
                        )

                except Exception as email_err:
                    print(">>> STUDENT EMAIL ERROR =", str(email_err))
                    messages.warning(
                        request,
                        f"⚠️ Student bana lekin email fail hui: {str(email_err)} | "
                        f"Username: {username} | Password: {password}"
                    )
            else:
                print(">>> NO STUDENT EMAIL ADDRESS FOUND")
                messages.warning(
                    request,
                    f"⚠️ Student approved lekin email address nahi mila. "
                    f"Username: {username} | Password: {password}"
                )

            # =========================================================
            # ✅ PARENT ACCOUNT BANAO + CREDENTIALS EMAIL BHEJO
            # =========================================================
            parent_email = getattr(admission, 'father_email', None)

            if parent_email:
                print(">>> Processing parent account for:", parent_email)
                try:
                    from parent_dashboard.models import Parent as ParentModel

                    # ── Parent ka Django User account banao ya existing use karo ──
                    parent_user_exists = User.objects.filter(email__iexact=parent_email).exists()

                    if parent_user_exists:
                        # Pehle se ek ya zyada accounts hain
                        # Sab users dhundo, pehla wala use karo, baqi delete karo
                        all_parent_users = User.objects.filter(email__iexact=parent_email).order_by('id')
                        parent_user = all_parent_users.first()
                        # Duplicate users delete karo
                        duplicate_ids = list(all_parent_users.values_list('id', flat=True)[1:])
                        if duplicate_ids:
                            User.objects.filter(id__in=duplicate_ids).delete()
                            print(f">>> {len(duplicate_ids)} duplicate parent users deleted")
                        parent_password = get_random_string(8)
                        parent_user.set_password(parent_password)
                        parent_user.save()
                        parent_username = parent_user.username
                        # Parent group assign karo agar nahi hai
                        parent_group, _ = Group.objects.get_or_create(name='Parent')
                        if not parent_user.groups.filter(name='Parent').exists():
                            parent_user.groups.add(parent_group)
                        print(">>> Existing parent user found — password reset:", parent_username)
                    else:
                        # Naya parent user banao
                        parent_username = f"PAR{get_random_string(6)}"
                        parent_password = get_random_string(8)
                        parent_user = User.objects.create_user(
                            username=parent_username,
                            email=parent_email,
                            password=parent_password,
                            first_name=admission.father_name or "Parent",
                        )
                        parent_group, _ = Group.objects.get_or_create(name='Parent')
                        parent_user.groups.set([parent_group])
                        print(">>> New parent user created:", parent_username)

                    # ── ParentModel bhi banao/update karo ──
                    parent_obj, parent_created = ParentModel.objects.get_or_create(
                        email=parent_email,
                        defaults={
                            'full_name': admission.father_name or "Parent",
                            'phone':     admission.contact or "",
                            'email':     parent_email,
                        }
                    )
                    # Parent user se link karo agar field available hai
                    if hasattr(parent_obj, 'user'):
                        parent_obj.user = parent_user
                        parent_obj.save()

                    # ── Student ko parent se link karo ──
                    student_obj = Student.objects.filter(student_id=username).first()
                    if student_obj:
                        parent_obj.students.add(student_obj)
                        print(">>> Student linked to parent ✓")

                    # ── Parent ko credentials ke saath email bhejo ──
                    parent_email_body = (
                        f"Dear {admission.father_name or 'Parent'} / Guardian,\n\n"
                        f"Assalam o Alaikum,\n\n"
                        f"We are pleased to inform you that your child's admission has been successfully approved.\n\n"
                        f"=============================\n"
                        f"  STUDENT DETAILS\n"
                        f"=============================\n"
                        f"Student Name  : {admission.name}\n"
                        f"Class         : {admission.class_fk}\n"
                        f"=============================\n\n"
                        f"  YOUR PARENT PORTAL LOGIN\n"
                        f"=============================\n"
                        f"Username  : {parent_username}\n"
                        f"Password  : {parent_password}\n"
                        f"Login URL : {login_url}\n"
                        f"=============================\n\n"
                        f"Please use the above credentials to login to the Parent Portal\n"
                        f"and track your child's progress, attendance, and results.\n\n"
                        f"For any queries, please contact the school administration.\n\n"
                        f"Regards,\n"
                        f"{school_name}\n"
                    )

                    parent_result = send_mail(
                        subject=f"Child Admission Approved & Parent Portal Access - {school_name}",
                        message=parent_email_body,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[parent_email],
                        fail_silently=False,
                    )
                    print(">>> PARENT EMAIL RESULT =", parent_result)

                    if parent_result == 1:
                        print(">>> PARENT EMAIL WITH CREDENTIALS SENT SUCCESSFULLY ✓")
                        messages.success(
                            request,
                            f"✅ Parent portal credentials '{parent_email}' par send ho gaye. "
                            f"Parent Username: {parent_username}"
                        )
                    else:
                        print(">>> PARENT EMAIL NOT SENT (result != 1)")
                        messages.warning(
                            request,
                            f"⚠️ Parent account bana lekin email nahi gayi. "
                            f"Parent Username: {parent_username} | Password: {parent_password}"
                        )

                except Exception as parent_err:
                    print(">>> PARENT ACCOUNT/EMAIL ERROR =", str(parent_err))
                    messages.warning(
                        request,
                        f"⚠️ Parent account ya email mein masla: {str(parent_err)}"
                    )
            else:
                print(">>> NO PARENT EMAIL (father_email) FOUND — skipping parent account creation")

        except Exception as main_err:
            print(">>> MAIN ERROR =", str(main_err))
            messages.error(request, f"❌ Student account nahi bana: {str(main_err)}")

        print("========== FUNCTION END ==========\n")
        return redirect('admission_list')

    # =========================================================
    # REJECTED / PENDING
    # =========================================================
    print(">>> STATUS is not approved — updating to:", status)
    admission.admission_status = status
    admission.save(update_fields=['admission_status'])
    print("========== FUNCTION END ==========\n")
    return redirect('admission_list')


# -------------------- query --------------------

@permission_required('admin_panel.view_admission', raise_exception=True)
def query(request):
    q_val        = request.GET.get('q', '').strip()
    status       = request.GET.get('status', '')
    academic_year = request.GET.get('academic_year', '')

    admissions = Admission.objects.select_related('section', 'class_fk', 'academic_year').order_by('id')

    if q_val:
        admissions = admissions.filter(
            Q(student_id__icontains=q_val)       |
            Q(name__icontains=q_val)             |
            Q(academic_year__year__icontains=q_val) |
            Q(admission_status__icontains=q_val) |
            Q(class_fk__class_name__icontains=q_val) |
            Q(section__section_name__icontains=q_val) |
            Q(email__icontains=q_val)            |
            Q(contact__icontains=q_val)          |
            Q(father_name__icontains=q_val)      |
            Q(mother_name__icontains=q_val)      |
            Q(campus__icontains=q_val)           |
            Q(branch__icontains=q_val)           |
            Q(ref_no__icontains=q_val)           |
            Q(dob__icontains=q_val)              |
            Q(gender__iexact=q_val)              |
            Q(address__icontains=q_val)          |
            Q(father_cnic__icontains=q_val)
        )

    if status:
        admissions = admissions.filter(admission_status=status)

    if academic_year:
        admissions = admissions.filter(academic_year=academic_year)

    paginator   = Paginator(admissions, 10)
    page_number = request.GET.get("page")
    page_obj    = paginator.get_page(page_number)

    context = {
        'admissions':   admissions,
        'query':        q_val,
        'status':       status,
        'academic_year': academic_year,
        'academic_years': AcademicYear.objects.all().order_by('-is_active', '-id'),
        'page_obj':     page_obj,
    }

    return render(request, 'admin_panel/query.html', context)


# -------------------- Classes CRUD --------------------
from django.db.models import Count, Q
from .models import Class
from .forms import ClassForm


@permission_required('admin_panel.view_class', raise_exception=True)
def class_list(request):
    from django.db.models import Avg, F
    from django.utils import timezone
    from student_profile.models import Student as PortalStudent
    from teacher_dashboard.models import Attendance, Assignment
    from .models import AssignedPeriod, ClassTeacher, ExamResult, FeeVoucher, Student as FeeStudent, Subject

    today = timezone.localdate()
    selected = {
        'q': request.GET.get('q', '').strip(),
        'class_name': request.GET.get('class_name', '').strip(),
        'grade': request.GET.get('grade', '').strip(),
        'section': request.GET.get('section', '').strip(),
        'teacher': request.GET.get('teacher', '').strip(),
        'campus': request.GET.get('campus', '').strip(),
        'academic_session': request.GET.get('academic_session', '').strip(),
        'curriculum': request.GET.get('curriculum', '').strip(),
        'house': request.GET.get('house', '').strip(),
        'shift': request.GET.get('shift', '').strip(),
        'status': request.GET.get('status', '').strip(),
        'strength': request.GET.get('strength', '').strip(),
        'timetable_status': request.GET.get('timetable_status', '').strip(),
        'attendance': request.GET.get('attendance', '').strip(),
        'fee': request.GET.get('fee', '').strip(),
        'performance': request.GET.get('performance', '').strip(),
    }

    classes_qs = Class.objects.select_related('group').order_by('class_name')
    if selected['q']:
        classes_qs = classes_qs.filter(
            Q(class_name__icontains=selected['q']) |
            Q(group__group_name__icontains=selected['q'])
        )
    if selected['class_name']:
        classes_qs = classes_qs.filter(class_name__icontains=selected['class_name'])
    if selected['grade']:
        classes_qs = classes_qs.filter(class_name__icontains=selected['grade'])
    if selected['curriculum']:
        classes_qs = classes_qs.filter(group__group_name__icontains=selected['curriculum'])

    academic_years = list(AcademicYear.objects.all().order_by('-is_active', '-id'))
    sections = list(Section.objects.select_related('academic_year', 'class_fk').order_by('class_fk__class_name', 'section_name'))
    class_teachers = list(ClassTeacher.objects.select_related('academic_year', 'class_fk', 'section', 'teacher'))
    assigned_periods = list(AssignedPeriod.objects.select_related('class_fk', 'section', 'subject', 'teacher'))

    if selected['academic_session']:
        sections = [s for s in sections if str(s.academic_year_id) == selected['academic_session']]
        class_teachers = [ct for ct in class_teachers if str(ct.academic_year_id) == selected['academic_session']]
    if selected['section']:
        sections = [s for s in sections if s.section_name.lower() == selected['section'].lower()]
    if selected['teacher']:
        class_teachers = [ct for ct in class_teachers if str(ct.teacher_id) == selected['teacher']]

    student_counts = PortalStudent.objects.values('class_fk_id').annotate(total=Count('id'))
    student_count_map = {item['class_fk_id']: item['total'] for item in student_counts}
    male_counts = PortalStudent.objects.filter(gender__iexact='Male').values('class_fk_id').annotate(total=Count('id'))
    female_counts = PortalStudent.objects.filter(gender__iexact='Female').values('class_fk_id').annotate(total=Count('id'))
    male_map = {item['class_fk_id']: item['total'] for item in male_counts}
    female_map = {item['class_fk_id']: item['total'] for item in female_counts}

    records = []
    ai_alerts = []
    totals = {
        'classes': 0,
        'active_classes': 0,
        'students': 0,
        'capacity': 0,
        'vacant_classes': 0,
        'without_teacher': 0,
        'without_timetable': 0,
        'low_attendance': 0,
        'ai_alerts': 0,
    }

    for class_obj in classes_qs:
        class_sections = [s for s in sections if s.class_fk_id == class_obj.id]
        if selected['section'] and not class_sections:
            continue

        teacher_assignments = [ct for ct in class_teachers if ct.class_fk_id == class_obj.id]
        if selected['teacher'] and not teacher_assignments:
            continue

        timetable_periods = [ap for ap in assigned_periods if ap.class_fk_id == class_obj.id]
        if selected['campus']:
            campus_exists = Admission.objects.filter(class_fk=class_obj, campus__icontains=selected['campus']).exists()
            if not campus_exists:
                continue

        section_names = sorted({s.section_name for s in class_sections}) or ['Not Assigned']
        session_names = sorted({s.academic_year.year for s in class_sections if s.academic_year}) or ['Not Assigned']
        teacher_names = sorted({ct.teacher.name for ct in teacher_assignments if ct.teacher}) or ['Not Assigned']
        capacity = sum(s.capacity for s in class_sections)
        students = student_count_map.get(class_obj.id, 0)
        male = male_map.get(class_obj.id, 0)
        female = female_map.get(class_obj.id, 0)
        vacant = max(capacity - students, 0) if capacity else 0
        capacity_percent = min(round((students / capacity) * 100), 100) if capacity else 0

        present_today = Attendance.objects.filter(class_fk=class_obj, date=today, status='present').count()
        attendance_marked = Attendance.objects.filter(class_fk=class_obj, date=today).count()
        attendance_percent = round((present_today / attendance_marked) * 100) if attendance_marked else 0

        result_avg = ExamResult.objects.filter(class_fk=class_obj).aggregate(
            avg=Avg(100.0 * F('marks_obtained') / F('total_marks'))
        )['avg']
        performance_percent = round(result_avg or 0)

        finance_students = FeeStudent.objects.filter(current_class__iexact=class_obj.class_name)
        fee_total = FeeVoucher.objects.filter(student__in=finance_students).count()
        fee_paid = FeeVoucher.objects.filter(student__in=finance_students, status='PAID').count()
        fee_percent = round((fee_paid / fee_total) * 100) if fee_total else 0

        homework_pending = Assignment.objects.filter(class_fk=class_obj, due_date__gte=today).count()
        subject_names = list(Subject.objects.filter(class_fk=class_obj).order_by('sort_order', 'name').values_list('name', flat=True)[:6])
        top_performers = list(PortalStudent.objects.filter(class_fk=class_obj).order_by('roll_no', 'name').values_list('name', flat=True)[:3])
        weak_students = list(PortalStudent.objects.filter(class_fk=class_obj).order_by('-id').values_list('name', flat=True)[:3])

        if not class_sections:
            status = 'Vacant'
        elif students >= capacity and capacity:
            status = 'Full'
        else:
            status = 'Active'

        if not timetable_periods:
            timetable_status = 'Not Assigned'
        elif len(timetable_periods) < max(len(class_sections), 1) * 5:
            timetable_status = 'Pending'
        else:
            timetable_status = 'Completed'

        if selected['status'] and selected['status'].lower() != status.lower():
            continue
        if selected['timetable_status'] and selected['timetable_status'].lower() != timetable_status.lower():
            continue

        if attendance_percent and attendance_percent < 75:
            ai_level = 'Critical'
            ai_tone = 'critical'
            recommendation = 'Review attendance and contact families today.'
        elif not teacher_assignments or timetable_status != 'Completed' or performance_percent < 70:
            ai_level = 'Attention Needed'
            ai_tone = 'warning'
            recommendation = 'Assign missing ownership and schedule a revision session.'
        else:
            ai_level = 'Excellent'
            ai_tone = 'healthy'
            recommendation = 'Keep current timetable, homework, and fee follow-up rhythm.'

        if selected['attendance'] == 'low' and attendance_percent >= 75:
            continue
        if selected['fee'] == 'low' and fee_percent >= 75:
            continue
        if selected['performance'] == 'low' and performance_percent >= 70:
            continue
        if selected['strength'] == 'full' and status != 'Full':
            continue
        if selected['strength'] == 'vacant' and not vacant:
            continue

        alert_text = ''
        if ai_tone == 'critical':
            alert_text = 'Low attendance detected'
        elif not teacher_assignments:
            alert_text = 'Class teacher not assigned'
        elif timetable_status != 'Completed':
            alert_text = 'Timetable needs completion'
        elif performance_percent < 70:
            alert_text = 'Performance needs review'

        if alert_text:
            ai_alerts.append(f"{class_obj.class_name}: {alert_text}")

        records.append({
            'class_obj': class_obj,
            'sections': ', '.join(section_names),
            'sessions': ', '.join(session_names),
            'campus': Admission.objects.filter(class_fk=class_obj).values_list('campus', flat=True).first() or 'Main Campus',
            'curriculum': class_obj.group.group_name if class_obj.group else 'Core',
            'teacher_names': ', '.join(teacher_names),
            'students': students,
            'male': male,
            'female': female,
            'capacity': capacity,
            'vacant': vacant,
            'capacity_percent': capacity_percent,
            'attendance_percent': attendance_percent,
            'performance_percent': performance_percent,
            'homework_percent': max(100 - min(homework_pending * 8, 100), 0),
            'homework_pending': homework_pending,
            'fee_percent': fee_percent,
            'timetable_status': timetable_status,
            'status': status,
            'ai_level': ai_level,
            'ai_tone': ai_tone,
            'recommendation': recommendation,
            'alert_text': alert_text or 'No operational alert',
            'subjects': subject_names,
            'top_performers': top_performers,
            'weak_students': weak_students,
            'period_count': len(timetable_periods),
            'section_count': len(class_sections),
        })

        totals['classes'] += 1
        totals['students'] += students
        totals['capacity'] += capacity
        totals['active_classes'] += 1 if status == 'Active' else 0
        totals['vacant_classes'] += 1 if not students else 0
        totals['without_teacher'] += 1 if not teacher_assignments else 0
        totals['without_timetable'] += 1 if timetable_status == 'Not Assigned' else 0
        totals['low_attendance'] += 1 if attendance_percent and attendance_percent < 75 else 0
        totals['ai_alerts'] += 1 if alert_text else 0

    totals['average_strength'] = round(totals['students'] / totals['classes']) if totals['classes'] else 0

    context = {
        'classes': records,
        'summary': totals,
        'selected': selected,
        'academic_years': academic_years,
        'section_options': sorted({s.section_name for s in Section.objects.all()}),
        'teacher_options': ClassTeacher.objects.select_related('teacher').order_by('teacher__name'),
        'campus_options': sorted({c for c in Admission.objects.exclude(campus='').values_list('campus', flat=True).distinct()}),
        'curriculum_options': Class.objects.exclude(group__isnull=True).select_related('group').values_list('group__group_name', flat=True).distinct(),
        'ai_alerts': ai_alerts[:6],
    }
    return render(request, 'admin_panel/class_list.html', context)


@permission_required('admin_panel.view_class', raise_exception=True)
def class_students(request, pk):
    from django.db.models import Avg, F
    from django.utils import timezone
    from student_profile.models import AssignmentSubmission, Student as PortalStudent
    from teacher_dashboard.models import Assignment, Attendance
    from .models import ExamResult, FeeVoucher, Student as FeeStudent

    class_obj = get_object_or_404(Class.objects.select_related('group'), pk=pk)
    today = timezone.localdate()
    selected = {
        'q': request.GET.get('q', '').strip(),
        'section': request.GET.get('section', '').strip(),
        'gender': request.GET.get('gender', '').strip(),
        'academic_year': request.GET.get('academic_year', '').strip(),
        'attendance': request.GET.get('attendance', '').strip(),
        'performance': request.GET.get('performance', '').strip(),
        'missing': request.GET.get('missing', '').strip(),
    }

    students_qs = PortalStudent.objects.select_related('academic_year', 'class_fk', 'section').filter(class_fk=class_obj)
    if selected['q']:
        students_qs = students_qs.filter(
            Q(name__icontains=selected['q']) |
            Q(student_id__icontains=selected['q']) |
            Q(roll_no__icontains=selected['q']) |
            Q(email__icontains=selected['q']) |
            Q(phone__icontains=selected['q']) |
            Q(father_name__icontains=selected['q']) |
            Q(mother_name__icontains=selected['q'])
        )
    if selected['section']:
        students_qs = students_qs.filter(section_id=selected['section'])
    if selected['gender']:
        students_qs = students_qs.filter(gender__iexact=selected['gender'])
    if selected['academic_year']:
        students_qs = students_qs.filter(academic_year_id=selected['academic_year'])
    if selected['missing'] == 'photo':
        students_qs = students_qs.filter(Q(photo='') | Q(photo__isnull=True))
    if selected['missing'] == 'contact':
        students_qs = students_qs.filter(Q(phone='') | Q(phone__isnull=True) | Q(email=''))

    sections = Section.objects.filter(class_fk=class_obj).select_related('academic_year').order_by('section_name')
    academic_years = AcademicYear.objects.filter(section__class_fk=class_obj).distinct().order_by('-is_active', '-id')
    assignments = Assignment.objects.filter(class_fk=class_obj)
    assignment_ids = list(assignments.values_list('id', flat=True))
    total_assignments = assignments.count()

    student_cards = []
    total_students = 0
    attendance_sum = 0
    performance_sum = 0
    fee_sum = 0
    risk_count = 0

    for student in students_qs.order_by('section__section_name', 'roll_no', 'name'):
        attendance_total = Attendance.objects.filter(student=student, class_fk=class_obj).count()
        attendance_present = Attendance.objects.filter(student=student, class_fk=class_obj, status='present').count()
        attendance_percent = round((attendance_present / attendance_total) * 100) if attendance_total else 0

        result_avg = ExamResult.objects.filter(student=student, class_fk=class_obj).aggregate(
            avg=Avg(100.0 * F('marks_obtained') / F('total_marks'))
        )['avg']
        performance_percent = round(result_avg or 0)

        submitted_assignments = AssignmentSubmission.objects.filter(
            student=student,
            assignment_id__in=assignment_ids
        ).count() if assignment_ids else 0
        pending_assignments = max(total_assignments - submitted_assignments, 0)

        finance_student = FeeStudent.objects.filter(student_id=student.student_id).first()
        if not finance_student:
            finance_student = FeeStudent.objects.filter(full_name__iexact=student.name, current_class__iexact=class_obj.class_name).first()
        fee_total = FeeVoucher.objects.filter(student=finance_student).count() if finance_student else 0
        fee_paid = FeeVoucher.objects.filter(student=finance_student, status='PAID').count() if finance_student else 0
        fee_percent = round((fee_paid / fee_total) * 100) if fee_total else 0
        fee_status = 'Paid' if fee_total and fee_paid == fee_total else 'Pending' if fee_total else 'Not Linked'

        if attendance_percent and attendance_percent < 75:
            ai_level = 'Critical'
            ai_tone = 'critical'
        elif pending_assignments or performance_percent < 70 or fee_status == 'Pending':
            ai_level = 'Attention Needed'
            ai_tone = 'warning'
        else:
            ai_level = 'Healthy'
            ai_tone = 'healthy'

        if selected['attendance'] == 'low' and attendance_percent >= 75:
            continue
        if selected['performance'] == 'low' and performance_percent >= 70:
            continue

        student_cards.append({
            'student': student,
            'attendance_percent': attendance_percent,
            'performance_percent': performance_percent,
            'total_assignments': total_assignments,
            'pending_assignments': pending_assignments,
            'fee_percent': fee_percent,
            'fee_status': fee_status,
            'ai_level': ai_level,
            'ai_tone': ai_tone,
        })
        total_students += 1
        attendance_sum += attendance_percent
        performance_sum += performance_percent
        fee_sum += fee_percent
        risk_count += 1 if ai_tone != 'healthy' else 0

    summary = {
        'students': total_students,
        'average_attendance': round(attendance_sum / total_students) if total_students else 0,
        'average_performance': round(performance_sum / total_students) if total_students else 0,
        'average_fee': round(fee_sum / total_students) if total_students else 0,
        'risk_count': risk_count,
    }

    context = {
        'class_obj': class_obj,
        'students': student_cards,
        'summary': summary,
        'selected': selected,
        'sections': sections,
        'academic_years': academic_years,
    }
    return render(request, 'admin_panel/class_students.html', context)


@permission_required('admin_panel.add_class', raise_exception=True)
def class_create(request):
    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('class_list')
    else:
        form = ClassForm()
    return render(request, 'admin_panel/class_form.html', {'form': form})


@permission_required('admin_panel.change_class', raise_exception=True)
def class_update(request, pk):
    class_obj = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        form = ClassForm(request.POST, instance=class_obj)
        if form.is_valid():
            form.save()
            return redirect('class_list')
    else:
        form = ClassForm(instance=class_obj)
    return render(request, 'admin_panel/class_form.html', {'form': form})


@permission_required('admin_panel.delete_class', raise_exception=True)
def class_delete(request, pk):
    class_obj = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        class_obj.delete()
        return redirect('class_list')
    return render(request, 'admin_panel/class_confirm_delete.html', {'class': class_obj})


# -------------------- Academic Year CRUD --------------------
from .models import AcademicYear
from .forms import AcademicYearForm


@permission_required('admin_panel.view_academicyear', raise_exception=True)
def academic_year_list(request):
    years = AcademicYear.objects.all()
    return render(request, 'admin_panel/academic_year_list.html', {'years': years})


@permission_required('admin_panel.add_academicyear', raise_exception=True)
def add_academic_year(request):
    if request.method == 'POST':
        form = AcademicYearForm(request.POST)
        if form.is_valid():
            if form.cleaned_data['is_active']:
                AcademicYear.objects.update(is_active=False)
            form.save()
            return redirect('academic_year_list')
    else:
        form = AcademicYearForm()
    return render(request, 'admin_panel/academic_year_form.html', {'form': form})


@permission_required('admin_panel.change_academicyear', raise_exception=True)
def update_academic_year(request, pk):
    year = get_object_or_404(AcademicYear, pk=pk)
    if request.method == 'POST':
        form = AcademicYearForm(request.POST, instance=year)
        if form.is_valid():
            if form.cleaned_data['is_active']:
                AcademicYear.objects.update(is_active=False)
            form.save()
            return redirect('academic_year_list')
    else:
        form = AcademicYearForm(instance=year)
    return render(request, 'admin_panel/academic_year_form.html', {'form': form})


@permission_required('admin_panel.delete_academicyear', raise_exception=True)
def delete_academic_year(request, pk):
    year = get_object_or_404(AcademicYear, pk=pk)
    if request.method == 'POST':
        year.delete()
        return redirect('academic_year_list')
    return render(request, 'admin_panel/academic_year_confirm_delete.html', {'year': year})


# -------------------- Section CRUD --------------------
from .models import Section, AcademicYear, Class, Admission
from .forms import SectionForm


@permission_required('admin_panel.add_section', raise_exception=True)
def add_section(request):
    if request.method == 'POST':
        form = SectionForm(request.POST)
        form.fields['academic_year'].queryset = AcademicYear.objects.filter(is_active=True)
        if form.is_valid():
            form.save()
            return redirect('section_list')
    else:
        form = SectionForm()
        form.fields['academic_year'].queryset = AcademicYear.objects.filter(is_active=True)
    return render(request, 'admin_panel/add_section.html', {'form': form})


from django.db.models import F, IntegerField, ExpressionWrapper


@permission_required('admin_panel.view_section', raise_exception=True)
def section_list(request):
    sections = Section.objects.select_related('academic_year', 'class_fk').annotate(
        approved_admissions_count=Count(
            'admission',
            filter=Q(admission__admission_status='approved')
        ),
        available_seats=ExpressionWrapper(
            F('capacity') - Count(
                'admission',
                filter=Q(admission__admission_status='approved')
            ),
            output_field=IntegerField()
        )
    )
    return render(request, 'admin_panel/section_list.html', {'sections': sections})


@permission_required('admin_panel.change_section', raise_exception=True)
def edit_section(request, pk):
    section = get_object_or_404(Section, pk=pk)
    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            return redirect('section_list')
    else:
        form = SectionForm(instance=section)
    return render(request, 'admin_panel/edit_section.html', {'form': form})


@permission_required('admin_panel.delete_section', raise_exception=True)
def delete_section(request, pk):
    section = get_object_or_404(Section, pk=pk)
    if request.method == 'POST':
        section.delete()
        return redirect('section_list')
    return render(request, 'admin_panel/delete_section.html', {'section': section})


# -------------------- Subject CRUD --------------------
from .models import Subject
from .forms import SubjectForm


@permission_required('admin_panel.view_subject', raise_exception=True)
def subject_list(request):
    subjects = Subject.objects.all()
    return render(request, 'admin_panel/subject_list.html', {'subjects': subjects})


@permission_required('admin_panel.add_subject', raise_exception=True)
def add_subject(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('subject_list')
        else:
            print("Form errors:", form.errors)
    else:
        form = SubjectForm()
    return render(request, 'admin_panel/subject_form.html', {'form': form, 'subject': None})


@permission_required('admin_panel.change_subject', raise_exception=True)
def edit_subject(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            return redirect('subject_list')
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'admin_panel/subject_form.html', {'form': form, 'subject': subject})


@permission_required('admin_panel.delete_subject', raise_exception=True)
def delete_subject(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        return redirect('subject_list')
    return render(request, 'admin_panel/subject_confirm_delete.html', {'subject': subject})


# -------------------- Teacher CRUD --------------------
from teacher_dashboard.models import Teacher
from admin_panel.forms import TeacherForm
from django.core.paginator import Paginator
from django.utils.text import slugify
import uuid


@permission_required('teacher_dashboard.view_teacher', raise_exception=True)
def teacher_list(request):
    teachers = Teacher.objects.all()
    paginator = Paginator(teachers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'admin_panel/teacher_list.html', {
        'page_obj': page_obj,
        'teachers': page_obj.object_list
    })


@permission_required('teacher_dashboard.view_teacher', raise_exception=True)
def teacher_profile(request, pk):
    teacher = get_object_or_404(Teacher.objects.prefetch_related('subjects'), pk=pk)

    from admin_panel.models import (
        AssignedPeriod,
        ClassTeacher,
        TeacherAppraisalSubmission,
        TeacherFixture,
    )
    from teacher_dashboard.models import Attendance

    assigned_periods_qs = (
        AssignedPeriod.objects
        .filter(teacher=teacher)
        .select_related('class_fk', 'section', 'subject', 'period')
        .order_by('day', 'period__start_time', 'period__id')
    )
    substitute_fixtures_qs = (
        TeacherFixture.objects
        .filter(substitute_teacher=teacher)
        .select_related('class_fk', 'section', 'subject', 'period', 'absent_teacher')
        .order_by('-fixture_date', 'period__start_time')
    )
    absent_fixtures_qs = TeacherFixture.objects.filter(absent_teacher=teacher)
    class_teacher_assignments = (
        ClassTeacher.objects
        .filter(teacher=teacher)
        .select_related('class_fk', 'section', 'academic_year')
        .order_by('-assigned_date')
    )

    stats = {
        'subjects_count': teacher.subjects.count(),
        'assigned_periods_count': assigned_periods_qs.count(),
        'substitute_duties_count': substitute_fixtures_qs.count(),
        'absent_fixture_count': absent_fixtures_qs.count(),
        'attendance_marked_count': Attendance.objects.filter(marked_by=teacher).count(),
        'appraisal_submissions_count': TeacherAppraisalSubmission.objects.filter(teacher=teacher).count(),
        'class_teacher_count': class_teacher_assignments.count(),
        'employee_linked': bool(getattr(teacher, 'employee_id', None)),
    }

    context = {
        'teacher': teacher,
        'stats': stats,
        'assigned_periods': assigned_periods_qs[:12],
        'substitute_fixtures': substitute_fixtures_qs[:8],
        'class_teacher_assignments': class_teacher_assignments[:6],
    }
    return render(request, 'admin_panel/teacher_profile.html', context)


DEFAULT_TEACHER_PASSWORD = "ex"


@permission_required("teacher_dashboard.add_teacher", raise_exception=True)
def teacher_create(request):
    teacher_group, _ = Group.objects.get_or_create(name="Teacher")

    if request.method == "POST":
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            teacher = form.save(commit=False)

            if teacher.user:
                teacher.save()
                form.save_m2m()
                messages.success(request, "Teacher updated.")
                return redirect("teacher_list")

            if User.objects.filter(email__iexact=teacher.email).exists():
                messages.error(request, "A user with this email already exists.")
                return render(request, "admin_panel/teacher_form.html", {"form": form})

            username = form.cleaned_data["login_id"].strip()
            password = form.cleaned_data["password"]

            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    email=teacher.email,
                    password=password,
                    first_name=teacher.name
                )
                user.groups.set([teacher_group])
                teacher.user = user
                teacher.save()
                form.save_m2m()

            try:
                send_mail(
                    subject="Teacher account created",
                    message=f"Username: {user.username}\nPassword: {password}\nLogin URL: {getattr(settings, 'SITE_URL', 'http://127.0.0.1:8000')}/login/teacher/",
                    from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                    recipient_list=[teacher.email],
                    fail_silently=True
                )
            except Exception:
                pass

            messages.success(request, "Teacher added and user created.")
            return redirect("teacher_list")

        messages.error(request, "Please fix the errors below.")
    else:
        form = TeacherForm()

    return render(request, "admin_panel/teacher_form.html", {"form": form})


@permission_required('teacher_dashboard.change_teacher', raise_exception=True)
def teacher_update(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)

    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES, instance=teacher)
        selected_faculties = request.POST.getlist('faculty_group')
        if form.is_valid():
            form.save()
            messages.success(request, 'Teacher updated successfully.')
            return redirect('teacher_list')
    else:
        form = TeacherForm(instance=teacher)
        selected_faculties = teacher.faculty_group

    return render(request, 'admin_panel/teacher_form.html', {
        'form': form, 'title': 'Edit Teacher', 'selected_faculties': selected_faculties
    })


@permission_required('teacher_dashboard.delete_teacher', raise_exception=True)
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        teacher.delete()
        messages.success(request, 'Teacher deleted successfully.')
        return redirect('teacher_list')
    return render(request, 'admin_panel/teacher_confirm_delete.html', {'teacher': teacher})


# ==================== PERIOD PLANNING ====================
from .models import CreatePeriod, Subject, SubjectPeriod, Section
from .forms import CreatePeriodForm, AssignedPeriodForm
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt


@permission_required('admin_panel.add_createperiod', raise_exception=True)
def create_period_view(request):
    if request.method == 'POST':
        form = CreatePeriodForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('period_list')
    else:
        form = CreatePeriodForm()
    return render(request, 'admin_panel/create_period.html', {'form': form})


@permission_required('admin_panel.view_createperiod', raise_exception=True)
def period_list_view(request):
    periods = CreatePeriod.objects.all().order_by('day', 'start_time')
    return render(request, 'admin_panel/period_list.html', {'periods': periods})


@permission_required('admin_panel.change_createperiod', raise_exception=True)
def update_period_view(request, pk):
    period = get_object_or_404(CreatePeriod, pk=pk)
    if request.method == 'POST':
        form = CreatePeriodForm(request.POST, instance=period)
        if form.is_valid():
            form.save()
            return redirect('period_list')
    else:
        form = CreatePeriodForm(instance=period)
    return render(request, 'admin_panel/update_period.html', {'form': form})


@permission_required('admin_panel.delete_createperiod', raise_exception=True)
def delete_period_view(request, pk):
    period = get_object_or_404(CreatePeriod, pk=pk)
    if request.method == 'POST':
        period.delete()
        return redirect('period_list')
    return render(request, 'admin_panel/delete_period.html', {'period': period})


# -------------------- Timetable View --------------------
from .models import (
    Subject, SubjectPeriod, Section, Class, ClassGroup, AssignedPeriod,
    TeacherFixture, ClassTeacher, TeacherAbsence, TeacherFixtureCandidateLog,
    TeacherFixtureNotificationLog
)


@permission_required('admin_panel.view_subjectperiod', raise_exception=True)
def timetable_automation(request):
    groups = ClassGroup.objects.all().order_by('group_name')
    selected_group_id = request.GET.get('group')
    selected_group = None
    automation_result = None

    if request.method == 'POST':
        action = request.POST.get('action', 'run')

        if action in ['cancel_fixture', 'mark_uncovered', 'rerun_fixture', 'replace_substitute']:
            fixture = get_object_or_404(TeacherFixture.objects.select_related('assigned_period', 'substitute_teacher'), id=request.POST.get('fixture_id'))
            override_reason = request.POST.get('override_reason', '').strip() or f'Admin action: {action}'

            if action == 'cancel_fixture':
                fixture.fixture_status = 'cancelled'
                fixture.override_reason = override_reason
                fixture.overridden_by = request.user
                fixture.overridden_at = timezone.now()
                fixture.assignment_mode = 'overridden'
                fixture.save(update_fields=['fixture_status', 'override_reason', 'overridden_by', 'overridden_at', 'assignment_mode'])
                messages.success(request, 'Fixture cancelled and override logged.')
                return redirect('timetable_automation')

            if action == 'mark_uncovered':
                fixture.fixture_status = 'uncovered'
                fixture.override_reason = override_reason
                fixture.overridden_by = request.user
                fixture.overridden_at = timezone.now()
                fixture.assignment_mode = 'overridden'
                fixture.save(update_fields=['fixture_status', 'override_reason', 'overridden_by', 'overridden_at', 'assignment_mode'])
                messages.success(request, 'Fixture marked as uncovered.')
                return redirect('timetable_automation')

            if action == 'replace_substitute':
                substitute_id = request.POST.get('substitute_teacher_id')
                replacement = get_object_or_404(Teacher, id=substitute_id, status='active')
                if not fixture.original_substitute_teacher_id:
                    fixture.original_substitute_teacher = fixture.substitute_teacher
                fixture.substitute_teacher = replacement
                fixture.fixture_status = 'reassigned'
                fixture.assignment_mode = 'overridden'
                fixture.override_reason = override_reason
                fixture.overridden_by = request.user
                fixture.overridden_at = timezone.now()
                fixture.save()
                messages.success(request, f'Fixture reassigned to {replacement.name}.')
                return redirect('timetable_automation')

            if action == 'rerun_fixture' and fixture.assigned_period:
                fixture.fixture_status = 'cancelled'
                fixture.override_reason = override_reason
                fixture.overridden_by = request.user
                fixture.overridden_at = timezone.now()
                fixture.assignment_mode = 'overridden'
                fixture.save(update_fields=['fixture_status', 'override_reason', 'overridden_by', 'overridden_at', 'assignment_mode'])
                automation_result = FixtureAutomationService.run_for_absence(
                    teacher=fixture.absent_teacher,
                    target_date=fixture.fixture_date,
                    source='manual',
                    triggered_by=request.user,
                )
                messages.success(request, 'Fixture cancelled and automation re-run.')
            elif action == 'rerun_fixture':
                messages.error(request, 'This fixture cannot be re-run because it is not linked to an assigned period.')
                return redirect('timetable_automation')

        teacher_id = request.POST.get('teacher_id')
        absence_date = request.POST.get('absence_date')

        if automation_result is None and (not teacher_id or not absence_date):
            messages.error(request, 'Please select absent teacher and date.')
            return redirect('timetable_automation')

        if automation_result is None:
            absent_teacher = get_object_or_404(Teacher, id=teacher_id)
            automation_result = FixtureAutomationService.run_for_absence(
                teacher=absent_teacher,
                target_date=absence_date,
                source='manual',
                triggered_by=request.user,
            )

        assigned_count = len(automation_result.get('assigned', []))
        unassigned_count = len(automation_result.get('unassigned', []))
        skipped_count = len(automation_result.get('skipped', []))
        if assigned_count:
            messages.success(
                request,
                f'Auto fixture completed: {assigned_count} assigned, {unassigned_count} unassigned, {skipped_count} skipped.'
            )
        else:
            messages.warning(
                request,
                f'Auto fixture completed with no new assignments: {unassigned_count} unassigned, {skipped_count} skipped.'
            )

    if selected_group_id:
        selected_group = get_object_or_404(ClassGroup, id=selected_group_id)
    elif groups.exists():
        selected_group = groups.first()

    group_subject_count = 0
    configured_subject_count = 0
    missing_allocation_count = 0
    planned_periods = 0
    assigned_group_periods = 0

    if selected_group:
        group_class_ids = Class.objects.filter(group=selected_group).values_list('id', flat=True)
        group_subjects = Subject.objects.filter(class_fk_id__in=group_class_ids).order_by('name')
        subject_names = set(group_subjects.values_list('name', flat=True))
        group_subject_count = len(subject_names)

        allocation_map = {}
        for sp in SubjectPeriod.objects.filter(group=selected_group, subject__class_fk_id__in=group_class_ids).select_related('subject'):
            key = (sp.subject.name, sp.day)
            allocation_map[key] = max(allocation_map.get(key, 0), sp.periods)

        planned_by_subject = {}
        for (subject_name, _day), periods in allocation_map.items():
            planned_by_subject[subject_name] = planned_by_subject.get(subject_name, 0) + periods

        configured_subject_count = len([name for name, total in planned_by_subject.items() if total > 0])
        missing_allocation_count = max(group_subject_count - configured_subject_count, 0)
        planned_periods = sum(planned_by_subject.values())
        assigned_group_periods = AssignedPeriod.objects.filter(class_fk_id__in=group_class_ids).count()

    duplicate_teacher_slots = AssignedPeriod.objects.values(
        'teacher_id', 'day', 'period_id'
    ).annotate(total=Count('id')).filter(total__gt=1).count()

    recent_fixtures = TeacherFixture.objects.select_related(
        'class_fk', 'section', 'period', 'subject', 'absent_teacher', 'substitute_teacher', 'assigned_period'
    ).order_by('-id')[:5]

    today = date.today()
    fixture_stats = {
        'total': TeacherFixture.objects.count(),
        'today': TeacherFixture.objects.filter(fixture_date=today).count(),
        'absent_teachers': TeacherFixture.objects.values('absent_teacher_id').distinct().count(),
        'substitute_teachers': TeacherFixture.objects.values('substitute_teacher_id').distinct().count(),
    }

    classes = Class.objects.all().order_by('class_name')
    sections = Section.objects.select_related('class_fk').order_by('class_fk__class_name', 'section_name')
    class_teacher_count = ClassTeacher.objects.count()
    teachers = Teacher.objects.filter(status='active').order_by('name')
    today_absences = TeacherAbsence.objects.filter(absence_date=today).select_related('teacher').order_by('-created_at')
    latest_absence = None
    if automation_result:
        latest_absence = automation_result.get('absence')
    elif today_absences.exists():
        latest_absence = today_absences.first()
    latest_absence_results = []
    if latest_absence:
        latest_absence_results = latest_absence.results.select_related(
            'assigned_period__class_fk', 'assigned_period__section', 'assigned_period__subject',
            'assigned_period__period', 'fixture__substitute_teacher'
        )
    visible_fixtures = TeacherFixture.objects.select_related(
        'class_fk', 'section', 'period', 'subject', 'absent_teacher', 'substitute_teacher', 'assigned_period'
    ).order_by('-fixture_date', '-id')[:10]
    uncovered_results = TeacherAbsence.objects.filter(
        results__status='unassigned'
    ).select_related('teacher').distinct().order_by('-absence_date')[:10]
    candidate_logs = TeacherFixtureCandidateLog.objects.select_related(
        'teacher', 'assigned_period__subject', 'assigned_period__period'
    ).order_by('-created_at')[:20]
    notification_logs = TeacherFixtureNotificationLog.objects.select_related(
        'recipient_teacher', 'fixture'
    ).order_by('-created_at')[:10]
    workflow_steps = FixtureAutomationService.workflow_steps()

    return render(request, 'admin_panel/timetable_automation.html', {
        'groups': groups,
        'selected_group': selected_group,
        'group_subject_count': group_subject_count,
        'configured_subject_count': configured_subject_count,
        'missing_allocation_count': missing_allocation_count,
        'planned_periods': planned_periods,
        'assigned_group_periods': assigned_group_periods,
        'unassigned_periods': max(planned_periods - assigned_group_periods, 0),
        'assignment_count': AssignedPeriod.objects.count(),
        'duplicate_teacher_slots': duplicate_teacher_slots,
        'fixture_stats': fixture_stats,
        'recent_fixtures': recent_fixtures,
        'classes': classes,
        'sections': sections,
        'class_teacher_count': class_teacher_count,
        'teachers': teachers,
        'today_absences': today_absences,
        'latest_absence': latest_absence,
        'latest_absence_results': latest_absence_results,
        'today': today,
        'visible_fixtures': visible_fixtures,
        'uncovered_results': uncovered_results,
        'candidate_logs': candidate_logs,
        'notification_logs': notification_logs,
        'workflow_steps': workflow_steps,
    })


@permission_required('admin_panel.view_subjectperiod', raise_exception=True)
def timetable_view(request):
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    groups = ClassGroup.objects.all()
    selected_group_id = request.GET.get('group') or request.POST.get('group')
    subjects = []
    saved_periods = {}

    if selected_group_id:
        selected_group = ClassGroup.objects.get(id=selected_group_id)
        classes = Class.objects.filter(group=selected_group)
        class_ids = classes.values_list('id', flat=True)
        all_subjects = Subject.objects.filter(class_fk_id__in=class_ids).order_by('name')

        seen = set()
        for subj in all_subjects:
            if subj.name not in seen:
                seen.add(subj.name)
                subjects.append(subj)

        all_sp = SubjectPeriod.objects.filter(group=selected_group)
        for sp in all_sp:
            saved_periods.setdefault(sp.subject_id, {})[sp.day] = sp.periods

    timetable_summary = {
        'selected_group_name': '',
        'total_subjects': len(subjects),
        'weekly_allocated_periods': 0,
        'remaining_periods': 36,
        'missing_subjects': 0,
        'is_over_limit': False,
    }

    if selected_group_id:
        selected_group = ClassGroup.objects.get(id=selected_group_id)
        timetable_summary['selected_group_name'] = selected_group.group_name
        for subject in subjects:
            subject_total = sum((saved_periods.get(subject.id, {}) or {}).values())
            timetable_summary['weekly_allocated_periods'] += subject_total
            if subject_total == 0:
                timetable_summary['missing_subjects'] += 1
        timetable_summary['remaining_periods'] = 36 - timetable_summary['weekly_allocated_periods']
        timetable_summary['is_over_limit'] = timetable_summary['remaining_periods'] < 0

    if request.method == 'POST' and selected_group_id:
        selected_group = ClassGroup.objects.get(id=selected_group_id)

        for subject in subjects:
            subject_name = subject.name
            for day in days:
                key = f'periods_{subject.id}_{day}'
                value = int(request.POST.get(key) or 0)

                matching_subjects = Subject.objects.filter(
                    name=subject_name,
                    class_fk__group=selected_group
                )
                for subj in matching_subjects:
                    SubjectPeriod.objects.update_or_create(
                        subject=subj,
                        group=selected_group,
                        day=day,
                        defaults={'periods': value}
                    )

        return redirect(f'{request.path}?group={selected_group_id}')

    return render(request, 'admin_panel/timetable.html', {
        'groups': groups,
        'subjects': subjects,
        'days': days,
        'selected_group_id': int(selected_group_id) if selected_group_id else None,
        'saved_periods': saved_periods,
        'timetable_summary': timetable_summary
    })


# -------------------- Assign Period --------------------
@permission_required('admin_panel.add_assignedperiod', raise_exception=True)
def assign_period_view(request):
    form = AssignedPeriodForm()
    classes = Class.objects.all()

    if request.method == 'POST':
        form = AssignedPeriodForm(request.POST)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        else:
            print("Form errors:", form.errors)
            return JsonResponse({
                'success': False,
                'error': form.errors.get_json_data()
            }, status=400)

    return render(request, 'admin_panel/assign_period.html', {
        'form': form,
        'classes': classes
    })


# -------------------- AJAX Endpoints --------------------
from .models import (
    Subject, SubjectPeriod, Section, Class, ClassGroup,
    AssignedPeriod, CreatePeriod
)


def get_teachers_for_subject(request):
    subject_id = request.GET.get('subject_id')
    teachers = Teacher.objects.filter(subjects__id=subject_id).values('id', 'name')
    return JsonResponse(list(teachers), safe=False)


def filter_subject_periods(request):
    subject_id = request.GET.get('subject_id')
    if not subject_id:
        return JsonResponse({'error': 'No subject provided'}, status=400)

    assigned_days = AssignedPeriod.objects.filter(subject_id=subject_id).values_list('day', flat=True).distinct()
    teacher_ids = AssignedPeriod.objects.filter(subject_id=subject_id).values_list('teacher_id', flat=True).distinct()
    teachers = list(Teacher.objects.filter(id__in=teacher_ids).values('id', 'name'))

    return JsonResponse({'days': list(assigned_days), 'teachers': teachers})


def get_time_slots_for_subject_day(request):
    subject_id = request.GET.get('subject_id')
    day = request.GET.get('day')

    if not subject_id or not day:
        return JsonResponse({'error': 'Missing subject or day'}, status=400)

    assigned_periods = AssignedPeriod.objects.filter(subject_id=subject_id, day=day)
    period_ids = assigned_periods.values_list('period_id', flat=True)
    periods = CreatePeriod.objects.filter(id__in=period_ids).order_by('start_time')

    period_data = [
        {
            'id': p.id,
            'label': f"{p.period_name} ({p.start_time.strftime('%I:%M %p')} - {p.end_time.strftime('%I:%M %p')})"
        }
        for p in periods
    ]
    return JsonResponse({'periods': period_data})


def get_days_for_subject(request):
    subject_id = request.GET.get('subject_id')
    section_id = request.GET.get('section_id')

    if subject_id and section_id:
        try:
            section = Section.objects.get(id=section_id)
            group = section.class_fk.group
            subject = Subject.objects.get(id=subject_id)
            class_ids = Class.objects.filter(group=group).values_list('id', flat=True)
            subject_ids = Subject.objects.filter(class_fk_id__in=class_ids, name=subject.name).values_list('id', flat=True)

            subject_periods = SubjectPeriod.objects.filter(
                subject_id__in=subject_ids,
                group=group,
                periods__gt=0
            )
            days = subject_periods.values_list('day', 'periods').distinct()
            result = [{'day': day, 'label': f"{day} ({periods})"} for day, periods in days]
            return JsonResponse(result, safe=False)
        except Section.DoesNotExist:
            return JsonResponse([], safe=False)
    return JsonResponse([], safe=False)


def ajax_subject_periods(request):
    subject_id = request.GET.get('subject_id')
    try:
        subject = Subject.objects.get(id=subject_id)
        teachers = subject.teachers.all()
        data = {"teachers": [{"id": t.id, "name": t.name} for t in teachers]}
    except Subject.DoesNotExist:
        data = {"teachers": []}
    return JsonResponse(data)


def ajax_time_slots(request):
    subject_id = request.GET.get('subject_id')
    day = request.GET.get('day')
    section_id = request.GET.get('section_id')
    print("Incoming Data: ", subject_id, day, section_id)

    if not subject_id or not day or not section_id:
        return JsonResponse({'error': 'Missing data'}, status=400)

    try:
        section = Section.objects.get(id=section_id)
        group = section.class_fk.group
        subject = Subject.objects.get(id=subject_id)
        class_ids = Class.objects.filter(group=group).values_list('id', flat=True)
        subject_ids = Subject.objects.filter(
            class_fk_id__in=class_ids,
            name=subject.name
        ).values_list('id', flat=True)

        subject_period = SubjectPeriod.objects.filter(
            subject_id__in=subject_ids,
            group=group,
            day=day,
            periods__gt=0
        ).first()

        all_periods = CreatePeriod.objects.filter(day=day).order_by('start_time')

        data = {
            "assignable": subject_period.periods if subject_period else 0,
            "periods": [
                {
                    "id": p.id,
                    "label": f"{p.period_name} ({p.start_time.strftime('%I:%M %p')} - {p.end_time.strftime('%I:%M %p')})"
                } for p in all_periods
            ]
        }
        return JsonResponse(data)

    except (Section.DoesNotExist, Subject.DoesNotExist):
        return JsonResponse({'error': 'Invalid subject or section'}, status=400)


def get_subjects_for_section(request):
    section_id = request.GET.get('section_id')
    if not section_id:
        return JsonResponse({'error': 'Missing section id'}, status=400)
    section = get_object_or_404(Section, id=section_id)
    subjects = Subject.objects.filter(class_fk=section.class_fk).values('id', 'name')
    return JsonResponse(list(subjects), safe=False)


def get_sections_for_class(request):
    class_id = request.GET.get('class_id')
    if not class_id:
        return JsonResponse({'error': 'Missing class_id'}, status=400)
    sections = Section.objects.filter(class_fk_id=class_id).values('id', 'section_name')
    data = [{'id': s['id'], 'name': s['section_name']} for s in sections]
    return JsonResponse(data, safe=False)


def get_periods_for_day(request):
    subject_id = request.GET.get('subject_id')
    day = request.GET.get('day')
    section_id = request.GET.get('section_id')

    if subject_id and day and section_id:
        periods = SubjectPeriod.objects.filter(
            subject_id=subject_id, section_id=section_id, day=day
        ).values('period_id', 'period__label')
        period_list = [{'id': p['period_id'], 'label': p['period__label']} for p in periods]
        return JsonResponse({'periods': period_list})

    return JsonResponse({'periods': []})


# -------------------- Class Group --------------------
@permission_required('admin_panel.view_classgroup', raise_exception=True)
def add_class_group(request):
    if request.method == 'POST':
        name = request.POST.get('group_name')
        ClassGroup.objects.create(group_name=name)
        return redirect('class_group_list')
    return render(request, 'admin_panel/add_group.html')


@permission_required('admin_panel.view_classgroup', raise_exception=True)
def class_group_list(request):
    groups = ClassGroup.objects.all()
    return render(request, 'admin_panel/class_group_list.html', {'groups': groups})


@require_http_methods(["DELETE"])
@csrf_exempt
def delete_assignment(request, assignment_id):
    try:
        assignment = AssignedPeriod.objects.get(id=assignment_id)
        assignment.delete()
        return JsonResponse({'success': True})
    except AssignedPeriod.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Assignment not found'}, status=404)


def get_assigned_classes(request):
    assignments = AssignedPeriod.objects.select_related('class_fk', 'section', 'period', 'subject', 'teacher').all()
    data = []
    for a in assignments:
        data.append({
            'id': a.id,
            'class_name': a.class_fk.class_name if a.class_fk else 'N/A',
            'section_name': a.section.section_name if a.section else 'N/A',
            'subject_name': a.subject.name if a.subject else 'N/A',
            'day_label': a.get_day_display() if hasattr(a, 'get_day_display') else a.day,
            'period_label': f"{a.period.period_name} ({a.period.start_time.strftime('%I:%M %p')} - {a.period.end_time.strftime('%I:%M %p')})" if a.period else 'N/A',
            'teacher_name': a.teacher.name if a.teacher else 'N/A',
        })
    return JsonResponse(data, safe=False)


def edit_class_group(request, group_id):
    group = get_object_or_404(ClassGroup, id=group_id)
    if request.method == 'POST':
        name = request.POST.get('group_name')
        group.group_name = name
        group.save()
        return redirect('class_group_list')
    return render(request, 'admin_panel/edit_group.html', {'group': group})


def delete_class_group(request, group_id):
    group = get_object_or_404(ClassGroup, id=group_id)
    if request.method == 'POST':
        group.delete()
        return redirect('class_group_list')
    return render(request, 'admin_panel/delete_class_group.html', {'class_group': group})


# -------------------- Class Teacher --------------------
from admin_panel.models import ClassTeacher, AcademicYear, Class, Section
from teacher_dashboard.models import Teacher


@permission_required('admin_panel.view_classteacher', raise_exception=True)
def class_teacher_list(request):
    assignments = ClassTeacher.objects.all()
    return render(request, 'admin_panel/class_teacher_list.html', {'assignments': assignments})


def class_teacher_create(request):
    classes = Class.objects.all()
    academic_years = AcademicYear.objects.all()

    if request.method == 'POST':
        academic_year = request.POST.get('academic_year')
        class_id = request.POST.get('class_fk')
        section_id = request.POST.get('section')
        teacher_id = request.POST.get('teacher')

        if academic_year and class_id and section_id and teacher_id:
            existing = ClassTeacher.objects.filter(
                academic_year_id=academic_year,
                class_fk_id=class_id,
                section_id=section_id,
                teacher_id=teacher_id
            ).first()

            if existing:
                teacher = Teacher.objects.get(id=teacher_id)
                return render(request, 'admin_panel/class_teacher_form.html', {
                    'classes': classes,
                    'academic_years': academic_years,
                    'error': f"❌ Teacher '{teacher.name}' is already assigned to this class and section."
                })

            ClassTeacher.objects.create(
                academic_year_id=academic_year,
                class_fk_id=class_id,
                section_id=section_id,
                teacher_id=teacher_id
            )
            return redirect('class_teacher_list')
        else:
            return render(request, 'admin_panel/class_teacher_form.html', {
                'classes': classes,
                'academic_years': academic_years,
                'error': 'Please fill all required fields.'
            })

    return render(request, 'admin_panel/class_teacher_form.html', {
        'classes': classes,
        'academic_years': academic_years
    })


def class_teacher_update(request, pk):
    assignment = get_object_or_404(ClassTeacher, pk=pk)
    academic_years = AcademicYear.objects.all()
    classes = Class.objects.all()
    sections = Section.objects.all()
    teachers = Teacher.objects.all()

    if request.method == 'POST':
        assignment.academic_year_id = request.POST.get('academic_year')
        assignment.class_fk_id = request.POST.get('class_fk')
        assignment.section_id = request.POST.get('section')
        assignment.teacher_id = request.POST.get('teacher')
        assignment.save()
        return redirect('class_teacher_list')

    context = {
        'assignment': assignment,
        'academic_years': academic_years,
        'classes': classes,
        'sections': sections,
        'teachers': teachers,
    }
    return render(request, 'admin_panel/class_teacher_form.html', context)


def class_teacher_delete(request, pk):
    assignment = get_object_or_404(ClassTeacher, pk=pk)
    if request.method == 'POST':
        assignment.delete()
        return redirect('class_teacher_list')
    return render(request, 'admin_panel/class_teacher_confirm_delete.html', {'assignment': assignment})


def ajax_load_sections(request):
    class_id = request.GET.get('class_id')
    if class_id:
        sections = Section.objects.filter(class_fk_id=class_id).values('id', 'section_name')
        return JsonResponse({'sections': list(sections)})
    return JsonResponse({'sections': []})


def ajax_load_teachers_by_class(request):
    class_id = request.GET.get('class_id')
    section_id = request.GET.get('section_id')

    assigned = AssignedPeriod.objects.filter(
        class_fk_id=class_id,
        section_id=section_id
    ).select_related('teacher')

    unique_teachers = {a.teacher for a in assigned}
    data = [{'id': t.id, 'name': t.name} for t in unique_teachers]
    return JsonResponse({'teachers': data})


# ==================== PORTFOLIO / TIMETABLE PDF ====================
from django.template.loader import render_to_string
from django.http import HttpResponse, JsonResponse


def _get_portfolio_timetable_data(class_obj, section_obj):
    from admin_panel.models import ClassTeacher

    all_periods = CreatePeriod.objects.all().order_by('day', 'start_time')
    monday_periods = all_periods.filter(day='Monday')
    period_order = [p.period_name for p in monday_periods]
    for p in all_periods:
        if p.period_name not in period_order:
            period_order.append(p.period_name)

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    teacher_obj = ClassTeacher.objects.filter(class_fk=class_obj, section=section_obj).first()
    teacher_name = teacher_obj.teacher.name if teacher_obj else None

    slots = {}
    for day in days:
        assigned_periods = AssignedPeriod.objects.filter(
            class_fk=class_obj, section=section_obj, day=day
        ).select_related('period', 'subject', 'teacher')
        assigned_map = {ap.period.period_name: ap for ap in assigned_periods}
        day_slots = {}
        for pname in period_order:
            ap = assigned_map.get(pname)
            if ap:
                day_slots[pname] = f"{ap.subject.name} ({ap.teacher.name})<br><small>{ap.period.start_time.strftime('%I:%M %p')} - {ap.period.end_time.strftime('%I:%M %p')}</small>"
            else:
                day_slots[pname] = "---"
        slots[day] = day_slots

    return period_order, slots, days, teacher_name


def get_sections(request):
    class_id = request.GET.get('class_id')
    if class_id:
        sections = Section.objects.filter(class_fk_id=class_id).values('id', 'section_name')
        return JsonResponse({'sections': list(sections)})
    return JsonResponse({'sections': []})


def extract_period_number(period_name):
    try:
        left = period_name.split(':')[0]
        parts = left.strip().split()
        for part in parts:
            digits = ''.join(filter(str.isdigit, part))
            if digits:
                return int(digits)
    except Exception:
        pass
    return 9999


@permission_required('admin_panel.view_assignedperiod', raise_exception=True)
def portfolio_timetable(request):
    class_id = request.GET.get('class_id')
    section_id = request.GET.get('section_id')

    classes = Class.objects.all()
    timetable_data = None
    class_info = section_info = teacher_name = None
    timetable_summary = {
        'class_name': '-',
        'section_name': '-',
        'class_teacher': 'Not Assigned',
        'assigned_periods': 0,
        'empty_periods': 0,
    }

    if class_id and section_id:
        class_obj = get_object_or_404(Class, id=class_id)
        section_obj = get_object_or_404(Section, id=section_id)

        period_order, slots, days, teacher_name = _get_portfolio_timetable_data(class_obj, section_obj)

        timetable_data = {
            'period_order': period_order,
            'slots': slots,
            'days': days
        }
        class_info = class_obj
        section_info = section_obj
        assigned_periods = AssignedPeriod.objects.filter(class_fk=class_obj, section=section_obj).count()
        total_slots = len(period_order) * len(days)
        timetable_summary = {
            'class_name': class_obj.class_name,
            'section_name': section_obj.section_name,
            'class_teacher': teacher_name or 'Not Assigned',
            'assigned_periods': assigned_periods,
            'empty_periods': max(total_slots - assigned_periods, 0),
        }

    return render(request, 'admin_panel/portfolio_timetable.html', {
        'classes': classes,
        'timetable_data': timetable_data,
        'class_info': class_info,
        'section_info': section_info,
        'teacher_name': teacher_name,
        'timetable_summary': timetable_summary
    })


@permission_required('admin_panel.view_assignedperiod', raise_exception=True)
def timetable_pdf(request):
    class_id = request.GET.get('class_id')
    section_id = request.GET.get('section_id')

    if not class_id or not section_id:
        return HttpResponse("Class or Section not provided", status=400)

    class_obj = get_object_or_404(Class, id=class_id)
    section_obj = get_object_or_404(Section, id=section_id)

    period_order, slots, days, teacher_name = _get_portfolio_timetable_data(class_obj, section_obj)

    html_string = render_to_string('admin_panel/timetable_pdf.html', {
        'class_info': class_obj,
        'section_info': section_obj,
        'period_order': period_order,
        'slots': slots,
        'days': days,
        'teacher_name': teacher_name
    })

    try:
        from weasyprint import HTML as WeasyHTML
        pdf = WeasyHTML(string=html_string).write_pdf()
    except OSError as e:
        return HttpResponse(
            f"PDF generation failed: WeasyPrint libraries not installed. {e}",
            status=500
        )

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timetable_{class_obj.class_name}_{section_obj.section_name}.pdf"'
    return response


# ==================== ID CARD ====================
from student_profile.models import Student


@permission_required('student_profile.view_studentidcard', raise_exception=True)
def generate_student_id_card(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    return render(request, 'admin_panel/id_card.html', {'student': student})


@permission_required('admin_panel.view_admission', raise_exception=True)
def student_id_card_list(request):
    students = Student.objects.all()
    return render(request, 'admin_panel/student_id_card_list.html', {'students': students})


def upload_student_photo(request, student_id):
    if request.method == "POST" and request.FILES.get("photo"):
        student = get_object_or_404(Student, id=student_id)
        student.photo = request.FILES["photo"]
        student.save()
    return redirect('student_id_card_list')


# ==================== STREAM ====================
from .models import Stream


@permission_required('admin_panel.view_stream', raise_exception=True)
def stream_list(request):
    streams = Stream.objects.all()
    return render(request, "admin_panel/stream_list.html", {"streams": streams})


@permission_required('admin_panel.add_stream', raise_exception=True)
def add_stream(request):
    if request.method == "POST":
        name = request.POST.get("stream_name")
        if name:
            Stream.objects.create(stream_name=name)
            return redirect("stream_list")
    return render(request, "admin_panel/stream_form.html")


@permission_required('admin_panel.change_stream', raise_exception=True)
def edit_stream(request, pk):
    stream = get_object_or_404(Stream, pk=pk)
    if request.method == "POST":
        stream.stream_name = request.POST.get("stream_name")
        stream.save()
        return redirect("stream_list")
    return render(request, "admin_panel/stream_form.html", {"stream": stream})


@permission_required('admin_panel.delete_stream', raise_exception=True)
def delete_stream(request, pk):
    stream = get_object_or_404(Stream, pk=pk)
    stream.delete()
    return redirect("stream_list")


def my_view(request):
    return HttpResponse("Hello, this is my_view!")


# ==================== EXAMINATION ====================
from .models import ExamFormat, Question


@login_required
@permission_required('admin_panel.view_examformat', raise_exception=True)
def format_list(request):
    classes = Class.objects.all()
    subjects = Subject.objects.all()
    class_id = request.GET.get("class_id")
    subject_id = request.GET.get("subject_id")
    formats = ExamFormat.objects.select_related("class_obj", "subject").all()
    if class_id:
        formats = formats.filter(class_obj_id=class_id)
    if subject_id:
        formats = formats.filter(subject_id=subject_id)
    return render(request, "admin_panel/format_list.html", {
        "formats": formats, "classes": classes, "subjects": subjects,
    })


@login_required
@permission_required('admin_panel.change_examformat', raise_exception=True)
def edit_format(request, format_id):
    exam_format = get_object_or_404(ExamFormat, id=format_id)
    if request.method == "POST":
        exam_format.format_name = request.POST.get("format_name")
        exam_format.num_mcqs = request.POST.get("num_mcqs")
        exam_format.num_short = request.POST.get("num_short")
        exam_format.num_long = request.POST.get("num_long")
        exam_format.class_obj_id = request.POST.get("class")
        exam_format.subject_id = request.POST.get("subject")
        exam_format.save()
        messages.success(request, "Exam format updated successfully!")
        return redirect("format_list")
    classes = Class.objects.all()
    subjects = Subject.objects.all()
    return render(request, "admin_panel/edit_format.html", {
        "exam_format": exam_format, "classes": classes, "subjects": subjects,
    })


@login_required
@permission_required('admin_panel.delete_examformat', raise_exception=True)
def delete_format(request, format_id):
    exam_format = get_object_or_404(ExamFormat, id=format_id)
    exam_format.delete()
    messages.success(request, "Exam format deleted successfully!")
    return redirect("format_list")


@permission_required('admin_panel.add_examformat', raise_exception=True)
def create_format(request):
    if request.method == "POST":
        ExamFormat.objects.create(
            class_obj_id=request.POST.get("class"),
            subject_id=request.POST.get("subject"),
            format_name=request.POST.get("format_name"),
            num_mcqs=request.POST.get("num_mcqs"),
            num_short=request.POST.get("num_short"),
            num_long=request.POST.get("num_long"),
        )
        return redirect("format_list")
    classes = Class.objects.all()
    subjects = Subject.objects.all()
    return render(request, "admin_panel/create_format.html", {"classes": classes, "subjects": subjects})


@login_required
@permission_required('admin_panel.view_examformat', raise_exception=True)
def format_questions(request, format_id):
    exam_format = get_object_or_404(
        ExamFormat.objects.select_related("class_obj", "subject"), id=format_id
    )
    questions = Question.objects.select_related("teacher", "section").filter(exam_format=exam_format)
    return render(request, "admin_panel/format_questions.html", {
        "exam_format": exam_format, "questions": questions,
    })


@permission_required('admin_panel.view_question', raise_exception=True)
def all_questions(request):
    questions = Question.objects.select_related("exam_format", "teacher", "section")
    return render(request, "admin_panel/all_questions.html", {"questions": questions})


# ==================== PAPER GENERATION ====================
import random
from math import ceil
from django.db.models import Count
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


@login_required
@permission_required('admin_panel.view_examformat', raise_exception=True)
def generate_paper_confirm(request, format_id):
    exam_format = get_object_or_404(
        ExamFormat.objects.select_related("class_obj", "subject"), id=format_id
    )
    if request.method == "POST":
        return generate_question_paper_pdf(request, format_id)
    return render(request, "admin_panel/generate_paper_confirm.html", {"exam_format": exam_format})


def _pick_most_common_questions(queryset, needed):
    if needed <= 0:
        return []
    grouped = (
        queryset.values("text")
        .annotate(cnt=Count("id"))
        .order_by("-cnt")
    )
    selected = []
    used_texts = set()
    for g in grouped:
        if len(selected) >= needed:
            break
        txt = g["text"]
        if txt in used_texts:
            continue
        q = queryset.filter(text=txt).first()
        if q:
            selected.append(q)
            used_texts.add(txt)
    return selected


def _pick_random_questions(queryset, needed, exclude_ids=None):
    if needed <= 0:
        return []
    qs = queryset
    if exclude_ids:
        qs = qs.exclude(id__in=exclude_ids)
    total = qs.count()
    if total == 0:
        return []
    if total <= needed:
        return list(qs)
    ids = list(qs.values_list("id", flat=True))
    sampled_ids = random.sample(ids, needed)
    return list(qs.filter(id__in=sampled_ids))


def generate_question_paper_pdf(request, format_id):
    exam_format = get_object_or_404(
        ExamFormat.objects.select_related("class_obj", "subject"), id=format_id
    )
    all_questions = Question.objects.filter(exam_format=exam_format)
    mcq_needed = getattr(exam_format, "num_mcqs", 0) or 0
    short_needed = getattr(exam_format, "num_short", 0) or 0
    long_needed = getattr(exam_format, "num_long", 0) or 0

    selected_questions = []

    mcq_qs = all_questions.filter(question_type__iexact="MCQ")
    mcq_most_common = _pick_most_common_questions(mcq_qs, ceil(mcq_needed * 0.5))
    selected_ids = [q.id for q in mcq_most_common]
    mcq_random = _pick_random_questions(mcq_qs, mcq_needed - len(mcq_most_common), exclude_ids=selected_ids)
    selected_questions.extend(mcq_most_common)
    selected_questions.extend(mcq_random)

    short_qs = all_questions.filter(question_type__iexact="SHORT")
    short_most_common = _pick_most_common_questions(short_qs, ceil(short_needed * 0.5))
    selected_ids = selected_ids + [q.id for q in short_most_common]
    short_random = _pick_random_questions(short_qs, short_needed - len(short_most_common), exclude_ids=selected_ids)
    selected_questions.extend(short_most_common)
    selected_questions.extend(short_random)

    long_qs = all_questions.filter(question_type__iexact="LONG")
    long_most_common = _pick_most_common_questions(long_qs, ceil(long_needed * 0.5))
    selected_ids = selected_ids + [q.id for q in long_most_common]
    long_random = _pick_random_questions(long_qs, long_needed - len(long_most_common), exclude_ids=selected_ids)
    selected_questions.extend(long_most_common)
    selected_questions.extend(long_random)

    response = HttpResponse(content_type="application/pdf")
    filename = f"Exam_Paper_{exam_format.format_name.replace(' ', '_')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 60

    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width / 2, y, f"Exam Paper — {exam_format.format_name}")
    y -= 30
    p.setFont("Helvetica", 12)
    p.drawCentredString(width / 2, y, f"Class: {exam_format.class_obj.class_name}    Subject: {exam_format.subject.name}")
    y -= 40

    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y, "Section A: MCQs")
    y -= 20
    p.setFont("Helvetica", 11)
    idx = 1
    for q in [qq for qq in selected_questions if qq.question_type.upper() == "MCQ"]:
        if y < 100:
            p.showPage()
            y = height - 60
        p.drawString(50, y, f"{idx}. {q.text}")
        y -= 16
        for attr, letter in [("option_a", "A"), ("option_b", "B"), ("option_c", "C"), ("option_d", "D")]:
            if hasattr(q, attr):
                p.drawString(70, y, f"{letter}) {getattr(q, attr)}")
                y -= 14
        y -= 8
        idx += 1

    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y, "Section B: Short Questions")
    y -= 20
    p.setFont("Helvetica", 11)
    idx = 1
    for q in [qq for qq in selected_questions if qq.question_type.upper() == "SHORT"]:
        if y < 100:
            p.showPage()
            y = height - 60
        p.drawString(50, y, f"{idx}. {q.text}")
        y -= 30
        idx += 1

    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y, "Section C: Long Questions")
    y -= 20
    p.setFont("Helvetica", 11)
    idx = 1
    for q in [qq for qq in selected_questions if qq.question_type.upper() == "LONG"]:
        if y < 100:
            p.showPage()
            y = height - 60
        p.drawString(50, y, f"{idx}. {q.text}")
        y -= 50
        idx += 1

    p.save()
    return response


def generate_question_paper(request, format_id=None):
    if format_id:
        exam_format = get_object_or_404(ExamFormat, id=format_id)
        context = {"format": exam_format}
    else:
        context = {"format": None}
    return render(request, "admin_panel/generate_question_paper.html", context)


# ==================== BOOK UPLOAD ====================
from .models import Book
from .forms import BookUploadForm
from .ml_toc_parser import parse_book_toc_ml


@permission_required('admin_panel.add_book', raise_exception=True)
def upload_book(request):
    if request.method == "POST":
        form = BookUploadForm(request.POST, request.FILES)
        if form.is_valid():
            book = form.save(commit=False)
            book.uploaded_by = request.user
            book.save()
            try:
                parse_book_toc_ml(book)
                messages.success(request, f"Book '{book.title}' uploaded and TOC parsed successfully!")
            except Exception as e:
                messages.warning(request, f"Book uploaded but TOC parsing failed: {e}")
            return redirect('upload_book')
        else:
            messages.error(request, "Form invalid. Please check the uploaded file.")
    else:
        form = BookUploadForm()

    books = Book.objects.all().order_by('-uploaded_at')
    return render(request, 'admin_panel/upload_book.html', {'form': form, 'books': books})


def book_detail(request, pk):
    book = get_object_or_404(Book.objects.prefetch_related('chapters__topics__subtopics'), pk=pk)
    return render(request, 'admin_panel/book_detail.html', {'book': book})


def book_list(request):
    books = Book.objects.all().order_by('-id')
    return render(request, "admin_panel/book_list.html", {"books": books})


# ==================== TEACHER DATA ====================
from teacher_dashboard.models import AssignedPeriod as TeacherAssignedPeriod


@login_required
def get_teacher_data(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    periods = TeacherAssignedPeriod.objects.filter(teacher=teacher).select_related(
        'class_fk', 'section', 'subject'
    )
    data = []
    for p in periods:
        data.append({
            'class_id': p.class_fk.id,
            'class_name': p.class_fk.class_name,
            'section_id': p.section.id,
            'section_name': p.section.section_name,
            'subject_id': p.subject.id,
            'subject_name': p.subject.name
        })
    return JsonResponse({'assigned': data})


# ==================== FIXTURE MANAGEMENT ====================
from .models import TeacherFixture


@permission_required('admin_panel.add_teacherfixture', raise_exception=True)
def manage_fixture(request):
    classes = Class.objects.all()
    sections = Section.objects.all()
    periods = CreatePeriod.objects.all()
    subjects = Subject.objects.all()
    teachers = Teacher.objects.filter(status='active')
    fixtures = TeacherFixture.objects.select_related(
        "class_fk", "section", "period",
        "absent_teacher", "substitute_teacher", "subject"
    ).order_by("-id")
    today = date.today()
    fixture_summary = {
        'total': fixtures.count(),
        'today': TeacherFixture.objects.filter(fixture_date=today).count(),
        'workload_limit': 3,
        'active_teachers': teachers.count(),
    }

    if request.method == 'POST':
        class_id = request.POST.get("class_fk")
        section_id = request.POST.get("section")
        subject_id = request.POST.get("subject")
        period_id = request.POST.get("period")
        day = request.POST.get("day")
        fixture_date = request.POST.get("fixture_date") or today
        absent_teacher_id = request.POST.get("absent_teacher")
        substitute_teacher_id = request.POST.get("substitute_teacher")

        if TeacherFixture.objects.filter(
            class_fk_id=class_id, section_id=section_id,
            period_id=period_id, day=day, fixture_date=fixture_date
        ).exists():
            messages.error(request, "❌ Fixture already assigned for this class/period.")
            return redirect("manage_fixture")

        if subject_id:
            subject_ok = Teacher.objects.filter(id=substitute_teacher_id, subjects__id=subject_id).exists()
            if not subject_ok:
                messages.error(request, "❌ Substitute teacher cannot teach this subject.")
                return redirect("manage_fixture")

        if TeacherFixture.objects.filter(substitute_teacher_id=substitute_teacher_id, fixture_date=fixture_date).count() >= 3:
            messages.error(request, "❌ Teacher workload limit reached for today.")
            return redirect("manage_fixture")

        TeacherFixture.objects.create(
            absent_teacher_id=absent_teacher_id,
            substitute_teacher_id=substitute_teacher_id,
            class_fk_id=class_id,
            section_id=section_id,
            subject_id=subject_id or None,
            period_id=period_id,
            day=day,
            fixture_date=fixture_date,
            source='manual',
            automation_status='assigned'
        )
        messages.success(request, "✅ Fixture Assigned Successfully!")
        return redirect("manage_fixture")

    return render(request, "admin_panel/manage_fixture.html", {
        'classes': classes, 'sections': sections, 'periods': periods,
        'subjects': subjects, 'teachers': teachers, 'fixtures': fixtures,
        'fixture_summary': fixture_summary,
    })


def get_free_teachers(request):
    try:
        day = request.GET.get('day')
        period_id = request.GET.get('period_id')
        subject_id = request.GET.get('subject_id')
        fixture_date = request.GET.get('fixture_date') or date.today()

        if not period_id:
            return JsonResponse({'error': 'Missing period'}, status=400)

        busy_from_timetable = AssignedPeriod.objects.filter(
            day=day, period_id=period_id
        ).values_list('teacher_id', flat=True)

        busy_from_fixtures = TeacherFixture.objects.filter(
            day=day, period_id=period_id, fixture_date=fixture_date
        ).values_list('substitute_teacher_id', flat=True)

        all_busy_ids = set(busy_from_timetable) | set(busy_from_fixtures)
        free_teachers = Teacher.objects.filter(status='active').exclude(id__in=all_busy_ids)

        if subject_id:
            free_teachers = free_teachers.filter(subjects__id=subject_id)

        free_teachers = [
            t for t in free_teachers
            if TeacherFixture.objects.filter(substitute_teacher=t, fixture_date=fixture_date).count() < 3
        ]

        free_teachers_sorted = sorted(
            free_teachers,
            key=lambda t: (
                TeacherFixture.objects.filter(substitute_teacher=t).count(),
                AssignedPeriod.objects.filter(teacher=t).count()
            )
        )

        data = [{'id': t.id, 'name': t.name} for t in free_teachers_sorted]
        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==================== HR / EMPLOYEE ====================
from .models import (Department, Designation, Employee, StaffCategory, JobType, LeaveType, LeaveApplication)


@permission_required('admin_panel.view_department', raise_exception=True)
def department_list(request):
    departments = Department.objects.all()
    return render(request, 'admin_panel/department_list.html', {'departments': departments})


@permission_required('admin_panel.add_department', raise_exception=True)
def department_create(request):
    if request.method == 'POST':
        Department.objects.create(
            name=request.POST.get('name'),
            description=request.POST.get('description', '')
        )
        return redirect('department_list')
    return render(request, 'admin_panel/department_form.html')


@permission_required('admin_panel.view_designation', raise_exception=True)
def designation_list(request):
    designations = Designation.objects.select_related('department')
    return render(request, 'admin_panel/designation_list.html', {'designations': designations})


@permission_required('admin_panel.add_designation', raise_exception=True)
def designation_create(request):
    departments = Department.objects.all()
    if request.method == 'POST':
        Designation.objects.create(
            department_id=request.POST.get('department'),
            title=request.POST.get('title')
        )
        return redirect('designation_list')
    return render(request, 'admin_panel/designation_form.html', {'departments': departments})


@permission_required('admin_panel.view_employee', raise_exception=True)
def employee_list(request):
    employees = Employee.objects.all()
    return render(request, 'admin_panel/employee_list.html', {'employees': employees})


from .forms import EmployeeForm, TeacherFromEmployeeForm


def employee_create(request):
    if request.method == "POST":
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            with transaction.atomic():
                employee = form.save()
                is_teacher = form.cleaned_data.get("is_teacher")

                if is_teacher:
                    teacher_group, _ = Group.objects.get_or_create(name="Teacher")
                    user = None
                    if employee.user:
                        user = employee.user
                    else:
                        username = slugify(employee.name) + "_" + str(uuid.uuid4())[:5]
                        user = User.objects.create_user(
                            username=username,
                            email=employee.email,
                            password="default123",
                            first_name=employee.name
                        )
                        user.groups.set([teacher_group])
                        employee.user = user
                        employee.save(update_fields=["user"])

                    teacher, _created = Teacher.objects.get_or_create(
                        email=employee.email,
                        defaults={
                            "user": user,
                            "name": employee.name,
                            "phone": employee.phone,
                            "gender": form.cleaned_data.get("teacher_gender") or "Male",
                            "qualification": form.cleaned_data.get("teacher_qualification") or "Matric",
                            "experience": form.cleaned_data.get("teacher_experience") or 0,
                            "Address": form.cleaned_data.get("teacher_address") or "",
                            "department": form.cleaned_data.get("teacher_department") or "",
                            "faculty_group": form.cleaned_data.get("teacher_faculty_group") or [],
                        }
                    )

                    teacher.user = user
                    teacher.name = employee.name
                    teacher.phone = employee.phone
                    teacher.gender = form.cleaned_data.get("teacher_gender") or teacher.gender
                    teacher.qualification = form.cleaned_data.get("teacher_qualification") or teacher.qualification
                    exp = form.cleaned_data.get("teacher_experience")
                    teacher.experience = exp if exp is not None else teacher.experience
                    teacher.Address = form.cleaned_data.get("teacher_address") or teacher.Address
                    teacher.department = form.cleaned_data.get("teacher_department") or teacher.department
                    teacher.faculty_group = form.cleaned_data.get("teacher_faculty_group") or teacher.faculty_group

                    if request.FILES.get("teacher_image"):
                        teacher.image = request.FILES.get("teacher_image")

                    teacher.save()

                    teacher_subjects = form.cleaned_data.get("teacher_subjects")
                    if teacher_subjects is not None:
                        teacher.subjects.set(teacher_subjects)

            messages.success(request, "Employee created successfully.")
            return redirect("employee_list")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = EmployeeForm()

    return render(request, "admin_panel/employee_form.html", {"form": form})


@permission_required('admin_panel.change_employee', raise_exception=True)
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'admin_panel/employee_form.html', {'form': form, 'title': 'Edit Employee'})


# ==================== LEAVE ====================
from django.utils import timezone
from datetime import date


def calculate_leave_days(start_date, end_date):
    return (end_date - start_date).days + 1


@permission_required('admin_panel.add_leaveapplication', raise_exception=True)
def leave_create(request):
    employees = Employee.objects.all()
    leave_types = LeaveType.objects.all()

    if request.method == 'POST':
        employee = get_object_or_404(Employee, id=request.POST.get('employee'))
        leave_type = get_object_or_404(LeaveType, id=request.POST.get('leave_type'))
        start_date = date.fromisoformat(request.POST.get('start_date'))
        end_date = date.fromisoformat(request.POST.get('end_date'))
        reason = request.POST.get('reason')

        if not employee.can_apply_leave():
            messages.error(request, "Employee is not eligible for leave (probation not completed).")
            return redirect('leave_create')

        if leave_type not in employee.job_type.allowed_leave_types.all():
            messages.error(request, "This leave type is not allowed for this job type.")
            return redirect('leave_create')

        overlap = LeaveApplication.objects.filter(
            employee=employee,
            start_date__lte=end_date,
            end_date__gte=start_date,
            status__in=['pending', 'approved']
        )
        if overlap.exists():
            messages.error(request, "Employee already has leave during these dates.")
            return redirect('leave_create')

        requested_days = calculate_leave_days(start_date, end_date)

        approved_leaves = LeaveApplication.objects.filter(
            employee=employee,
            leave_type=leave_type,
            status='approved',
            start_date__year=date.today().year
        )
        used_days = sum(calculate_leave_days(l.start_date, l.end_date) for l in approved_leaves)
        remaining_days = leave_type.yearly_quota - used_days

        if requested_days > remaining_days:
            messages.error(request, f"Leave quota exceeded. Remaining days: {remaining_days}")
            return redirect('leave_create')

        LeaveApplication.objects.create(
            employee=employee,
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date,
            reason=reason
        )
        messages.success(request, "Leave applied successfully.")
        return redirect('leave_list')

    return render(request, 'admin_panel/leave_form.html', {
        'employees': employees, 'leave_types': leave_types
    })


@permission_required('admin_panel.view_leaveapplication', raise_exception=True)
def leave_list(request):
    leaves = LeaveApplication.objects.select_related('employee', 'leave_type').order_by('-applied_at')
    return render(request, 'admin_panel/leave_list.html', {'leaves': leaves})


@permission_required('admin_panel.change_leaveapplication', raise_exception=True)
def leave_action(request, pk, action):
    leave = get_object_or_404(LeaveApplication, pk=pk)
    if action in ['approved', 'rejected']:
        leave.status = action
        leave.action_date = timezone.now()
        leave.save()
        if action == 'approved':
            automation_results = FixtureAutomationService.run_for_leave(leave, triggered_by=request.user)
            assigned_count = sum(len(result.get('assigned', [])) for result in automation_results)
            unassigned_count = sum(len(result.get('unassigned', [])) for result in automation_results)
            skipped_count = sum(len(result.get('skipped', [])) for result in automation_results)
            if automation_results:
                messages.success(
                    request,
                    f'Leave approved. Auto fixture result: {assigned_count} assigned, {unassigned_count} unassigned, {skipped_count} skipped.'
                )
            else:
                messages.warning(
                    request,
                    'Leave approved, but no linked teacher profile was found for fixture automation.'
                )
    return redirect('leave_list')


# ==================== STAFF CATEGORY & JOB TYPE ====================
from .forms import StaffCategoryForm, JobTypeForm


@permission_required('admin_panel.view_staffcategory', raise_exception=True)
def staff_category_list(request):
    categories = StaffCategory.objects.all()
    return render(request, 'admin_panel/staff_category_list.html', {'categories': categories})


@permission_required('admin_panel.add_staffcategory', raise_exception=True)
def staff_category_create(request):
    if request.method == 'POST':
        form = StaffCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('staff_category_list')
    else:
        form = StaffCategoryForm()
    return render(request, 'admin_panel/staff_category_form.html', {'form': form})


@permission_required('admin_panel.view_jobtype', raise_exception=True)
def job_type_list(request):
    job_types = JobType.objects.all()
    return render(request, 'admin_panel/job_type_list.html', {'job_types': job_types})


@permission_required('admin_panel.add_jobtype', raise_exception=True)
def job_type_create(request):
    if request.method == 'POST':
        form = JobTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('job_type_list')
    else:
        form = JobTypeForm()
    return render(request, 'admin_panel/job_type_form.html', {'form': form})


# ==================== LEAVE TYPE ====================
from .forms import LeaveTypeForm


@permission_required('admin_panel.view_leavetype', raise_exception=True)
def leave_type_list(request):
    leave_types = LeaveType.objects.all()
    return render(request, 'admin_panel/leave_type_list.html', {'leave_types': leave_types})


@permission_required('admin_panel.add_leavetype', raise_exception=True)
def leave_type_create(request):
    if request.method == 'POST':
        form = LeaveTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('leave_type_list')
    else:
        form = LeaveTypeForm()
    return render(request, 'admin_panel/leave_type_form.html', {'form': form})


# ==================== APPRAISAL ====================
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from admin_panel.models import AppraisalCycle, TeacherAppraisalSubmission, KpiTemplate, KpiRule
from .appraisal_services import generate_score, predict_band, train_random_forest


def is_admin(user):
    return user.is_superuser or user.is_staff or user.groups.filter(name__iexact="Admin").exists()


@login_required
@user_passes_test(is_admin)
def admin_kpi_builder(request):
    cycle = AppraisalCycle.objects.filter(is_open=True).order_by("-start_date").first()
    if not cycle:
        messages.error(request, "No open appraisal cycle found.")
        return redirect("admin_dashboard")

    template = KpiTemplate.objects.filter(cycle=cycle).order_by("-id").first()
    if not template:
        template = KpiTemplate.objects.create(name=f"KPI Template - {cycle.name}", cycle=cycle)

    auto_keys = [k for k, _ in KpiRule.AUTO_KPI_KEYS]
    for k, label in KpiRule.AUTO_KPI_KEYS:
        KpiRule.objects.get_or_create(
            template=template,
            kpi_key=k,
            defaults={
                "title": label,
                "weight": 0,
                "target_value": 0,
                "scoring_method": "linear",
                "is_active": True,
                "is_manual": False,
            },
        )

    if request.method == "POST" and request.POST.get("save_kpis") == "1":
        for k in auto_keys:
            rule = template.rules.filter(kpi_key=k, is_manual=False).first()
            if not rule:
                continue
            active = request.POST.get(f"auto_active_{k}") == "on"
            w = request.POST.get(f"auto_weight_{k}") or "0"
            try:
                rule.weight = float(w)
            except Exception:
                rule.weight = 0.0
            rule.is_active = active
            rule.save(update_fields=["weight", "is_active"])

        manual_rules = list(template.rules.filter(is_manual=True).order_by("id"))
        for r in manual_rules:
            if request.POST.get(f"manual_delete_{r.id}") == "on":
                r.delete()
                continue
            r.is_active = request.POST.get(f"manual_active_{r.id}") == "on"
            r.title = (request.POST.get(f"manual_title_{r.id}") or r.title).strip()
            w = request.POST.get(f"manual_weight_{r.id}") or str(r.weight or 0)
            try:
                r.weight = float(w)
            except Exception:
                pass
            r.save()

        new_titles = request.POST.getlist("manual_title[]")
        new_weights = request.POST.getlist("manual_weight[]")
        for idx, t in enumerate(new_titles):
            t = (t or "").strip()
            if not t:
                continue
            w = 0.0
            if idx < len(new_weights):
                try:
                    w = float(new_weights[idx] or 0)
                except Exception:
                    w = 0.0
            KpiRule.objects.create(
                template=template,
                title=t,
                weight=w,
                is_active=True,
                is_manual=True,
                scoring_method="linear",
                target_value=10,
            )

        messages.success(request, "KPI rules saved successfully.")
        return redirect("admin_kpi_builder")

    auto_kpis = template.rules.filter(is_manual=False, kpi_key__in=auto_keys).order_by("id")
    manual_rules = template.rules.filter(is_manual=True).order_by("id")

    return render(request, "admin_panel/appraisal_admin_kpis.html", {
        "cycle": cycle,
        "template": template,
        "auto_kpis": auto_kpis,
        "manual_rules": manual_rules,
    })


@login_required
@user_passes_test(is_admin)
def admin_appraisal_list(request):
    cycle = AppraisalCycle.objects.filter(is_open=True).order_by("-start_date").first()
    qs = TeacherAppraisalSubmission.objects.select_related("teacher", "cycle").order_by("-updated_at")
    if cycle:
        qs = qs.filter(cycle=cycle)
    return render(request, "admin_panel/appraisal_admin_list.html", {
        "cycle": cycle, "submissions": qs
    })


@login_required
@user_passes_test(is_admin)
def admin_appraisal_detail(request, pk):
    submission = get_object_or_404(TeacherAppraisalSubmission, pk=pk)
    generate_score(submission)
    pred = predict_band(submission)

    if request.method == "POST":

        if request.POST.get("save_final_band"):
            submission.final_band = request.POST.get("final_band", "")
            submission.save(update_fields=["final_band"])
            messages.success(request, "Final band saved.")
            return redirect("admin_appraisal_detail", pk=pk)

        if request.POST.get("train_ml"):
            labeled = TeacherAppraisalSubmission.objects.exclude(
                final_band=""
            ).exclude(auto_metrics={})

            template = submission.kpi_template

            if not template and submission.cycle:
                template = KpiTemplate.objects.filter(
                    cycle=submission.cycle
                ).order_by("-id").first()

            if not template:
                messages.error(
                    request,
                    "❌ KPI Template nahi mila. Pehle KPI Builder mein template banao."
                )
                return redirect("admin_appraisal_detail", pk=pk)

            try:
                train_random_forest(labeled, template)
                messages.success(request, "✅ RandomForest trained successfully.")
            except RuntimeError as e:
                messages.error(request, f"❌ Training failed: {str(e)}")
            except Exception as e:
                messages.error(request, f"❌ Unexpected error: {str(e)}")

            return redirect("admin_appraisal_detail", pk=pk)

        if request.POST.get("regenerate"):
            generate_score(submission)
            messages.success(request, "Score regenerated.")
            return redirect("admin_appraisal_detail", pk=pk)

    return render(request, "admin_panel/appraisal_admin_detail.html", {
        "submission": submission,
        "pred": pred,
    })


# ==================== ACADEMIC CALENDAR ====================
import json
from datetime import datetime
from django.utils.dateparse import parse_datetime
from django.utils.timezone import localtime
from .models import AcademicCalendarEvent


def _type_to_color(event_type: str) -> str:
    return {
        "gazetted": "#34a853",
        "college_assess": "#1a73e8",
        "college_event": "#7c6f31",
        "mat_camb_assess": "#8ab4f8",
        "pre_primary": "#d56a1f",
        "speaker_train": "#7b1fa2",
        "eca_cca": "#fbbc05",
        "ptm": "#ea4335",
        "other": "#5f6368",
    }.get(event_type or "other", "#1a73e8")


@login_required
@permission_required("admin_panel.view_academiccalendarevent", raise_exception=True)
def academic_calendar_page(request):
    return render(request, "admin_panel/academic_calendar.html")


@login_required
@permission_required("admin_panel.view_academiccalendarevent", raise_exception=True)
def academic_calendar_events(request):
    events = AcademicCalendarEvent.objects.all().order_by("start")
    data = []
    for e in events:
        color = e.color or _type_to_color(e.event_type)
        data.append({
            "id": e.id,
            "title": e.title,
            "start": e.start.isoformat(),
            "end": e.end.isoformat() if e.end else None,
            "allDay": e.all_day,
            "extendedProps": {
                "event_type": e.event_type,
                "description": e.description,
                "streams": e.streams,
                "responsibility": e.responsibility,
                "color": color,
            },
            "backgroundColor": color,
            "borderColor": color,
        })
    return JsonResponse(data, safe=False)


@login_required
@permission_required("admin_panel.add_academiccalendarevent", raise_exception=True)
@require_http_methods(["POST"])
def academic_calendar_create(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
        title = (payload.get("title") or "").strip()
        start = parse_datetime(payload.get("start") or "")
        end = parse_datetime(payload.get("end") or "") if payload.get("end") else None

        if not title or not start:
            return HttpResponseBadRequest("title/start required")

        event_type = payload.get("event_type") or "other"
        e = AcademicCalendarEvent.objects.create(
            title=title,
            start=start,
            end=end,
            all_day=bool(payload.get("allDay", False)),
            event_type=event_type,
            description=payload.get("description") or "",
            streams=payload.get("streams") or "",
            responsibility=payload.get("responsibility") or "",
            color=payload.get("color") or _type_to_color(event_type),
        )
        return JsonResponse({"ok": True, "id": e.id})
    except Exception as ex:
        return HttpResponseBadRequest(str(ex))


@login_required
@permission_required("admin_panel.change_academiccalendarevent", raise_exception=True)
@require_http_methods(["POST", "PATCH"])
def academic_calendar_update(request, event_id):
    try:
        e = AcademicCalendarEvent.objects.get(id=event_id)
        payload = json.loads(request.body.decode("utf-8"))

        if "title" in payload:
            e.title = (payload.get("title") or "").strip()
        if "start" in payload:
            dt = parse_datetime(payload.get("start") or "")
            if dt:
                e.start = dt
        if "end" in payload:
            e.end = parse_datetime(payload.get("end") or "") if payload.get("end") else None
        if "allDay" in payload:
            e.all_day = bool(payload.get("allDay"))
        if "event_type" in payload:
            e.event_type = payload.get("event_type") or "other"
        if "description" in payload:
            e.description = payload.get("description") or ""
        if "streams" in payload:
            e.streams = payload.get("streams") or ""
        if "responsibility" in payload:
            e.responsibility = payload.get("responsibility") or ""
        if "color" in payload:
            e.color = payload.get("color") or ""
        else:
            if not e.color:
                e.color = _type_to_color(e.event_type)

        if not e.title or not e.start:
            return HttpResponseBadRequest("title/start required")

        e.save()
        return JsonResponse({"ok": True})
    except AcademicCalendarEvent.DoesNotExist:
        return HttpResponseBadRequest("event not found")
    except Exception as ex:
        return HttpResponseBadRequest(str(ex))


@login_required
@permission_required("admin_panel.delete_academiccalendarevent", raise_exception=True)
@require_http_methods(["POST", "DELETE"])
def academic_calendar_delete(request, event_id):
    AcademicCalendarEvent.objects.filter(id=event_id).delete()
    return JsonResponse({"ok": True})


@login_required
@permission_required("admin_panel.view_academiccalendarevent", raise_exception=True)
def academic_calendar_export_pdf(request):
    year_str = request.GET.get("year")
    try:
        year = int(year_str) if year_str else datetime.now().year
    except Exception:
        year = datetime.now().year

    qs = AcademicCalendarEvent.objects.filter(start__year=year).order_by("start")

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="ACADEMIC_CALENDAR_{year}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=18, bottomMargin=18
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal = styles["BodyText"]

    story = []
    story.append(Paragraph(f"ACADEMIC CALENDAR {year}", title_style))
    story.append(Spacer(1, 10))

    legend_items = [
        ("Gazetted Holidays", _type_to_color("gazetted")),
        ("College Assess/Admin", _type_to_color("college_assess")),
        ("College Events", _type_to_color("college_event")),
        ("Mat/Camb Assess", _type_to_color("mat_camb_assess")),
        ("Pre-Primary", _type_to_color("pre_primary")),
        ("Speaker/Ses/Train", _type_to_color("speaker_train")),
        ("ECA/CCA", _type_to_color("eca_cca")),
        ("PTM", _type_to_color("ptm")),
    ]

    legend_data = []
    row = []
    for label, hexcol in legend_items:
        swatch = Table([[" "]], colWidths=10, rowHeights=10)
        swatch.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(hexcol)),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        cell = Table([[swatch, Paragraph(label, normal)]], colWidths=[14, 120])
        cell.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        row.append(cell)
        if len(row) == 3:
            legend_data.append(row)
            row = []
    if row:
        while len(row) < 3:
            row.append("")
        legend_data.append(row)

    legend_table = Table(legend_data, colWidths=[180, 180, 180])
    legend_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(legend_table)
    story.append(Spacer(1, 10))

    header = ["S.No", "Date", "Day", "Event", "Streams", "Responsibility"]
    table_data = [header]

    for idx, e in enumerate(qs, start=1):
        dt = localtime(e.start)
        table_data.append([
            str(idx),
            dt.strftime("%d-%m-%y"),
            dt.strftime("%A"),
            Paragraph(e.title or "", normal),
            Paragraph(e.streams or "", normal),
            Paragraph(e.responsibility or "", normal),
        ])

    if len(table_data) == 1:
        table_data.append(["", "", "", Paragraph("No events found for this year.", normal), "", ""])

    col_widths = [40, 70, 90, 280, 220, 220]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eeeeee")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cfcfcf")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]

    for row_i, e in enumerate(qs, start=1):
        col = colors.HexColor(e.color or _type_to_color(e.event_type))
        style_cmds.append(("BACKGROUND", (0, row_i), (-1, row_i), col))
        style_cmds.append(("TEXTCOLOR", (0, row_i), (-1, row_i), colors.white))

    t.setStyle(TableStyle(style_cmds))
    story.append(t)

    doc.build(story)
    return response


# ==================== BULK UPLOAD ====================
import openpyxl
import uuid
from datetime import datetime
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.models import Group, User
from django.contrib import messages
from django.shortcuts import redirect, render
from django.db import transaction
from django.utils.text import slugify


DEFAULT_TEACHER_PASSWORD = "Teacher@123"


def _parse_date(val):
    if val is None:
        return None
    if hasattr(val, 'date') and callable(val.date):
        return val.date()
    if hasattr(val, 'year'):
        return val
    val = str(val).strip()
    if not val or val.lower() in ('none', 'nan', ''):
        return None
    for fmt in ('%d-%b-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d %b %Y'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_experience(val):
    if val is None:
        return 0
    val = str(val).strip().lower()
    val = val.replace('years', '').replace('year', '').strip()
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _str(val):
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.lower() in ('none', 'nan') else s


@permission_required('admin_panel.add_admission', raise_exception=True)
def bulk_upload_students(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        except Exception as e:
            messages.error(request, f"❌ Excel file padhne mein masla: {e}")
            return redirect('bulk_upload_students')

        ws = wb.active
        active_year = AcademicYear.objects.filter(is_active=True).first()

        if not active_year:
            messages.error(request, "❌ Koi active academic year nahi mila. Pehle academic year set karein.")
            return redirect('bulk_upload_students')

        success_count = 0
        skip_count    = 0
        error_rows    = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            if not any(row):
                continue

            try:
                student_id     = _str(row[1])
                campus         = _str(row[2])
                branch         = _str(row[3])
                name           = _str(row[4])
                dob            = _parse_date(row[5])
                gender         = _str(row[6]) or 'Male'
                email          = _str(row[7])
                contact        = _str(row[8])
                address        = _str(row[9])
                admission_date = _parse_date(row[10])
                father_name    = _str(row[11])
                father_email   = _str(row[12])
                class_name     = _str(row[14])
                father_occ     = _str(row[15])
                mother_name    = _str(row[16])
                father_contact = _str(row[17])
                father_cnic    = _str(row[18])
                nationality    = _str(row[19])
                raw_status     = _str(row[20]).lower() or 'pending'
                login_id       = _str(row[21])
                password       = _str(row[22])

                if not name or not login_id or not password:
                    skip_count += 1
                    error_rows.append(
                        f"Row {row_num}: Student Name, Login_Id, and Password are required."
                    )
                    continue

                if User.objects.filter(username__iexact=login_id).exists():
                    skip_count += 1
                    error_rows.append(
                        f"Row {row_num}: Login ID '{login_id}' pehle se exist karta hai — skip kiya."
                    )
                    continue
                try:
                    validate_password(password)
                except Exception as exc:
                    skip_count += 1
                    error_rows.append(
                        f"Row {row_num}: Password invalid — {' '.join(exc.messages) if hasattr(exc, 'messages') else str(exc)}"
                    )
                    continue

                status_map = {
                    'active':   'approved',
                    'inactive': 'rejected',
                    'pending':  'pending',
                    'approved': 'approved',
                    'rejected': 'rejected',
                }
                admission_status = status_map.get(raw_status, 'pending')

                if student_id and Admission.objects.filter(student_id=student_id).exists():
                    skip_count += 1
                    error_rows.append(
                        f"Row {row_num}: Student ID '{student_id}' pehle se exist karta hai — skip kiya."
                    )
                    continue

                class_obj = Class.objects.filter(class_name__iexact=class_name).first() if class_name else None

                if admission_status == 'approved':
                    student_email = email or f"{login_id}@edupilot.local"
                    if Student.objects.filter(email__iexact=student_email).exists():
                        skip_count += 1
                        error_rows.append(
                            f"Row {row_num}: Student email '{student_email}' pehle se exist karta hai — skip kiya."
                        )
                        continue

                admission = Admission.objects.create(
                    student_id       = student_id or None,
                    campus           = campus,
                    branch           = branch,
                    name             = name,
                    dob              = dob,
                    gender           = gender,
                    email            = email,
                    contact          = contact,
                    address          = address,
                    admission_date   = admission_date,
                    father_name      = father_name,
                    father_email     = father_email,
                    mother_name      = mother_name,
                    father_contact   = father_contact,
                    father_cnic      = father_cnic,
                    father_occupation= father_occ,
                    academic_year    = active_year,
                    class_fk         = class_obj,
                    admission_status = admission_status,
                    nationality      = nationality,
                )

                if admission_status == 'approved':
                    with transaction.atomic():
                        student_group, _ = Group.objects.get_or_create(name="Student")
                        student_user = User.objects.create_user(
                            username=login_id,
                            email=email or "",
                            password=password,
                            first_name=name,
                        )
                        student_user.groups.add(student_group)
                        Student.objects.create(
                            user=student_user,
                            academic_year=active_year,
                            student_id=admission.student_id or login_id,
                            name=name,
                            father_name=father_name,
                            mother_name=mother_name or "Unknown",
                            class_fk=admission.class_fk,
                            section=admission.section,
                            roll_no=student_id or login_id,
                            phone=contact,
                            gender=(gender or "Male").title(),
                            date_of_birth=dob or date.today(),
                            email=email or f"{login_id}@edupilot.local",
                        )
                success_count += 1

            except Exception as e:
                error_rows.append(f"Row {row_num}: {str(e)}")

        if success_count:
            messages.success(request, f"✅ {success_count} students successfully upload ho gaye.")
        if skip_count:
            messages.warning(request, f"⚠️ {skip_count} rows skip ki gayi (duplicate ya empty).")
        for err in error_rows:
            messages.error(request, err)

        return redirect('admission_list')

    return render(request, 'admin_panel/bulk_upload_students.html')


@permission_required('teacher_dashboard.add_teacher', raise_exception=True)
def bulk_upload_teachers(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        except Exception as e:
            messages.error(request, f"❌ Excel file padhne mein masla: {e}")
            return redirect('bulk_upload_teachers')

        ws = wb.active
        teacher_group, _ = Group.objects.get_or_create(name="Teacher")

        success_count = 0
        skip_count    = 0
        error_rows    = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            if not any(row):
                continue

            try:
                teacher_id    = _str(row[0])
                name          = _str(row[1])
                email         = _str(row[2])
                phone         = _str(row[3])
                gender        = _str(row[4]) or 'Male'
                dob           = _parse_date(row[5])
                qualification = _str(row[6]) or 'BS'
                experience    = _parse_experience(row[7])
                address       = _str(row[8])
                faculty_group = _str(row[9])
                department    = _str(row[10])
                subject_name  = _str(row[11])
                joining_date  = _parse_date(row[13])
                raw_status    = _str(row[14]).lower() or 'active'
                login_id      = _str(row[15])
                password      = _str(row[16])

                if not name or not email or not login_id or not password:
                    skip_count += 1
                    error_rows.append(
                        f"Row {row_num}: Name, Email, Login_Id, and Password are required."
                    )
                    continue

                if User.objects.filter(email__iexact=email).exists():
                    skip_count += 1
                    error_rows.append(
                        f"Row {row_num}: Email '{email}' pehle se exist karta hai — skip kiya."
                    )
                    continue
                if User.objects.filter(username__iexact=login_id).exists():
                    skip_count += 1
                    error_rows.append(
                        f"Row {row_num}: Login ID '{login_id}' pehle se exist karta hai — skip kiya."
                    )
                    continue
                try:
                    validate_password(password)
                except Exception as exc:
                    skip_count += 1
                    error_rows.append(
                        f"Row {row_num}: Password invalid — {' '.join(exc.messages) if hasattr(exc, 'messages') else str(exc)}"
                    )
                    continue

                status_map = {
                    'active':   'active',
                    'inactive': 'inactive',
                    'on leave': 'inactive',
                }
                teacher_status = status_map.get(raw_status, 'active')

                with transaction.atomic():
                    user = User.objects.create_user(
                        username   = login_id,
                        email      = email,
                        password   = password,
                        first_name = name,
                    )
                    user.groups.set([teacher_group])

                    teacher = Teacher.objects.create(
                        user          = user,
                        name          = name,
                        email         = email,
                        phone         = phone,
                        gender        = gender,
                        qualification = qualification,
                        experience    = experience,
                        Address       = address,
                        department    = department,
                        faculty_group = [faculty_group] if faculty_group else [],
                        status        = teacher_status,
                    )

                    if subject_name:
                        subject_obj = Subject.objects.filter(name__iexact=subject_name).first()
                        if subject_obj:
                            teacher.subjects.add(subject_obj)

                success_count += 1

            except Exception as e:
                error_rows.append(f"Row {row_num}: {str(e)}")

        if success_count:
            messages.success(request, f"✅ {success_count} teachers successfully upload ho gaye.")
        if skip_count:
            messages.warning(request, f"⚠️ {skip_count} rows skip ki gayi (duplicate ya empty).")
        for err in error_rows:
            messages.error(request, err)

        return redirect('teacher_list')

    return render(request, 'admin_panel/bulk_upload_teachers.html')


# ==================== BULK DELETE ====================
from django.views.decorators.http import require_POST


@permission_required('admin_panel.delete_admission', raise_exception=True)
@require_POST
def bulk_delete_students(request):
    delete_all = request.POST.get('delete_all') == 'on'
    if delete_all:
        deleted_count, _ = Admission.objects.all().delete()
        messages.success(request, f"✅ Saare {deleted_count} students delete ho gaye.")
    else:
        student_ids = request.POST.getlist('student_ids')
        if not student_ids:
            messages.warning(request, "⚠️ Koi student select nahi kiya gaya.")
            return redirect('admission_list')
        deleted_count, _ = Admission.objects.filter(id__in=student_ids).delete()
        messages.success(request, f"✅ {deleted_count} students successfully delete ho gaye.")
    return redirect('admission_list')


@permission_required('teacher_dashboard.delete_teacher', raise_exception=True)
@require_POST
def bulk_delete_teachers(request):
    delete_all = request.POST.get('delete_all') == 'on'
    if delete_all:
        user_ids = Teacher.objects.values_list('user_id', flat=True)
        deleted_count, _ = Teacher.objects.all().delete()
        User.objects.filter(id__in=user_ids).delete()
        messages.success(request, f"✅ Saare {deleted_count} teachers delete ho gaye.")
    else:
        teacher_ids = request.POST.getlist('teacher_ids')
        if not teacher_ids:
            messages.warning(request, "⚠️ Koi teacher select nahi kiya gaya.")
            return redirect('teacher_list')
        user_ids = Teacher.objects.filter(id__in=teacher_ids).values_list('user_id', flat=True)
        deleted_count, _ = Teacher.objects.filter(id__in=teacher_ids).delete()
        User.objects.filter(id__in=user_ids).delete()
        messages.success(request, f"✅ {deleted_count} teachers successfully delete ho gaye.")
    return redirect('teacher_list')





# --- LOGIN/LOGOUT ---
@csrf_protect
def login_view(request):
    if request.user.is_authenticated:
        # User pehle se authenticated hai toh role ke mutabiq redirect karein
        if request.user.role == 'admin': return redirect('admin_dashboard')
        elif request.user.role == 'student': return redirect('student-dashboard')
        elif request.user.role == 'teacher': return redirect('teacher-dashboard')
        elif request.user.role == 'parent': return redirect('parent-dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role') # UI dropdown se milne wala role
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Simple validation: user ka role check karein
            if user.role.lower() == role:
                login(request, user)
                if role == 'admin': return redirect('admin_dashboard')
                elif role == 'student': return redirect('student-dashboard')
                elif role == 'teacher': return redirect('teacher-dashboard')
                elif role == 'parent': return redirect('parent-dashboard')
                else: return redirect('admin_dashboard') # Default fallback
            else:
                messages.error(request, f"Invalid login for {role.upper()} portal.")
        else:
            messages.error(request, "Invalid username ya password.")
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

# --- ADMIN DASHBOARD ---
@login_required(login_url='login')
def admin_dashboard_view(request):
    fee_settings = FeeGenerationSettings.objects.first() or FeeGenerationSettings.objects.create()
    salary_settings = SalaryAutomationSettings.objects.first() or SalaryAutomationSettings.objects.create()

    if request.method == 'POST':
        action = request.POST.get('action')

        try:
            if action == 'save_fee_settings':
                fee_settings.auto_enabled = request.POST.get('fee_auto_enabled') == 'on'
                fee_settings.generation_day = int(request.POST.get('fee_generation_day') or 1)
                fee_settings.generation_time = datetime.strptime(
                    request.POST.get('fee_generation_time') or '09:00', '%H:%M'
                ).time()
                fee_settings.send_notifications = request.POST.get('fee_send_notifications') == 'on'
                fee_settings.save()
                messages.success(request, 'Fee automation settings saved successfully.')

            elif action == 'save_salary_settings':
                salary_settings.auto_enabled = request.POST.get('salary_auto_enabled') == 'on'
                salary_settings.generation_day = int(request.POST.get('salary_generation_day') or 30)
                salary_settings.generation_time = datetime.strptime(
                    request.POST.get('salary_generation_time') or '18:00', '%H:%M'
                ).time()
                salary_settings.send_notifications = request.POST.get('salary_send_notifications') == 'on'
                salary_settings.save()
                messages.success(request, 'Salary automation settings saved successfully.')

            elif action == 'generate_fees':
                today = date.today()
                month_name = request.POST.get('fee_month') or today.strftime('%B')
                year = int(request.POST.get('fee_year') or today.year)
                job = AutomationJob.objects.create(job_type='MANUAL_GENERATION', status='PENDING')
                count = FeeGenerationService.generate_monthly_fees(month_name, year, job_id=job.id)
                messages.success(request, f'Fee generation completed. {count} voucher(s) generated. Job ID: {job.id}.')

            elif action == 'generate_salaries':
                today = date.today()
                month_name = request.POST.get('salary_month') or today.strftime('%B')
                year = int(request.POST.get('salary_year') or today.year)
                SalaryAutomationService.generate_salaries(month_name, year)
                messages.success(request, 'Salary generation process completed.')

            elif action == 'send_notifications':
                pending_count = NotificationQueue.objects.filter(status='PENDING').count()
                NotificationDispatcherService.send_pending_notifications()
                messages.success(request, f'Notification dispatcher processed {pending_count} pending notification(s).')

            elif action == 'retry_failed_fees':
                today = date.today()
                month_name = request.POST.get('retry_fee_month') or today.strftime('%B')
                year = int(request.POST.get('retry_fee_year') or today.year)
                failed_jobs = AutomationJob.objects.filter(details__status='FAILED').distinct()
                retry_count = 0
                for job in failed_jobs:
                    FeeGenerationService.retry_failed_records(job.id, month_name, year)
                    retry_count += 1
                messages.success(request, f'Retry initiated for {retry_count} fee automation job(s).')

            elif action == 'retry_failed_salaries':
                SalaryAutomationService.generate_salaries()
                messages.success(request, 'Salary retry process completed.')

        except Exception as e:
            messages.error(request, f'Action failed: {str(e)}')

        return redirect('admin_dashboard')

    # Ab hum models.py mein banaye gaye helper function se stats utha rahe hain
    context = get_dashboard_stats()
    context.update({
        'fee_settings': fee_settings,
        'salary_settings': salary_settings,
        'fee_logs': FeeGenerationLog.objects.all().order_by('-started_at')[:5],
        'automation_jobs': AutomationJob.objects.all().order_by('-started_at')[:8],
        'automation_details': AutomationJobDetail.objects.filter(status='FAILED').order_by('-id')[:8],
        'salary_jobs': SalaryAutomationJob.objects.all().order_by('-started_at')[:8],
        'salary_details': SalaryAutomationJobDetail.objects.filter(status='FAILED').order_by('-id')[:8],
        'notifications': NotificationQueue.objects.all().order_by('-created_at')[:8],
        'pending_notifications': NotificationQueue.objects.filter(status='PENDING').count(),
        'failed_notifications': NotificationQueue.objects.filter(status='FAILED').count(),
        'fee_vouchers_count': FeeVoucher.objects.count(),
        'salary_vouchers_count': SalaryVoucher.objects.count(),
        'months': [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ],
        'current_month': date.today().strftime('%B'),
        'current_year': date.today().year,
    })
    return render(request, 'dashboard.html', context)

# API for Admin Dashboard (Dynamic Front-end ke liye)
class AdminDashboardAPI(APIView):
    def get(self, request):
        data = get_dashboard_stats()
        # API response format
        return Response({
            "stats": {
                "total_students": data['total_students'],
                "monthly_revenue": data['monthly_revenue'],
                "total_expenses": data['total_expenses'],
                "high_risk_students": data['high_risk_students'],
                "total_teachers": data['total_teachers'],
                "total_staff": data['total_staff']
            },
            "recent_transactions": list(data['recent_transactions'].values('title', 'amount', 'type', 'date'))
        })

# --- STUDENT & PARENT VIEWS ---
def student_registration_view(request):
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')
    else:
        form = StudentRegistrationForm()
    return render(request, 'register.html', {'form': form})

def generate_fees_view(request):
    # Current month aur year nikal kar service call karna
    today = date.today()
    month_name = today.strftime("%B")
    year = today.year
    
    count = FeeGenerationService.generate_monthly_fees(month_name, year)
    messages.success(request, f"{count} vouchers generated successfully for {month_name}-{year}!")
    return redirect('/admin/edupilot_core/feevoucher/')

@login_required
def student_dashboard(request):
    try:
        student = Student.objects.get(admission_number=request.user.username)
        vouchers = FeeVoucher.objects.filter(student=student).order_by('-id')
        balance = StudentBalance.objects.get(student=student)
        
        # ✅ GET NOTIFICATIONS
        notifications = NotificationQueue.objects.filter(student=student).order_by('-created_at')[:10]
        
        context = {
            'student': student,
            'vouchers': vouchers,
            'balance': balance,
            'notifications': notifications  # ← ADD YE
        }
    except:
        context = {
            'student': None,
            'vouchers': [],
            'balance': None,
            'notifications': [],
            'error': 'Student data not found'
        }
    return render(request, 'student_profile/dashboard.html', context)

@login_required
def teacher_dashboard(request):
    try:
        # Teacher model mein 'user' field nahi hai
        # Username ko teacher_id se match karo
        teacher = Teacher.objects.get(teacher_id=request.user.username)
        
        # Get salary vouchers
        salary_vouchers = SalaryVoucher.objects.filter(teacher=teacher).order_by('-id')
        
        # Get notifications (empty - teacher notifications not in system)
        notifications = []
        
        context = {
            'teacher': teacher,
            'salary_vouchers': salary_vouchers,
            'notifications': notifications
        }
    except Teacher.DoesNotExist:
        context = {
            'teacher': None,
            'salary_vouchers': [],
            'notifications': [],
            'error': 'Teacher record not found'
        }
    except Exception as e:
        context = {
            'teacher': None,
            'salary_vouchers': [],
            'notifications': [],
            'error': f'Error: {str(e)}'
        }
    
    return render(request, 'teacher_dashboard/teacher_dashboard.html', context)

@login_required
def parent_dashboard(request):
    context = {'children': [], 'vouchers': []}
    return render(request, 'parent/dashboard.html', context)

def automation_logs(request):
    logs = AutomationJobDetail.objects.all().order_by('-id')
    return render(request, 'automation/logs.html', {'logs': logs})


# ==================== OPERATIONS ====================
# Admin-only procurement, inventory and fleet screens. TransportRoute remains
# the same model used by fee assignments, so its existing behavior is preserved.
from django.db.models.deletion import ProtectedError
from .models import (
    ProcurementCategory, Vendor, PurchaseRequest, InventoryItem, StockMovement,
    Vehicle, RouteVehicleAssignment, TransportTrip, VehicleMaintenance, TransportRoute,
)
from .forms import (
    ProcurementCategoryForm, VendorForm, PurchaseRequestForm, InventoryItemForm,
    StockMovementForm, TransportRouteOperationForm, VehicleForm,
    RouteVehicleAssignmentForm, TransportTripForm, VehicleMaintenanceForm,
)


def _operation_configs():
    return {
        "procurement_category": {
            "model": ProcurementCategory, "form": ProcurementCategoryForm,
            "title": "Categories", "singular": "Procurement Category",
            "subtitle": "Organize purchasing and inventory records.", "icon": "fa-tags",
            "list_url": "operation_procurement_category_list",
            "create_url": "operation_procurement_category_create",
            "edit_url": "operation_procurement_category_edit",
            "delete_url": "operation_procurement_category_delete",
            "headers": ["Name", "Description", "Status", "Created"],
            "cells": lambda obj: [obj.name, obj.description or "-", "Active" if obj.is_active else "Inactive", obj.created_at.strftime("%d %b %Y")],
            "search": ["name", "description"],
        },
        "vendor": {
            "model": Vendor, "form": VendorForm, "title": "Vendors", "singular": "Vendor",
            "subtitle": "Manage supplier contacts and availability.", "icon": "fa-handshake",
            "list_url": "operation_vendor_list", "create_url": "operation_vendor_create",
            "edit_url": "operation_vendor_edit", "delete_url": "operation_vendor_delete",
            "headers": ["Vendor", "Contact", "Phone", "Email", "Status"],
            "cells": lambda obj: [obj.name, obj.contact_person or "-", obj.phone or "-", obj.email or "-", obj.get_status_display()],
            "search": ["name", "contact_person", "phone", "email"],
        },
        "purchase_request": {
            "model": PurchaseRequest, "form": PurchaseRequestForm,
            "title": "Purchase Requests", "singular": "Purchase Request",
            "subtitle": "Track requested purchases from approval to receipt.", "icon": "fa-clipboard-list",
            "list_url": "operation_purchase_request_list", "create_url": "operation_purchase_request_create",
            "edit_url": "operation_purchase_request_edit", "delete_url": "operation_purchase_request_delete",
            "headers": ["Request", "Category", "Priority", "Estimated Cost", "Needed By", "Status"],
            "cells": lambda obj: [obj.title, str(obj.category or "-"), obj.get_priority_display(), f"PKR {obj.estimated_cost:,.2f}", obj.needed_by.strftime("%d %b %Y") if obj.needed_by else "-", obj.get_status_display()],
            "search": ["title", "description", "category__name", "vendor__name"],
            "select_related": ["category", "vendor", "requested_by"],
        },
        "inventory_item": {
            "model": InventoryItem, "form": InventoryItemForm,
            "title": "Inventory Items", "singular": "Inventory Item",
            "subtitle": "Monitor quantities, reorder levels and unit costs.", "icon": "fa-boxes",
            "list_url": "operation_inventory_item_list", "create_url": "operation_inventory_item_create",
            "edit_url": "operation_inventory_item_edit", "delete_url": "operation_inventory_item_delete",
            "headers": ["Item", "SKU", "Category", "Quantity", "Reorder Level", "Unit Cost", "Status"],
            "cells": lambda obj: [obj.name, obj.sku or "-", str(obj.category or "-"), f"{obj.quantity} {obj.unit}", obj.reorder_level, f"PKR {obj.unit_cost:,.2f}", obj.get_status_display()],
            "search": ["name", "sku", "category__name", "vendor__name"],
            "select_related": ["category", "vendor"],
        },
        "stock_movement": {
            "model": StockMovement, "form": StockMovementForm,
            "title": "Stock Movements", "singular": "Stock Movement",
            "subtitle": "Review stock additions, removals and adjustments.", "icon": "fa-exchange-alt",
            "list_url": "operation_stock_movement_list", "create_url": "operation_stock_movement_create",
            "edit_url": "operation_stock_movement_edit", "delete_url": "operation_stock_movement_delete",
            "headers": ["Item", "Movement", "Quantity", "Note", "Created By", "Date"],
            "cells": lambda obj: [obj.item.name, obj.get_movement_type_display(), obj.quantity, obj.note or "-", obj.created_by.get_username() if obj.created_by else "-", obj.created_at.strftime("%d %b %Y, %I:%M %p")],
            "search": ["item__name", "item__sku", "note"],
            "select_related": ["item", "created_by"],
        },
        "vehicle": {
            "model": Vehicle, "form": VehicleForm, "title": "Vehicles", "singular": "Vehicle",
            "subtitle": "Manage the school transport fleet and drivers.", "icon": "fa-shuttle-van",
            "list_url": "operation_vehicle_list", "create_url": "operation_vehicle_create",
            "edit_url": "operation_vehicle_edit", "delete_url": "operation_vehicle_delete",
            "headers": ["Vehicle No.", "Type", "Capacity", "Driver", "Phone", "Registration Expiry", "Status"],
            "cells": lambda obj: [obj.vehicle_no, obj.get_vehicle_type_display(), obj.capacity, obj.driver_name or "-", obj.driver_phone or "-", obj.registration_expiry.strftime("%d %b %Y") if obj.registration_expiry else "-", obj.get_status_display()],
            "search": ["vehicle_no", "driver_name", "driver_phone"],
        },
        "transport_route": {
            "model": TransportRoute, "form": TransportRouteOperationForm,
            "title": "Transport Routes", "singular": "Transport Route",
            "subtitle": "Manage routes while preserving their fee assignments.", "icon": "fa-route",
            "list_url": "operation_transport_route_list", "create_url": "operation_transport_route_create",
            "edit_url": "operation_transport_route_edit", "delete_url": "operation_transport_route_delete",
            "headers": ["Route", "Monthly Fee", "Assigned Vehicles"],
            "cells": lambda obj: [obj.route_name, f"PKR {obj.amount:,.2f}", obj.vehicle_assignments.count()],
            "search": ["route_name"],
        },
        "route_assignment": {
            "model": RouteVehicleAssignment, "form": RouteVehicleAssignmentForm,
            "title": "Route Assignments", "singular": "Route Assignment",
            "subtitle": "Assign vehicles and drivers to transport routes.", "icon": "fa-map-marked-alt",
            "list_url": "operation_route_assignment_list", "create_url": "operation_route_assignment_create",
            "edit_url": "operation_route_assignment_edit", "delete_url": "operation_route_assignment_delete",
            "headers": ["Route", "Vehicle", "Driver", "Start Date", "End Date", "Status"],
            "cells": lambda obj: [obj.route.route_name, obj.vehicle.vehicle_no, obj.driver_name or obj.vehicle.driver_name or "-", obj.start_date.strftime("%d %b %Y"), obj.end_date.strftime("%d %b %Y") if obj.end_date else "-", "Active" if obj.is_active else "Inactive"],
            "search": ["route__route_name", "vehicle__vehicle_no", "driver_name"],
            "select_related": ["route", "vehicle"],
        },
        "transport_trip": {
            "model": TransportTrip, "form": TransportTripForm,
            "title": "Transport Trips", "singular": "Transport Trip",
            "subtitle": "Record route service, departure performance and transported students.", "icon": "fa-road",
            "list_url": "operation_transport_trip_list", "create_url": "operation_transport_trip_create",
            "edit_url": "operation_transport_trip_edit", "delete_url": "operation_transport_trip_delete",
            "headers": ["Service Date", "Route", "Vehicle", "Scheduled", "Actual", "Students", "Status"],
            "cells": lambda obj: [obj.service_date.strftime("%d %b %Y"), obj.route.route_name, obj.vehicle.vehicle_no, obj.scheduled_departure.strftime("%I:%M %p"), obj.actual_departure.strftime("%I:%M %p") if obj.actual_departure else "-", obj.students_transported, obj.get_status_display()],
            "search": ["route__route_name", "vehicle__vehicle_no", "notes", "status"],
            "select_related": ["route", "vehicle"],
        },
        "vehicle_maintenance": {
            "model": VehicleMaintenance, "form": VehicleMaintenanceForm,
            "title": "Maintenance Records", "singular": "Maintenance Record",
            "subtitle": "Track scheduled and completed vehicle servicing.", "icon": "fa-tools",
            "list_url": "operation_vehicle_maintenance_list", "create_url": "operation_vehicle_maintenance_create",
            "edit_url": "operation_vehicle_maintenance_edit", "delete_url": "operation_vehicle_maintenance_delete",
            "headers": ["Vehicle", "Service", "Service Date", "Cost", "Status", "Notes"],
            "cells": lambda obj: [obj.vehicle.vehicle_no, obj.maintenance_type, obj.service_date.strftime("%d %b %Y"), f"PKR {obj.cost:,.2f}", obj.get_status_display(), obj.notes or "-"],
            "search": ["vehicle__vehicle_no", "maintenance_type", "notes"],
            "select_related": ["vehicle"],
        },
    }


def _operation_list(request, key):
    config = _operation_configs()[key]
    queryset = config["model"].objects.all()
    if config.get("select_related"):
        queryset = queryset.select_related(*config["select_related"])
    query = request.GET.get("q", "").strip()
    if query:
        condition = Q()
        for field in config["search"]:
            condition |= Q(**{f"{field}__icontains": query})
        queryset = queryset.filter(condition)
    rows = [{"object": obj, "cells": config["cells"](obj)} for obj in queryset]
    return render(request, "admin_panel/operations_list.html", {**config, "rows": rows, "query": query})


def _operation_save(request, key, pk=None):
    config = _operation_configs()[key]
    instance = get_object_or_404(config["model"], pk=pk) if pk else None
    form = config["form"](request.POST or None, instance=instance)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        if key == "purchase_request" and not obj.requested_by_id:
            obj.requested_by = request.user
        if key == "stock_movement" and not obj.created_by_id:
            obj.created_by = request.user
        obj.save()
        messages.success(request, f"{config['singular']} saved successfully.")
        return redirect(config["list_url"])
    return render(request, "admin_panel/operations_form.html", {
        **config, "form": form, "is_edit": bool(instance),
    })


def _operation_delete(request, key, pk):
    config = _operation_configs()[key]
    obj = get_object_or_404(config["model"], pk=pk)
    if request.method == "POST":
        try:
            obj.delete()
            messages.success(request, f"{config['singular']} deleted successfully.")
        except ProtectedError:
            messages.error(request, f"{config['singular']} is in use and cannot be deleted.")
        return redirect(config["list_url"])
    return render(request, "admin_panel/operations_confirm_delete.html", {**config, "object": obj})


@login_required
@user_passes_test(is_admin)
def operation_procurement_dashboard(request):
    items = list(InventoryItem.objects.all())
    pending = PurchaseRequest.objects.filter(status="pending").count()
    approved_value = PurchaseRequest.objects.filter(status__in=["approved", "ordered", "received"]).aggregate(total=Sum("estimated_cost"))["total"] or 0
    context = {
        "module": "procurement", "title": "Procurement & Inventory",
        "stats": [
            ("Purchase Requests", PurchaseRequest.objects.count(), "fa-clipboard-list", "blue"),
            ("Pending Approvals", pending, "fa-hourglass-half", "amber"),
            ("Inventory Items", len(items), "fa-boxes", "green"),
            ("Low Stock Items", sum(1 for item in items if item.quantity <= item.reorder_level), "fa-exclamation-triangle", "red"),
        ],
        "approved_value": approved_value,
        "recent_requests": PurchaseRequest.objects.select_related("category").all()[:6],
    }
    return render(request, "admin_panel/operations_dashboard.html", context)


@login_required
@user_passes_test(is_admin)
def operation_transportation_dashboard(request):
    context = {
        "module": "fleet", "title": "Transportation & Fleet",
        "stats": [
            ("Total Vehicles", Vehicle.objects.count(), "fa-bus", "blue"),
            ("Active Vehicles", Vehicle.objects.filter(status="active").count(), "fa-check-circle", "green"),
            ("Transport Routes", TransportRoute.objects.count(), "fa-route", "amber"),
            ("Under Maintenance", Vehicle.objects.filter(status="maintenance").count(), "fa-tools", "red"),
        ],
        "active_assignments": RouteVehicleAssignment.objects.filter(is_active=True).select_related("route", "vehicle")[:6],
        "upcoming_maintenance": VehicleMaintenance.objects.exclude(status="completed").select_related("vehicle")[:6],
    }
    return render(request, "admin_panel/operations_dashboard.html", context)


def _operation_view_set(key):
    @login_required
    @user_passes_test(is_admin)
    def list_view(request):
        return _operation_list(request, key)

    @login_required
    @user_passes_test(is_admin)
    def create_view(request):
        return _operation_save(request, key)

    @login_required
    @user_passes_test(is_admin)
    def edit_view(request, pk):
        return _operation_save(request, key, pk)

    @login_required
    @user_passes_test(is_admin)
    def delete_view(request, pk):
        return _operation_delete(request, key, pk)

    return list_view, create_view, edit_view, delete_view


(operation_procurement_category_list, operation_procurement_category_create,
 operation_procurement_category_edit, operation_procurement_category_delete) = _operation_view_set("procurement_category")
(operation_vendor_list, operation_vendor_create,
 operation_vendor_edit, operation_vendor_delete) = _operation_view_set("vendor")
(operation_purchase_request_list, operation_purchase_request_create,
 operation_purchase_request_edit, operation_purchase_request_delete) = _operation_view_set("purchase_request")
(operation_inventory_item_list, operation_inventory_item_create,
 operation_inventory_item_edit, operation_inventory_item_delete) = _operation_view_set("inventory_item")
(operation_stock_movement_list, operation_stock_movement_create,
 operation_stock_movement_edit, operation_stock_movement_delete) = _operation_view_set("stock_movement")
(operation_vehicle_list, operation_vehicle_create,
 operation_vehicle_edit, operation_vehicle_delete) = _operation_view_set("vehicle")
(operation_transport_route_list, operation_transport_route_create,
 operation_transport_route_edit, operation_transport_route_delete) = _operation_view_set("transport_route")
(operation_route_assignment_list, operation_route_assignment_create,
 operation_route_assignment_edit, operation_route_assignment_delete) = _operation_view_set("route_assignment")
(operation_transport_trip_list, operation_transport_trip_create,
 operation_transport_trip_edit, operation_transport_trip_delete) = _operation_view_set("transport_trip")
(operation_vehicle_maintenance_list, operation_vehicle_maintenance_create,
 operation_vehicle_maintenance_edit, operation_vehicle_maintenance_delete) = _operation_view_set("vehicle_maintenance")
